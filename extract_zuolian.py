# -*- coding: utf-8 -*-
"""
从《左联回忆录》PDF中提取人际关系数据并写入Excel
自动合并重复记录
"""

import pdfplumber
import openpyxl
import re
from collections import defaultdict

# 完整人物对照表 - 包含所有别名映射
PERSON_MAP = {
    # ZLH-001 鲁迅
    "鲁迅": "ZLH-001", "周树人": "ZLH-001", "L.X.": "ZLH-001", "豫才": "ZLH-001",
    # ZLH-002 茅盾
    "茅盾": "ZLH-002", "沈雁冰": "ZLH-002", "玄珠": "ZLH-002", "方璧": "ZLH-002", "微明": "ZLH-002",
    # ZLH-003 瞿秋白
    "瞿秋白": "ZLH-003", "维它": "ZLH-003", "史铁儿": "ZLH-003", "宋阳": "ZLH-003", "何凝": "ZLH-003",
    # ZLH-004 夏衍
    "夏衍": "ZLH-004", "沈端先": "ZLH-004", "黄子布": "ZLH-004", "蔡叔亮": "ZLH-004", "丁一": "ZLH-004",
    # ZLH-005 冯雪峰
    "冯雪峰": "ZLH-005", "画室": "ZLH-005", "洛扬": "ZLH-005", "成休": "ZLH-005", "雪峰": "ZLH-005",
    # ZLH-006 潘汉年
    "潘汉年": "ZLH-006", "萧叔安": "ZLH-006", "胡越": "ZLH-006", "小潘": "ZLH-006", "胡阿毛": "ZLH-006",
    # ZLH-007 周扬
    "周扬": "ZLH-007", "周起应": "ZLH-007", "宜振": "ZLH-007", "公甫": "ZLH-007",
    # ZLH-008 田汉
    "田汉": "ZLH-008", "寿昌": "ZLH-008", "陈瑜": "ZLH-008", "伯鸿": "ZLH-008", "汉仙": "ZLH-008",
    # ZLH-009 郑伯奇
    "郑伯奇": "ZLH-009", "伯奇": "ZLH-009", "君嘉": "ZLH-009", "席珍": "ZLH-009",
    # ZLH-010 洪深
    "洪深": "ZLH-010", "洪伯飞": "ZLH-010", "浅哉": "ZLH-010", "伯飞": "ZLH-010",
    # ZLH-011 钱杏邨
    "钱杏邨": "ZLH-011", "阿英": "ZLH-011", "魏如晦": "ZLH-011", "钱谦": "ZLH-011",
    # ZLH-012 蒋光慈
    "蒋光慈": "ZLH-012", "蒋光赤": "ZLH-012", "光赤": "ZLH-012", "光慈": "ZLH-012",
    # ZLH-013 阳翰笙
    "阳翰笙": "ZLH-013", "欧阳继修": "ZLH-013", "华汉": "ZLH-013", "翰笙": "ZLH-013",
    # ZLH-014 冯乃超
    "冯乃超": "ZLH-014", "冯其英": "ZLH-014", "乃超": "ZLH-014",
    # ZLH-015 郁达夫
    "郁达夫": "ZLH-015", "郁文": "ZLH-015", "达夫": "ZLH-015",
    # ZLH-016 柔石
    "柔石": "ZLH-016", "赵平复": "ZLH-016", "平复": "ZLH-016",
    # ZLH-017 胡也频
    "胡也频": "ZLH-017", "胡崇轩": "ZLH-017", "也频": "ZLH-017",
    # ZLH-018 冯铿
    "冯铿": "ZLH-018", "岭梅": "ZLH-018", "梅子": "ZLH-018",
    # ZLH-019 殷夫
    "殷夫": "ZLH-019", "徐白": "ZLH-019", "白莽": "ZLH-019",
    # ZLH-020 李求实
    "李求实": "ZLH-020", "李伟森": "ZLH-020", "仲才": "ZLH-020", "求实": "ZLH-020",
    # ZLH-021 丁玲
    "丁玲": "ZLH-021", "蒋冰之": "ZLH-021", "彬芷": "ZLH-021",
    # ZLH-022 张天翼
    "张天翼": "ZLH-022", "张元定": "ZLH-022", "铁池": "ZLH-022", "天翼": "ZLH-022",
    # ZLH-023 叶紫
    "叶紫": "ZLH-023", "余鹤林": "ZLH-023",
    # ZLH-024 艾芜
    "艾芜": "ZLH-024", "汤道耕": "ZLH-024", "道耕": "ZLH-024",
    # ZLH-025 沙汀
    "沙汀": "ZLH-025", "杨朝熙": "ZLH-025", "朝熙": "ZLH-025",
    # ZLH-026 胡风
    "胡风": "ZLH-026", "张光人": "ZLH-026", "谷非": "ZLH-026",
    # ZLH-027 萧军
    "萧军": "ZLH-027", "刘鸿霖": "ZLH-027", "三郎": "ZLH-027",
    # ZLH-028 萧红
    "萧红": "ZLH-028", "张迺莹": "ZLH-028", "悄吟": "ZLH-028",
    # ZLH-029 聂绀弩
    "聂绀弩": "ZLH-029", "绀弩": "ZLH-029", "耳耶": "ZLH-029",
    # ZLH-030 吴奚如
    "吴奚如": "ZLH-030", "吴席儒": "ZLH-030", "奚如": "ZLH-030",
    # ZLH-031 葛琴
    "葛琴": "ZLH-031", "葛爱琴": "ZLH-031", "爱琴": "ZLH-031",
    # ZLH-032 徐懋庸
    "徐懋庸": "ZLH-032", "徐茂荣": "ZLH-032", "懋庸": "ZLH-032", "茂荣": "ZLH-032",
    # ZLH-033 周文
    "周文": "ZLH-033", "第一修": "ZLH-033",
    # ZLH-034 彭康
    "彭康": "ZLH-034", "彭坚": "ZLH-034",
    # ZLH-035 李辉英
    "李辉英": "ZLH-035", "李公白": "ZLH-035", "辉英": "ZLH-035", "公白": "ZLH-035",
    # ZLH-036 穆木天
    "穆木天": "ZLH-036", "穆敬熙": "ZLH-036", "木天": "ZLH-036", "敬熙": "ZLH-036",
    # ZLH-037 楼适夷
    "楼适夷": "ZLH-037", "楼锡春": "ZLH-037", "适夷": "ZLH-037", "锡春": "ZLH-037",
    # ZLH-038 陈荒煤
    "陈荒煤": "ZLH-038", "陈光亮": "ZLH-038", "荒煤": "ZLH-038", "光亮": "ZLH-038",
    # ZLH-039 欧阳山
    "欧阳山": "ZLH-039", "杨凤岐": "ZLH-039", "凤岐": "ZLH-039",
    # ZLH-040 草明
    "草明": "ZLH-040", "郑晓光": "ZLH-040", "晓光": "ZLH-040",
    # ZLH-041 骆宾基
    "骆宾基": "ZLH-041", "张璞君": "ZLH-041", "璞君": "ZLH-041",
    # ZLH-042 魏金枝
    "魏金枝": "ZLH-042", "魏金志": "ZLH-042", "金志": "ZLH-042",
    # ZLH-043 端木蕻良
    "端木蕻良": "ZLH-043", "曹京平": "ZLH-043", "曹汉文": "ZLH-043", "端木": "ZLH-043", "蕻良": "ZLH-043",
    # ZLH-044 蒋锡金
    "蒋锡金": "ZLH-044", "蒋锡珍": "ZLH-044", "锡金": "ZLH-044", "锡珍": "ZLH-044",
    # ZLH-045 周立波
    "周立波": "ZLH-045", "周绍仪": "ZLH-045", "立波": "ZLH-045",
    # ZLH-046 邵荃麟
    "邵荃麟": "ZLH-046", "邵全麟": "ZLH-046", "荃麟": "ZLH-046", "全麟": "ZLH-046",
    # ZLH-047 林淡秋
    "林淡秋": "ZLH-047", "林秉钧": "ZLH-047", "淡秋": "ZLH-047", "秉钧": "ZLH-047",
    # ZLH-048 王任叔
    "王任叔": "ZLH-048", "巴人": "ZLH-048", "任叔": "ZLH-048",
    # ZLH-049 舒群
    "舒群": "ZLH-049", "许春元": "ZLH-049", "春元": "ZLH-049",
    # ZLH-050 于黑丁
    "于黑丁": "ZLH-050", "于黑庭": "ZLH-050", "黑庭": "ZLH-050",
    # ZLH-051 邱东平
    "邱东平": "ZLH-051", "邱东萍": "ZLH-051", "东萍": "ZLH-051",
    # ZLH-052 梅志
    "梅志": "ZLH-052", "屠稚眉": "ZLH-052", "稚眉": "ZLH-052",
    # ZLH-053 黄源
    "黄源": "ZLH-053", "黄河清": "ZLH-053", "河清": "ZLH-053",
    # ZLH-054 许幸之
    "许幸之": "ZLH-054", "许兴之": "ZLH-054", "幸之": "ZLH-054", "兴之": "ZLH-054",
    # ZLH-055 郑君里
    "郑君里": "ZLH-055", "郑君礼": "ZLH-055", "君里": "ZLH-055", "君礼": "ZLH-055",
    # ZLH-056 白杨
    "白杨": "ZLH-056", "杨成芳": "ZLH-056", "成芳": "ZLH-056",
    # ZLH-057 赵丹
    "赵丹": "ZLH-057", "赵凤翱": "ZLH-057", "凤翱": "ZLH-057",
    # ZLH-058 郭沫若
    "郭沫若": "ZLH-058", "郭开贞": "ZLH-058", "独武": "ZLH-058", "沫若": "ZLH-058", "开贞": "ZLH-058",
    # ZLH-059 白薇
    "白薇": "ZLH-059", "黄彰": "ZLH-059",
    # ZLH-060 鲁彦
    "鲁彦": "ZLH-060", "王鲁彦": "ZLH-060", "王衡": "ZLH-060",
    # ZLH-061 朱镜我
    "朱镜我": "ZLH-061", "朱德源": "ZLH-061", "金修": "ZLH-061",
    # ZLH-062 李一芒
    "李一芒": "ZLH-062", "李太白": "ZLH-062",
    # ZLH-063 任白戈
    "任白戈": "ZLH-063", "任全": "ZLH-063", "任以鲁": "ZLH-063",
    # ZLH-064 杜埃
    "杜埃": "ZLH-064", "杜秉正": "ZLH-064",
    # ZLH-065 柯仲平
    "柯仲平": "ZLH-065", "柯维翰": "ZLH-065",
    # ZLH-066 黄其奎
    "黄其奎": "ZLH-066", "黄日葵": "ZLH-066",
    # ZLH-067 任钧
    "任钧": "ZLH-067", "任君": "ZLH-067", "任均": "ZLH-067",
    # ZLH-068 温梓川
    "温梓川": "ZLH-068", "温志": "ZLH-068", "梓川": "ZLH-068",
    # ZLH-069 张庚
    "张庚": "ZLH-069", "张三庚": "ZLH-069",
    # ZLH-070 塞克
    "塞克": "ZLH-070", "陈凝秋": "ZLH-070",
    # ZLH-071 李霁野
    "李霁野": "ZLH-071", "霁野": "ZLH-071",
    # ZLH-072 丛深
    "丛深": "ZLH-072",
    # ZLH-073 王余杞
    "王余杞": "ZLH-073", "王余": "ZLH-073",
    # ZLH-074 严文井
    "严文井": "ZLH-074", "严文锦": "ZLH-074",
    # ZLH-075 黄肃秋
    "黄肃秋": "ZLH-075", "黄秋秋": "ZLH-075",
    # ZLH-076 马宁
    "马宁": "ZLH-076", "马云": "ZLH-076",
    # ZLH-077 柳倩
    "柳倩": "ZLH-077", "柳自强": "ZLH-077",
    # ZLH-078 王一知
    "王一知": "ZLH-078",
    # ZLH-079 谢和赓
    "谢和赓": "ZLH-079", "谢和": "ZLH-079",
    # ZLH-080 陈企霞
    "陈企霞": "ZLH-080", "陈企": "ZLH-080",
    # ZLH-081 陈白尘
    "陈白尘": "ZLH-081", "陈炳泉": "ZLH-081",
    # ZLH-082 宋之的
    "宋之的": "ZLH-082",
    # ZLH-083 林焕平
    "林焕平": "ZLH-083", "林焕": "ZLH-083",
    # ZLH-084 陆万美
    "陆万美": "ZLH-084",
    # ZLH-085 方殷
    "方殷": "ZLH-085",
    # ZLH-086 葛一虹
    "葛一虹": "ZLH-086",
    # ZLH-087 于伶
    "于伶": "ZLH-087", "任郁伶": "ZLH-087",
    # ZLH-088 徐芳
    "徐芳": "ZLH-088",
    # ZLH-089 唐弢
    "唐弢": "ZLH-089",
    # ZLH-090 吴晓邦
    "吴晓邦": "ZLH-090",
    # ZLH-091 叶以群
    "叶以群": "ZLH-091",
    # ZLH-092 蒋天佐
    "蒋天佐": "ZLH-092",
    # ZLH-093 王亚平
    "王亚平": "ZLH-093",
    # ZLH-094 吴亮平
    "吴亮平": "ZLH-094", "吴亮": "ZLH-094",
    # ZLH-095 何稼夫
    "何稼夫": "ZLH-095", "何稼": "ZLH-095",
    # ZLH-096 王统照
    "王统照": "ZLH-096", "剑三": "ZLH-096",
    # ZLH-097 王西彦
    "王西彦": "ZLH-097", "王西": "ZLH-097",
    # ZLH-098 沈从文
    "沈从文": "ZLH-098", "上官碧": "ZLH-098",
    # ZLH-099 吴奔星
    "吴奔星": "ZLH-099", "吴奔": "ZLH-099",
    # ZLH-100 姚蓬子
    "姚蓬子": "ZLH-100", "蓬子": "ZLH-100",
    # ZLH-101 徐迟
    "徐迟": "ZLH-101",
    # ZLH-102 张映纶
    "张映纶": "ZLH-102", "张映": "ZLH-102",
    # ZLH-103 王莹
    "王莹": "ZLH-103", "喻志华": "ZLH-103",
    # ZLH-104 叶灵凤
    "叶灵凤": "ZLH-104", "灵凤": "ZLH-104",
    # ZLH-105 沈起予
    "沈起予": "ZLH-105", "起予": "ZLH-105",
    # ZLH-106 张光年
    "张光年": "ZLH-106", "光未然": "ZLH-106",
    # ZLH-107 孙师毅
    "孙师毅": "ZLH-107", "师毅": "ZLH-107",
    # ZLH-108 陈鲤庭
    "陈鲤庭": "ZLH-108", "鲤庭": "ZLH-108",
    # ZLH-109 王林
    "王林": "ZLH-109",
    # ZLH-110 章泯
    "章泯": "ZLH-110",
    # ZLH-111 朱静子
    "朱静子": "ZLH-111", "静子": "ZLH-111",
    # ZLH-112 戴望舒
    "戴望舒": "ZLH-112", "望舒": "ZLH-112",
    # ZLH-113 施蛰存
    "施蛰存": "ZLH-113", "施蛰": "ZLH-113",
    # ZLH-114 刘呐鸥
    "刘呐鸥": "ZLH-114", "呐鸥": "ZLH-114",
    # ZLH-115 穆时英
    "穆时英": "ZLH-115", "时英": "ZLH-115",
    # ZLH-116 孙席珍
    "孙席珍": "ZLH-116",
    # ZLH-117 陈波儿
    "陈波儿": "ZLH-117", "波儿": "ZLH-117",
    # ZLH-118 吴组缃
    "吴组缃": "ZLH-118", "组缃": "ZLH-118",
    # ZLH-119 何家槐
    "何家槐": "ZLH-119", "家槐": "ZLH-119",
    # ZLH-120 夏征农
    "夏征农": "ZLH-120", "征农": "ZLH-120",
    # ZLH-121 内山完造
    "内山完造": "ZLH-121", "邬其山": "ZLH-121", "内山": "ZLH-121",
    # ZLH-122 许广平
    "许广平": "ZLH-122", "景宋": "ZLH-122", "广平": "ZLH-122",
    # ZLH-123 邹韬奋
    "邹韬奋": "ZLH-123", "邹恩润": "ZLH-123", "韬奋": "ZLH-123",
    # ZLH-124 李小峰
    "李小峰": "ZLH-124", "李致广": "ZLH-124", "小峰": "ZLH-124",
    # ZLH-125 陆志庠
    "陆志庠": "ZLH-125", "陆志祥": "ZLH-125",
    # ZLH-126 沈兼士
    "沈兼士": "ZLH-126", "兼士": "ZLH-126",
    # ZLH-127 曹靖华
    "曹靖华": "ZLH-127", "曹联亚": "ZLH-127", "靖华": "ZLH-127",
    # ZLH-128 林语堂
    "林语堂": "ZLH-128", "语堂": "ZLH-128",
    # ZLH-129 胡愈之
    "胡愈之": "ZLH-129", "愈之": "ZLH-129",
    # ZLH-130 杨杏佛
    "杨杏佛": "ZLH-130", "杏佛": "ZLH-130",
    # ZLH-131 艾格尼丝·史沫特莱
    "艾格尼丝·史沫特莱": "ZLH-131", "史沫特莱": "ZLH-131", "Smedley": "ZLH-131",
    # ZLH-132 哈罗德·伊罗生
    "哈罗德·伊罗生": "ZLH-132", "伊罗生": "ZLH-132", "Isaacs": "ZLH-132",
    # ZLH-133 埃德加·斯诺
    "埃德加·斯诺": "ZLH-133", "斯诺": "ZLH-133", "Snow": "ZLH-133",
    # ZLH-134 鹿地亘
    "鹿地亘": "ZLH-134", "Wataru": "ZLH-134",
    # ZLH-135 池田幸子
    "池田幸子": "ZLH-135", "幸子": "ZLH-135",
    # ZLH-136 宋庆龄
    "宋庆龄": "ZLH-136", "苏菲": "ZLH-136",
    # ZLH-137 蔡元培
    "蔡元培": "ZLH-137", "孑民": "ZLH-137",
    # ZLH-138 何香凝
    "何香凝": "ZLH-138", "香凝": "ZLH-138",
    # ZLH-139 陈望道
    "陈望道": "ZLH-139", "望道": "ZLH-139",
    # ZLH-140 鲁维查
    "鲁维查": "ZLH-140", "Rouwicha": "ZLH-140",
    # ZLH-141 高尔基
    "高尔基": "ZLH-141", "Gorky": "ZLH-141",
    # ZLH-142 法捷耶夫
    "法捷耶夫": "ZLH-142", "Fadeyev": "ZLH-142",
    # ZLH-143 罗曼·罗兰
    "罗曼·罗兰": "ZLH-143", "罗曼罗兰": "ZLH-143", "Rolland": "ZLH-143",
    # ZLH-144 巴比塞
    "巴比塞": "ZLH-144", "Barbusse": "ZLH-144",
    # ZLH-145 萧伯纳
    "萧伯纳": "ZLH-145", "Shaw": "ZLH-145",
    # ZLH-146 马尔罗
    "马尔罗": "ZLH-146", "Malraux": "ZLH-146",
    # ZLH-147 辛克莱
    "辛克莱": "ZLH-147", "Sinclair": "ZLH-147",
    # ZLH-148 片山潜
    "片山潜": "ZLH-148", "Katayama": "ZLH-148",
    # ZLH-149 路易·艾黎
    "路易·艾黎": "ZLH-149", "艾黎": "ZLH-149", "Alley": "ZLH-149",
    # ZLH-150 萧三
    "萧三": "ZLH-150", "萧植蕃": "ZLH-150", "艾三": "ZLH-150",
}

# ID到主要名字的映射
ID_TO_NAME = {
    "ZLH-001": "鲁迅", "ZLH-002": "茅盾", "ZLH-003": "瞿秋白", "ZLH-004": "夏衍",
    "ZLH-005": "冯雪峰", "ZLH-006": "潘汉年", "ZLH-007": "周扬", "ZLH-008": "田汉",
    "ZLH-009": "郑伯奇", "ZLH-010": "洪深", "ZLH-011": "钱杏邨", "ZLH-012": "蒋光慈",
    "ZLH-013": "阳翰笙", "ZLH-014": "冯乃超", "ZLH-015": "郁达夫", "ZLH-016": "柔石",
    "ZLH-017": "胡也频", "ZLH-018": "冯铿", "ZLH-019": "殷夫", "ZLH-020": "李求实",
    "ZLH-021": "丁玲", "ZLH-022": "张天翼", "ZLH-023": "叶紫", "ZLH-024": "艾芜",
    "ZLH-025": "沙汀", "ZLH-026": "胡风", "ZLH-027": "萧军", "ZLH-028": "萧红",
    "ZLH-029": "聂绀弩", "ZLH-030": "吴奚如", "ZLH-031": "葛琴", "ZLH-032": "徐懋庸",
    "ZLH-033": "周文", "ZLH-034": "彭康", "ZLH-035": "李辉英", "ZLH-036": "穆木天",
    "ZLH-037": "楼适夷", "ZLH-038": "陈荒煤", "ZLH-039": "欧阳山", "ZLH-040": "草明",
    "ZLH-041": "骆宾基", "ZLH-042": "魏金枝", "ZLH-043": "端木蕻良", "ZLH-044": "蒋锡金",
    "ZLH-045": "周立波", "ZLH-046": "邵荃麟", "ZLH-047": "林淡秋", "ZLH-048": "王任叔",
    "ZLH-049": "舒群", "ZLH-050": "于黑丁", "ZLH-051": "邱东平", "ZLH-052": "梅志",
    "ZLH-053": "黄源", "ZLH-054": "许幸之", "ZLH-055": "郑君里", "ZLH-056": "白杨",
    "ZLH-057": "赵丹", "ZLH-058": "郭沫若", "ZLH-059": "白薇", "ZLH-060": "鲁彦",
    "ZLH-061": "朱镜我", "ZLH-062": "李一芒", "ZLH-063": "任白戈", "ZLH-064": "杜埃",
    "ZLH-065": "柯仲平", "ZLH-066": "黄其奎", "ZLH-067": "任钧", "ZLH-068": "温梓川",
    "ZLH-069": "张庚", "ZLH-070": "塞克", "ZLH-071": "李霁野", "ZLH-072": "丛深",
    "ZLH-073": "王余杞", "ZLH-074": "严文井", "ZLH-075": "黄肃秋", "ZLH-076": "马宁",
    "ZLH-077": "柳倩", "ZLH-078": "王一知", "ZLH-079": "谢和赓", "ZLH-080": "陈企霞",
    "ZLH-081": "陈白尘", "ZLH-082": "宋之的", "ZLH-083": "林焕平", "ZLH-084": "陆万美",
    "ZLH-085": "方殷", "ZLH-086": "葛一虹", "ZLH-087": "于伶", "ZLH-088": "徐芳",
    "ZLH-089": "唐弢", "ZLH-090": "吴晓邦", "ZLH-091": "叶以群", "ZLH-092": "蒋天佐",
    "ZLH-093": "王亚平", "ZLH-094": "吴亮平", "ZLH-095": "何稼夫", "ZLH-096": "王统照",
    "ZLH-097": "王西彦", "ZLH-098": "沈从文", "ZLH-099": "吴奔星", "ZLH-100": "姚蓬子",
    "ZLH-101": "徐迟", "ZLH-102": "张映纶", "ZLH-103": "王莹", "ZLH-104": "叶灵凤",
    "ZLH-105": "沈起予", "ZLH-106": "张光年", "ZLH-107": "孙师毅", "ZLH-108": "陈鲤庭",
    "ZLH-109": "王林", "ZLH-110": "章泯", "ZLH-111": "朱静子", "ZLH-112": "戴望舒",
    "ZLH-113": "施蛰存", "ZLH-114": "刘呐鸥", "ZLH-115": "穆时英", "ZLH-116": "孙席珍",
    "ZLH-117": "陈波儿", "ZLH-118": "吴组缃", "ZLH-119": "何家槐", "ZLH-120": "夏征农",
    "ZLH-121": "内山完造", "ZLH-122": "许广平", "ZLH-123": "邹韬奋", "ZLH-124": "李小峰",
    "ZLH-125": "陆志庠", "ZLH-126": "沈兼士", "ZLH-127": "曹靖华", "ZLH-128": "林语堂",
    "ZLH-129": "胡愈之", "ZLH-130": "杨杏佛", "ZLH-131": "史沫特莱", "ZLH-132": "伊罗生",
    "ZLH-133": "斯诺", "ZLH-134": "鹿地亘", "ZLH-135": "池田幸子", "ZLH-136": "宋庆龄",
    "ZLH-137": "蔡元培", "ZLH-138": "何香凝", "ZLH-139": "陈望道", "ZLH-140": "鲁维查",
    "ZLH-141": "高尔基", "ZLH-142": "法捷耶夫", "ZLH-143": "罗曼·罗兰", "ZLH-144": "巴比塞",
    "ZLH-145": "萧伯纳", "ZLH-146": "马尔罗", "ZLH-147": "辛克莱", "ZLH-148": "片山潜",
    "ZLH-149": "路易·艾黎", "ZLH-150": "萧三",
}

# 左联核心成员 - 用于判断组织隶属
LEFT_LEAGUE_MEMBERS = set([f"ZLH-{i:03d}" for i in range(1, 51)])


def match_person(name):
    """匹配人物,返回ID或None"""
    name = name.strip()
    if name in PERSON_MAP:
        return PERSON_MAP[name]
    return None


def determine_relation_type(source_id, target_id, context):
    """根据上下文确定关系类型和权重"""
    context_lower = context.lower()
    
    # 组织隶属关系关键词
    org_keywords = ["左联", "左翼作家联盟", "入盟", "加入", "成员", "盟员", "常委", "执委", "理事"]
    # 师生关系关键词
    teacher_keywords = ["师从", "老师", "学生", "教授", "指导", "学习"]
    # 亲属关系关键词
    family_keywords = ["妻子", "丈夫", "父亲", "母亲", "兄", "弟", "姐", "妹", "夫人", "爱人"]
    # 通信关系关键词
    comm_keywords = ["信", "函", "电报", "通信", "书信", "致信", "来信", "回信"]
    # 合作关系关键词
    coop_keywords = ["合作", "共同", "一起", "创办", "编辑", "主编", "合译", "合著"]
    # 论战关系关键词
    debate_keywords = ["论战", "批评", "批判", "争论", "辩论", "反驳", "驳斥"]
    # 时空共现关键词
    meet_keywords = ["会见", "见面", "拜访", "访问", "来访", "会晤", "相见", "同去", "同往", "聚会", "晚餐", "宴会"]
    
    for kw in org_keywords:
        if kw in context:
            return "强关联-组织隶属", 4
    
    for kw in teacher_keywords:
        if kw in context:
            return "强关联-师生", 4
    
    for kw in family_keywords:
        if kw in context:
            return "强关联-亲属", 5
    
    for kw in debate_keywords:
        if kw in context:
            return "弱关联-论战", 3
    
    for kw in coop_keywords:
        if kw in context:
            return "弱关联-合作", 3
    
    for kw in comm_keywords:
        if kw in context:
            return "弱关联-通信", 2
    
    for kw in meet_keywords:
        if kw in context:
            return "弱关联-时空共现", 2
    
    # 如果都是左联成员，默认为组织隶属
    if source_id in LEFT_LEAGUE_MEMBERS and target_id in LEFT_LEAGUE_MEMBERS:
        return "强关联-组织隶属", 3
    
    # 默认为时空共现
    return "弱关联-时空共现", 2


def extract_relationships_from_text(text, page_num):
    """从文本中提取人物关系"""
    relationships = []
    evidence_ref = f"左联回忆录 第{page_num}页"
    
    # 查找文本中出现的所有人物
    found_persons = []
    for name, person_id in PERSON_MAP.items():
        if len(name) >= 2 and name in text:
            # 记录位置
            for match in re.finditer(re.escape(name), text):
                found_persons.append({
                    'name': name,
                    'id': person_id,
                    'pos': match.start()
                })
    
    # 去重（同一人物可能多次出现）
    seen_ids = set()
    unique_persons = []
    for p in sorted(found_persons, key=lambda x: x['pos']):
        if p['id'] not in seen_ids:
            seen_ids.add(p['id'])
            unique_persons.append(p)
    
    # 如果找到多个不同的人物，建立两两关系
    if len(unique_persons) >= 2:
        for i in range(len(unique_persons)):
            for j in range(i + 1, len(unique_persons)):
                p1, p2 = unique_persons[i], unique_persons[j]
                
                # 提取上下文 (两人名字之间的文本)
                start = min(p1['pos'], p2['pos'])
                end = max(p1['pos'], p2['pos']) + len(unique_persons[j]['name'])
                context_start = max(0, start - 20)
                context_end = min(len(text), end + 20)
                context = text[context_start:context_end]
                
                # 确定关系类型
                rel_type, weight = determine_relation_type(p1['id'], p2['id'], context)
                
                relationships.append({
                    'source_id': p1['id'],
                    'target_id': p2['id'],
                    'relation_type': rel_type,
                    'context': context[:100],  # 限制长度
                    'evidence_ref': evidence_ref,
                    'weight': weight
                })
    
    return relationships


def load_existing_data(excel_path, sheet_name='Sheet2'):
    """加载现有Excel数据"""
    existing = {}
    try:
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in range(2, sheet.max_row + 1):
                source_id = sheet.cell(row=row, column=2).value
                target_id = sheet.cell(row=row, column=3).value
                if source_id and target_id:
                    key = (source_id, target_id)
                    existing[key] = {
                        'relation_type': sheet.cell(row=row, column=4).value,
                        'context': sheet.cell(row=row, column=5).value or "",
                        'evidence_ref': sheet.cell(row=row, column=6).value or "",
                        'weight': sheet.cell(row=row, column=7).value or 1
                    }
        wb.close()
    except Exception as e:
        print(f"加载现有数据时出错: {e}")
    return existing


def merge_relationships(existing, new_rels):
    """合并现有数据和新数据"""
    merged = dict(existing)  # 复制现有数据
    
    for rel in new_rels:
        key = (rel['source_id'], rel['target_id'])
        reverse_key = (rel['target_id'], rel['source_id'])
        
        # 检查是否已存在（包括反向）
        if key in merged:
            # 合并：取较高权重，合并上下文和证据
            old = merged[key]
            merged[key] = {
                'relation_type': rel['relation_type'] if rel['weight'] > old['weight'] else old['relation_type'],
                'context': old['context'] + "; " + rel['context'] if rel['context'] not in old['context'] else old['context'],
                'evidence_ref': old['evidence_ref'] + "; " + rel['evidence_ref'] if rel['evidence_ref'] not in old['evidence_ref'] else old['evidence_ref'],
                'weight': max(old['weight'], rel['weight'])
            }
        elif reverse_key in merged:
            # 存在反向关系，也合并
            old = merged[reverse_key]
            merged[reverse_key] = {
                'relation_type': rel['relation_type'] if rel['weight'] > old['weight'] else old['relation_type'],
                'context': old['context'] + "; " + rel['context'] if rel['context'] not in old['context'] else old['context'],
                'evidence_ref': old['evidence_ref'] + "; " + rel['evidence_ref'] if rel['evidence_ref'] not in old['evidence_ref'] else old['evidence_ref'],
                'weight': max(old['weight'], rel['weight'])
            }
        else:
            # 新关系
            merged[key] = {
                'relation_type': rel['relation_type'],
                'context': rel['context'],
                'evidence_ref': rel['evidence_ref'],
                'weight': rel['weight']
            }
    
    return merged


def main():
    pdf_path = r'd:\1大创\中国文学史资料全编 现代卷 左联回忆录 (《左联回忆录》编辑组编著).pdf'
    excel_path = r'd:\1大创\大创数据收集1.xlsx'
    
    print("=" * 60)
    print("从《左联回忆录》提取人际关系数据")
    print("=" * 60)
    
    # 1. 加载现有数据
    print("\n[1/4] 加载现有Excel数据...")
    existing_data = load_existing_data(excel_path)
    print(f"  现有记录: {len(existing_data)} 条")
    
    # 2. 提取PDF文本并识别关系
    print("\n[2/4] 读取PDF并提取关系...")
    all_new_rels = []
    
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"  PDF共 {total_pages} 页")
        
        for i, page in enumerate(pdf.pages):
            if (i + 1) % 50 == 0:
                print(f"  处理第 {i+1}/{total_pages} 页...")
            
            text = page.extract_text()
            if text:
                rels = extract_relationships_from_text(text, i + 1)
                all_new_rels.extend(rels)
    
    print(f"  从PDF提取到 {len(all_new_rels)} 条原始关系")
    
    # 3. 合并数据
    print("\n[3/4] 合并数据（去重）...")
    merged_data = merge_relationships(existing_data, all_new_rels)
    print(f"  合并后共 {len(merged_data)} 条关系")
    print(f"  新增 {len(merged_data) - len(existing_data)} 条")
    
    # 4. 写入Excel
    print("\n[4/4] 写入Excel...")
    wb = openpyxl.load_workbook(excel_path)
    
    # 确保Sheet2存在
    if 'Sheet2' not in wb.sheetnames:
        wb.create_sheet('Sheet2')
    sheet = wb['Sheet2']
    
    # 写入表头
    headers = ['序号', 'Source_ID', 'Target_ID', 'Relation_Type', 'Context', 'Evidence_Ref', 'Weight']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)
    
    # 清除旧数据
    for row in range(2, sheet.max_row + 1):
        for col in range(1, 8):
            sheet.cell(row=row, column=col, value=None)
    
    # 写入合并后的数据
    row_num = 2
    for seq, ((source_id, target_id), data) in enumerate(merged_data.items(), 1):
        sheet.cell(row=row_num, column=1, value=seq)
        sheet.cell(row=row_num, column=2, value=source_id)
        sheet.cell(row=row_num, column=3, value=target_id)
        sheet.cell(row=row_num, column=4, value=data['relation_type'])
        sheet.cell(row=row_num, column=5, value=data['context'][:200] if data['context'] else "")  # 限制长度
        sheet.cell(row=row_num, column=6, value=data['evidence_ref'][:200] if data['evidence_ref'] else "")
        sheet.cell(row=row_num, column=7, value=data['weight'])
        row_num += 1
    
    wb.save(excel_path)
    print(f"\n✓ 成功写入 {len(merged_data)} 条关系数据到 Sheet2")
    print(f"  保存到: {excel_path}")
    
    # 统计信息
    print("\n" + "=" * 60)
    print("人物出现频次统计 (前20):")
    person_count = defaultdict(int)
    for (s, t) in merged_data.keys():
        person_count[s] += 1
        person_count[t] += 1
    
    for person_id, count in sorted(person_count.items(), key=lambda x: -x[1])[:20]:
        name = ID_TO_NAME.get(person_id, person_id)
        print(f"  {name} ({person_id}): {count} 次")


if __name__ == '__main__':
    main()
