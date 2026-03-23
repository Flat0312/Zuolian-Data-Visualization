"""
fix_sheet2.py
════════════════════════════════════════════════════════════
一键修复 context_extracted.csv 的四类数据质量问题：
  P0-1. Relation_Type 剥离「强/弱关联-」前缀
  P0-2. 去除 262 行完全重复记录（保留 Weight 最大的行）
  P1.   Context OCR 空格残留再次清理
  P2.   多来源拼接（含 / ）只保留第一段
输出：fixed_Sheet2.csv，并重新写回 xlsx
════════════════════════════════════════════════════════════
"""
import re
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_CSV   = Path(r"d:\1大创\context_extracted.csv")
OUTPUT_CSV  = Path(r"d:\1大创\fixed_Sheet2.csv")
XLSX_BACKUP = Path(r"d:\1大创\《左联相关档案资源目录》_备份_20260225_111914.xlsx")
OUTPUT_XLSX = Path(r"d:\1大创\《左联相关档案资源目录》_更新版.xlsx")

# ── 正则 ──────────────────────────────────────────────────
PREFIX_RE = re.compile(r"^(强|弱)(关联|关系)\s*[-—–]\s*")
CJK = r"[\u4e00-\u9fff\u3400-\u4dbf\uf900-\ufaff]"

def strip_prefix(s: str) -> str:
    """剥离「强关联-」「弱关联-」等前缀。"""
    return PREFIX_RE.sub("", str(s)).strip()

def clean_ocr_spaces(s: str) -> str:
    """消除 CJK 字符之间 OCR 产生的多余空格。"""
    s = re.sub(rf"({CJK}) +({CJK})", r"\1\2", s)
    s = re.sub(rf"({CJK}) +({CJK})", r"\1\2", s)  # 再过一遍
    return s

def keep_first_source(s: str) -> str:
    """多来源拼接时只保留第一段（去掉 ' / ' 后的部分）。"""
    return s.split(" / ")[0].strip()

# ── 读取 Sheet1（用于亲属修正 role 映射） ─────────────────
_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

def cell_value(elem) -> str:
    is_t = elem.findall(".//main:is//main:t", _NS)
    if is_t:
        return "".join(e.text or "" for e in is_t)
    v = elem.find("main:v", _NS)
    return v.text if v is not None else ""

def read_sheet1(xlsx_path: Path) -> pd.DataFrame:
    with zipfile.ZipFile(xlsx_path) as z:
        with z.open("xl/worksheets/sheet1.xml") as f:
            root = ET.parse(f).getroot()
    rows = root.findall(".//main:sheetData/main:row", _NS)
    headers = [cell_value(c) for c in rows[0].findall("main:c", _NS)]
    data = []
    for r in rows[1:]:
        vals = [cell_value(c) for c in r.findall("main:c", _NS)]
        ncols = len(headers)
        vals = vals + [""] * (ncols - len(vals)) if len(vals) < ncols else vals[:ncols]
        data.append(dict(zip(headers, vals)))
    return pd.DataFrame(data)

INTERNAL = {"核心领导", "普通成员"}
EXTERNAL = {"外围联络人", "相关人士"}

def fix_kinship(row, role_map: dict) -> str:
    if row["Relation_Type"] != "亲属":
        return row["Relation_Type"]
    src, tgt = str(row["Source_ID"]), str(row["Target_ID"])
    sr, tr = role_map.get(src, "未知"), role_map.get(tgt, "未知")
    if sr in INTERNAL and tr in INTERNAL:
        return "组织隶属"
    return "交游"

# ── openpyxl 写入工具 ──────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)

def write_df_to_sheet(ws, df: pd.DataFrame):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            val = str(value) if value not in (None, "") else None
            if val and re.fullmatch(r"-?\d+(\.\d+)?", val):
                try:
                    val = float(val) if "." in val else int(val)
                except ValueError:
                    pass
            ws.cell(row=row_idx, column=col_idx, value=val)
            ws.cell(row=row_idx, column=col_idx).alignment = CELL_ALIGN
    for col_idx, col_name in enumerate(df.columns, start=1):
        sample = df.iloc[:200, col_idx - 1].astype(str)
        max_len = max(sample.str.len().max(), len(col_name))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"

# ════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  Sheet2 四类数据质量修复")
    print("=" * 60)

    # 读取
    df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig")
    print(f"\n原始行数: {len(df)}")

    # P0-1. 剥离 Relation_Type 前缀
    print("\n[P0-1] 剥离 Relation_Type 前缀...")
    before = df["Relation_Type"].value_counts().to_dict()
    df["Relation_Type"] = df["Relation_Type"].astype(str).apply(strip_prefix)
    after = df["Relation_Type"].value_counts().to_dict()
    print("  修复后分布:", after)

    # 亲属修正
    print("  执行亲属→角色逻辑修正...")
    df_s1    = read_sheet1(XLSX_BACKUP)
    role_map = dict(zip(df_s1["Entity_ID"], df_s1["Role"]))
    df["Relation_Type"] = df.apply(lambda r: fix_kinship(r, role_map), axis=1)
    after2 = df["Relation_Type"].value_counts().to_dict()
    print("  亲属修正后:", after2)

    # P1. OCR 空格清理
    print("\n[P1] 清理 Context OCR 空格残留...")
    before_ocr = df["Context"].fillna("").astype(str).str.contains(
        rf"{CJK} {CJK}", regex=True).sum()
    df["Context"] = df["Context"].fillna("").astype(str).apply(clean_ocr_spaces)
    after_ocr = df["Context"].str.contains(rf"{CJK} {CJK}", regex=True).sum()
    print(f"  OCR空格行: {before_ocr} → {after_ocr}")

    # P2. 多来源拼接只保留第一段
    print("\n[P2] 移除多来源拼接（只保留第一段）...")
    multi_before = df["Context"].str.contains(" / ").sum()
    df["Context"] = df["Context"].apply(keep_first_source)
    multi_after  = df["Context"].str.contains(" / ").sum()
    print(f"  多来源行: {multi_before} → {multi_after}")

    # P0-2. 去重（同 Source_ID + Target_ID + Relation_Type，保留 Weight 最大的行）
    print("\n[P0-2] 去除重复行...")
    before_len = len(df)
    df["Weight_num"] = pd.to_numeric(df["Weight"], errors="coerce").fillna(0)
    df = df.sort_values("Weight_num", ascending=False)
    df = df.drop_duplicates(subset=["Source_ID", "Target_ID", "Relation_Type"], keep="first")
    df = df.drop(columns=["Weight_num"])
    df = df.reset_index(drop=True)
    # 重建序号
    df["序号"] = range(1, len(df) + 1)
    after_len = len(df)
    print(f"  {before_len} → {after_len}（去除 {before_len - after_len} 行重复）")

    # 保存 CSV
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    print(f"\n  已保存: {OUTPUT_CSV}")

    # 写回 xlsx
    print("\n[写入 xlsx]")
    df_s1_clean = df_s1.copy()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    write_df_to_sheet(ws1, df_s1_clean)
    ws2 = wb.create_sheet("Sheet2")
    write_df_to_sheet(ws2, df)
    wb.save(OUTPUT_XLSX)
    print(f"  已保存: {OUTPUT_XLSX}")
    print(f"  Sheet1: {len(df_s1_clean)} 行 | Sheet2: {len(df)} 行")
    print("\n" + "=" * 60)

if __name__ == "__main__":
    main()
