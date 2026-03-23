from __future__ import annotations

import csv
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


TARGET_BASENAME = "《左联相关档案资源目录》.xlsx"
OUTPUT_BASENAME = "《左联相关档案资源目录》_修正版.xlsx"
LOG_BASENAME = "《左联相关档案资源目录》_修改日志.xlsx"
REVIEW_BASENAME = "review_needed.csv"

SHEET_ALIASES = {
    "sheet1": ["sheet1", "sheet 1", "节点", "人物", "entity", "entities"],
    "sheet2": ["sheet2", "sheet 2", "关系", "relation", "relations", "边"],
    "sheet3": ["sheet3", "sheet 3", "事件", "event", "events"],
}

SHEET1_COLS = {
    "entity_id": ["entity_id", "id", "实体id", "人物id", "entity id"],
    "true_name": ["true_name", "label", "name", "true name", "真实姓名", "姓名"],
    "birth_death": ["birth_death", "birth-death", "birth death", "生卒", "生卒年"],
    "role": ["role", "角色"],
}

SHEET2_COLS = {
    "seq": ["序号", "id", "relation_id"],
    "source_id": ["source_id", "source", "from_id"],
    "target_id": ["target_id", "target", "to_id"],
    "relation_type": ["relation_type", "relation", "关系类型"],
    "context": ["context", "证据", "说明"],
    "evidence_ref": ["evidence_ref", "evidence", "来源", "出处"],
    "weight": ["weight", "score", "权重"],
}

SHEET3_COLS = {
    "seq": ["序号", "id", "event_id"],
    "entity_id": ["entity_id", "entity", "person_id"],
    "timestamp": ["timestamp", "date", "time", "日期"],
    "hist_loc": ["hist_loc", "historical_location", "historical location", "历史地点"],
    "current_loc": ["current_loc", "current_location", "current location", "今址", "现址"],
    "coord_xy": ["coord_xy", "coord", "coordinates", "坐标"],
    "event": ["event", "event_name", "事件", "事件名称"],
}

ORG_KEYWORDS = ["左联", "联盟", "成员", "执行委", "执行委员会", "常委", "书记", "秘书处", "发起人", "同盟", "社"]
VISIT_KEYWORDS = ["来访", "会见", "会面", "拜访", "同往", "看屋", "看房", "同赴"]
COMM_KEYWORDS = ["来信", "收信", "寄信", "通信", "寄函", "收函"]
COOP_KEYWORDS = ["联名", "签署", "共同", "合编", "合作", "编辑", "创办", "发表"]
DEBATE_KEYWORDS = ["论战", "论争", "批判", "批评", "驳", "争论"]
MEMORIAL_KEYWORDS = ["悼", "纪念", "追悼", "悼念", "缅怀"]
KINSHIP_KEYWORDS = ["妻", "夫人", "丈夫", "母", "父", "子", "女", "兄", "弟", "姐", "妹", "伴侣"]
TEACH_KEYWORDS = ["老师", "先生", "学生", "授业", "师从", "弟子", "讲课", "指导"]
ACTIVE_EVENT_KEYWORDS = ["文学活动", "上海文学活动", "被捕事件", "被捕", "秘密会议", "成立大会", "活动", "交流会", "会面"]

VERIFIED_EVENT_RULES: Dict[str, Dict[str, Any]] = {
    "左联成立大会": {
        "standard_event_name": "中国左翼作家联盟成立大会",
        "corrected_date": "1930-03-02",
        "date_precision": "日",
        "historical_location": "中华艺术大学教室（原窦乐安路233弄）",
        "current_address": "上海市虹口区多伦路201弄2号",
        "corrected_persons": "",
        "correction_reason": "公开资料表明左联成立大会于1930年3月2日在中华艺术大学借用教室召开，今址为左联会址纪念馆；原表存在公啡咖啡馆、前哨、复旦大学等冲突地点。",
        "source_url": "https://www.shhk.gov.cn/xwzx/002008/002008040/20221031/bd8cb3ee-198a-431a-adf7-781e9fc5185d.html ; https://www.shhk.gov.cn/xwzx/002003/20250303/ec139c7d-8fa3-4970-a5dd-3248468989c8.html",
        "confidence": "high",
        "manual_review_default": False,
    },
    "五烈士遇难": {
        "standard_event_name": "左联五烈士遇难",
        "corrected_date": "1931-02-07",
        "date_precision": "日",
        "historical_location": "上海龙华淞沪警备司令部刑场",
        "current_address": "上海龙华烈士陵园（门牌待复核）",
        "corrected_persons": "柔石；胡也频；李伟森；冯铿；殷夫",
        "correction_reason": "官方公开资料明确左联五烈士于1931年2月7日在上海龙华遇难；原表将日期统一写成1931-01-01，且今址门牌存在多个冲突版本。",
        "source_url": "https://www.shhk.gov.cn/slh/038001/20260302/5f983b4c-74f7-4a6a-a2b4-930dedf99970.html ; https://www.shhk.gov.cn/xwzx/002006/20210722/07f62353-471f-40d9-880e-ea82be5da936.html",
        "confidence": "high",
        "manual_review_default": False,
        "participant_names": {"柔石", "胡也频", "李伟森", "冯铿", "殷夫"},
    },
    "龙华二十四烈士遇难": {
        "standard_event_name": "龙华二十四烈士遇难",
        "corrected_date": "1931-02-07",
        "date_precision": "日",
        "historical_location": "上海龙华淞沪警备司令部刑场",
        "current_address": "上海龙华烈士陵园（门牌待复核）",
        "corrected_persons": "",
        "correction_reason": "中国共产党新闻网资料表明龙华二十四烈士于1931年2月7日被秘密集体枪杀；原表日期为占位值1931-01-01，且部分关联人物并非可直接确认的二十四烈士成员。",
        "source_url": "https://cpc.people.com.cn/n1/2022/1209/c443712-32583679.html",
        "confidence": "medium",
        "manual_review_default": True,
    },
    "内山书店秘密会议": {
        "standard_event_name": "内山书店秘密会议",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "内山书店旧址",
        "current_address": "四川北路2050号",
        "corrected_persons": "",
        "correction_reason": "公开资料可确认内山书店旧址位于四川北路2050号；原表在2040、2044-2048、2050等多个地址之间冲突，但暂未检得同等级公开来源可精确确认该次“秘密会议”的具体日期与参与人。",
        "source_url": "https://www.shhk.gov.cn/xwzx/002006/20210930/96ecb0ec-79e3-49ef-a097-a89c5a5dbc40.html ; https://www.shhk.gov.cn/xwzx/002008/002008040/20240425/3e9546e4-0e0f-409d-a0e2-91bc115f8f66.html",
        "confidence": "medium",
        "manual_review_default": True,
    },
    "鲁迅与柔石会面": {
        "standard_event_name": "鲁迅与柔石会面",
        "corrected_date": None,
        "date_precision": None,
        "historical_location": "北四川路一带",
        "current_address": "",
        "corrected_persons": "鲁迅；柔石",
        "correction_reason": "公开转录材料可支持1930年3月28日与3月31日相关看屋活动；1929年8月29日与1930年3月30日条目仍需回查《鲁迅日记》原文，不宜强行改写。",
        "source_url": "https://al3tai.nenzhu.com/news-id-2373.html",
        "confidence": "medium",
        "manual_review_default": True,
    },
}

ROW_EVENT_OVERRIDES: Dict[int, Dict[str, Any]] = {
    7: {
        "corrected_date": "",
        "date_precision": "",
        "correction_reason": "公开可检索转录材料更明确出现的是1930-03-31同柔石看屋，当前1930-03-30记录未找到同等级公开佐证，保留待人工复核。",
        "source_url": "https://al3tai.nenzhu.com/news-id-2373.html",
        "confidence": "low",
        "needs_manual_review": "yes",
    },
    8: {
        "corrected_date": "",
        "date_precision": "",
        "correction_reason": "公开可检索转录材料更明确出现的是1930-03-31同柔石看屋，当前1930-03-30记录未找到同等级公开佐证，保留待人工复核。",
        "source_url": "https://al3tai.nenzhu.com/news-id-2373.html",
        "confidence": "low",
        "needs_manual_review": "yes",
    },
}

VERIFICATION_SOURCES = [
    ("左联会址纪念馆与成立大会旧址", "https://www.shhk.gov.cn/xwzx/002008/002008040/20221031/bd8cb3ee-198a-431a-adf7-781e9fc5185d.html"),
    ("左联成立95周年主题活动", "https://www.shhk.gov.cn/xwzx/002003/20250303/ec139c7d-8fa3-4970-a5dd-3248468989c8.html"),
    ("左联五烈士专题纪念", "https://www.shhk.gov.cn/slh/038001/20260302/5f983b4c-74f7-4a6a-a2b4-930dedf99970.html"),
    ("左联会址纪念馆五烈士介绍", "https://www.shhk.gov.cn/xwzx/002006/20210722/07f62353-471f-40d9-880e-ea82be5da936.html"),
    ("龙华二十四烈士资料", "https://cpc.people.com.cn/n1/2022/1209/c443712-32583679.html"),
    ("内山书店旧址说明", "https://www.shhk.gov.cn/xwzx/002006/20210930/96ecb0ec-79e3-49ef-a097-a89c5a5dbc40.html"),
    ("内山书店今址活动页", "https://www.shhk.gov.cn/xwzx/002008/002008040/20240425/3e9546e4-0e0f-409d-a0e2-91bc115f8f66.html"),
    ("鲁迅与柔石看屋转录材料", "https://al3tai.nenzhu.com/news-id-2373.html"),
    ("公啡咖啡馆与左联筹备会", "https://www.thepaper.cn/newsDetail_forward_8389375"),
]


@dataclass
class EntityInfo:
    entity_id: str
    name: str
    birth_year: Optional[int]
    death_year: Optional[int]
    role: str


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_\-]+", "", str(value or "").strip().lower())


def find_input_file(cwd: Path) -> Path:
    preferred = cwd / "知识库" / "data" / TARGET_BASENAME
    if preferred.exists():
        return preferred
    exact = [p for p in cwd.rglob(TARGET_BASENAME) if "_修正版" not in p.name and "_修改日志" not in p.name]
    if exact:
        return sorted(exact, key=lambda p: len(str(p)))[0]
    candidates = [p for p in cwd.rglob("*.xlsx") if "左联相关档案资源目录" in p.name and "_修正版" not in p.name and "_修改日志" not in p.name]
    if candidates:
        return sorted(candidates, key=lambda p: len(str(p)))[0]
    raise FileNotFoundError(f"未找到输入文件：{TARGET_BASENAME}")


def map_sheet_name(sheet_names: Iterable[str], aliases: List[str]) -> str:
    alias_set = {normalize_header(a) for a in aliases}
    for name in sheet_names:
        if normalize_header(name) in alias_set:
            return name
    raise KeyError(f"未找到匹配 sheet：{aliases}")


def map_columns(headers: List[Any], alias_map: Dict[str, List[str]]) -> Dict[str, int]:
    normalized_headers = {normalize_header(h): idx for idx, h in enumerate(headers)}
    result: Dict[str, int] = {}
    for logical_field, aliases in alias_map.items():
        found = None
        for alias in aliases:
            alias_norm = normalize_header(alias)
            if alias_norm in normalized_headers:
                found = normalized_headers[alias_norm]
                break
        if found is None:
            raise KeyError(f"未找到列：{logical_field}")
        result[logical_field] = found
    return result


def parse_birth_death(text: Any) -> Tuple[Optional[int], Optional[int]]:
    if text is None:
        return None, None
    m = re.match(r"^\s*(\d{4}|\?)\s*-\s*(\d{4}|\?)\s*$", str(text))
    if not m:
        return None, None
    birth = int(m.group(1)) if m.group(1).isdigit() else None
    death = int(m.group(2)) if m.group(2).isdigit() else None
    return birth, death


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def extract_year(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year
    s = as_text(value)
    m = re.match(r"^(\d{4})", s)
    return int(m.group(1)) if m else None


def infer_date_precision(value: Any) -> str:
    s = as_text(value)
    if not s:
        return ""
    if isinstance(value, datetime):
        return "日"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return "年" if s.endswith("-01-01") else "日"
    if re.fullmatch(r"\d{4}-\d{2}", s):
        return "月"
    if re.fullmatch(r"\d{4}", s):
        return "年"
    return ""


def is_placeholder_jan1(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, datetime):
        return value.month == 1 and value.day == 1
    return bool(re.fullmatch(r"\d{4}-01-01", as_text(value)))


def contains_any(text: str, keywords: Iterable[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def is_list_like(text: str) -> bool:
    if not text:
        return False
    score = text.count("、") + text.count(";") + text.count("；") + text.count(",")
    return score >= 6 or ("等" in text and text.count("、") >= 3)


def has_ocr_noise(text: str) -> bool:
    if not text:
        return False
    patterns = [r"[`]{2,}", r"[_]{2,}", r"[“”\"]{3,}", r"[A-Za-z]{6,}", r"[…]{2,}", r"[)\(]{2,}", r"�"]
    return any(re.search(pattern, text) for pattern in patterns)


def relation_risk_assessment(relation_type: str, context: str) -> Tuple[int, str, List[str]]:
    score = 0
    reasons: List[str] = []
    if relation_type in {"组织隶属", "亲属关系", "师生关系"}:
        score += 45
        reasons.append("人物-人物关系使用强证据型关系标签")
    if relation_type == "组织隶属":
        score += 15
        reasons.append("人物-人物之间出现“组织隶属”语义不自然")
    if is_list_like(context):
        score += 25
        reasons.append("证据更像长名单并列共现")
    if has_ocr_noise(context):
        score += 20
        reasons.append("存在明显OCR噪声")
    if relation_type == "亲属关系" and not contains_any(context, KINSHIP_KEYWORDS):
        score += 20
        reasons.append("缺少亲属关键词")
    if relation_type == "师生关系" and not contains_any(context, TEACH_KEYWORDS):
        score += 20
        reasons.append("缺少师生关键词")
    if relation_type == "组织隶属" and not contains_any(context, ORG_KEYWORDS):
        score += 15
        reasons.append("缺少组织归属关键词")
    if score >= 80:
        return score, "critical", reasons
    if score >= 60:
        return score, "high", reasons
    if score >= 35:
        return score, "medium", reasons
    return score, "low", reasons


def infer_relation_type(relation_type: str, context: str, risk_level: str) -> Tuple[str, str, str, str]:
    text = context or ""
    if relation_type == "组织隶属":
        if contains_any(text, COMM_KEYWORDS):
            return "通信", "relation_type_mismatch", "medium", "no"
        if contains_any(text, VISIT_KEYWORDS):
            return "交往", "relation_type_mismatch", "medium", "no"
        if contains_any(text, COOP_KEYWORDS):
            return "合作", "relation_type_mismatch", "medium", "no"
        if contains_any(text, DEBATE_KEYWORDS):
            return "论战", "relation_type_mismatch", "medium", "no"
        if contains_any(text, MEMORIAL_KEYWORDS):
            return "纪念/悼念", "relation_type_mismatch", "medium", "no"
        if contains_any(text, ORG_KEYWORDS):
            return "同属组织", "person_person_org_relation", "low", "yes"
        return "待核验", "weak_evidence", "low", "yes"
    if relation_type in {"亲属关系", "师生关系"}:
        if relation_type == "亲属关系" and contains_any(text, KINSHIP_KEYWORDS):
            return relation_type, "verified_direct_evidence", "medium", "no"
        if relation_type == "师生关系" and contains_any(text, TEACH_KEYWORDS):
            return relation_type, "verified_direct_evidence", "medium", "no"
        if contains_any(text, COMM_KEYWORDS):
            return "通信", "relation_type_mismatch", "low", "yes"
        if contains_any(text, VISIT_KEYWORDS):
            return "交往", "relation_type_mismatch", "low", "yes"
        return "待核验", "strong_claim_without_support", "low", "yes"
    if risk_level in {"critical", "high"} and is_list_like(text):
        return "待核验", "weak_evidence", "low", "yes"
    return relation_type, "", "", ""


def is_personal_activity_event(event_name: str) -> bool:
    return any(keyword in event_name for keyword in ACTIVE_EVENT_KEYWORDS)


def clone_sheet(wb, source_name: str, target_name: str):
    if target_name in wb.sheetnames:
        del wb[target_name]
    ws = wb.copy_worksheet(wb[source_name])
    ws.title = target_name
    return ws


def append_headers(ws, headers: List[str]) -> Dict[str, int]:
    start_col = ws.max_column + 1
    positions: Dict[str, int] = {}
    for offset, header in enumerate(headers):
        col = start_col + offset
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = copy(ws.cell(row=1, column=1).font) if ws.cell(row=1, column=1).font else Font(bold=True)
        positions[header] = col
    return positions


def write_value(ws, row_num: int, col_idx: int, value: Any):
    ws.cell(row=row_num, column=col_idx, value=value)


def add_log(logs: List[Dict[str, Any]], **kwargs):
    if as_text(kwargs["original_value"]) == as_text(kwargs["new_value"]):
        return
    logs.append(
        {
            "sheet_name": kwargs["sheet_name"],
            "row_number": kwargs["row_number"],
            "primary_key": kwargs["primary_key"],
            "column_name": kwargs["column_name"],
            "original_value": as_text(kwargs["original_value"]),
            "new_value": as_text(kwargs["new_value"]),
            "issue_type": kwargs["issue_type"],
            "correction_reason": kwargs["correction_reason"],
            "source_url": kwargs["source_url"],
            "evidence_ref_used": kwargs["evidence_ref_used"],
            "confidence": kwargs["confidence"],
            "needs_manual_review": kwargs["needs_manual_review"],
        }
    )


def create_aux_sheet(wb, title: str):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def build_entity_map(ws, mapping: Dict[str, int]) -> Dict[str, EntityInfo]:
    entities: Dict[str, EntityInfo] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        entity_id = as_text(row[mapping["entity_id"]])
        if not entity_id:
            continue
        birth_year, death_year = parse_birth_death(row[mapping["birth_death"]])
        entities[entity_id] = EntityInfo(entity_id, as_text(row[mapping["true_name"]]), birth_year, death_year, as_text(row[mapping["role"]]))
    return entities


def main():
    cwd = Path.cwd()
    input_path = find_input_file(cwd)
    output_dir = cwd
    output_path = output_dir / OUTPUT_BASENAME
    log_path = output_dir / LOG_BASENAME
    review_path = output_dir / REVIEW_BASENAME

    wb = load_workbook(input_path)
    sheet1_name = map_sheet_name(wb.sheetnames, SHEET_ALIASES["sheet1"])
    sheet2_name = map_sheet_name(wb.sheetnames, SHEET_ALIASES["sheet2"])
    sheet3_name = map_sheet_name(wb.sheetnames, SHEET_ALIASES["sheet3"])

    ws1 = wb[sheet1_name]
    ws2 = wb[sheet2_name]
    ws3 = wb[sheet3_name]

    headers1 = [c.value for c in ws1[1]]
    headers2 = [c.value for c in ws2[1]]
    headers3 = [c.value for c in ws3[1]]

    map1 = map_columns(headers1, SHEET1_COLS)
    map2 = map_columns(headers2, SHEET2_COLS)
    map3 = map_columns(headers3, SHEET3_COLS)
    entities = build_entity_map(ws1, map1)

    event_clusters: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"rows": [], "dates": set(), "locs": set(), "entities": set()})
    for row_num in range(2, ws3.max_row + 1):
        event_name = as_text(ws3.cell(row=row_num, column=map3["event"] + 1).value)
        timestamp = ws3.cell(row=row_num, column=map3["timestamp"] + 1).value
        hist_loc = as_text(ws3.cell(row=row_num, column=map3["hist_loc"] + 1).value)
        cur_loc = as_text(ws3.cell(row=row_num, column=map3["current_loc"] + 1).value)
        entity_id = as_text(ws3.cell(row=row_num, column=map3["entity_id"] + 1).value)
        event_clusters[event_name]["rows"].append(row_num)
        event_clusters[event_name]["dates"].add(as_text(timestamp))
        event_clusters[event_name]["locs"].add((hist_loc, cur_loc))
        event_clusters[event_name]["entities"].add(entity_id)
    duplicate_event_clusters = sum(1 for data in event_clusters.values() if len(data["dates"]) > 1 or len(data["locs"]) > 1)

    ws2c = clone_sheet(wb, sheet2_name, "Sheet2_corrected")
    ws3c = clone_sheet(wb, sheet3_name, "Sheet3_corrected")

    sheet2_cols = append_headers(
        ws2c,
        [
            "source_name",
            "target_name",
            "is_person_to_person",
            "context_quality_flags",
            "relation_quality_score",
            "relation_risk_level",
            "original_relation_type",
            "corrected_relation_type",
            "correction_reason",
            "source_url",
            "evidence_ref_used",
            "confidence",
            "needs_manual_review",
        ],
    )
    sheet3_cols = append_headers(
        ws3c,
        [
            "entity_name",
            "original_date",
            "corrected_date",
            "date_precision",
            "original_location",
            "historical_location",
            "current_address",
            "standard_event_name",
            "corrected_persons",
            "entity_role_in_event",
            "event_conflict_flags",
            "correction_reason",
            "source_url",
            "confidence",
            "needs_manual_review",
        ],
    )

    logs: List[Dict[str, Any]] = []
    review_rows: List[Dict[str, Any]] = []
    sheet2_auto_corrected_rows, sheet2_manual_review_rows = set(), set()
    sheet3_auto_corrected_rows, sheet3_manual_review_rows = set(), set()
    high_risk_relation_count = 0

    for row_num in range(2, ws2c.max_row + 1):
        source_id = as_text(ws2c.cell(row=row_num, column=map2["source_id"] + 1).value)
        target_id = as_text(ws2c.cell(row=row_num, column=map2["target_id"] + 1).value)
        relation_type = as_text(ws2c.cell(row=row_num, column=map2["relation_type"] + 1).value)
        context = as_text(ws2c.cell(row=row_num, column=map2["context"] + 1).value)
        evidence_ref = as_text(ws2c.cell(row=row_num, column=map2["evidence_ref"] + 1).value)
        seq = as_text(ws2c.cell(row=row_num, column=map2["seq"] + 1).value)
        source_name = entities.get(source_id).name if source_id in entities else ""
        target_name = entities.get(target_id).name if target_id in entities else ""
        is_person_to_person = "yes" if source_id in entities and target_id in entities else "no"
        score, risk_level, risk_reasons = relation_risk_assessment(relation_type, context)
        if risk_level in {"critical", "high"}:
            high_risk_relation_count += 1

        corrected_relation_type, issue_type, confidence, review_flag = infer_relation_type(relation_type, context, risk_level)
        quality_flags = []
        if is_list_like(context):
            quality_flags.append("list_like_evidence")
        if has_ocr_noise(context):
            quality_flags.append("ocr_noise")
        if relation_type in {"组织隶属", "亲属关系", "师生关系"}:
            quality_flags.append("strong_claim_type")
        if not context:
            quality_flags.append("empty_context")
        context_quality_flags = ",".join(quality_flags)

        correction_reason = "结构性检测未发现需要自动改写的明确证据。"
        source_url = ""
        if corrected_relation_type != relation_type:
            parts = list(risk_reasons)
            mapping_reason = {
                "通信": "上下文直接出现来信/收信/寄信类动作",
                "交往": "上下文直接出现来访/会见/共同行动类动作",
                "合作": "上下文直接出现联名/共同编辑/合作动作",
                "论战": "上下文直接出现论争/批评/论战动作",
                "纪念/悼念": "上下文直接出现纪念/悼念动作",
                "同属组织": "证据只能支持同属组织或共同组织关联",
                "待核验": "证据不足以支撑原关系类型",
            }
            if corrected_relation_type in mapping_reason:
                parts.append(mapping_reason[corrected_relation_type])
            correction_reason = "；".join(dict.fromkeys(parts))
            sheet2_auto_corrected_rows.add(row_num)
        else:
            confidence = confidence or "medium"
            review_flag = review_flag or ("yes" if risk_level in {"critical", "high"} else "no")
        if review_flag == "yes":
            sheet2_manual_review_rows.add(row_num)

        for header, value in {
            "source_name": source_name,
            "target_name": target_name,
            "is_person_to_person": is_person_to_person,
            "context_quality_flags": context_quality_flags,
            "relation_quality_score": score,
            "relation_risk_level": risk_level,
            "original_relation_type": relation_type,
            "corrected_relation_type": corrected_relation_type,
            "correction_reason": correction_reason,
            "source_url": source_url,
            "evidence_ref_used": evidence_ref,
            "confidence": confidence or "low",
            "needs_manual_review": review_flag or "no",
        }.items():
            write_value(ws2c, row_num, sheet2_cols[header], value)

        primary_key = f"{seq}|{source_id}|{target_id}"
        if corrected_relation_type != relation_type:
            add_log(
                logs,
                sheet_name=sheet2_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Relation_Type",
                original_value=relation_type,
                new_value=corrected_relation_type,
                issue_type=issue_type or "relation_type_mismatch",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used=evidence_ref,
                confidence=confidence or "low",
                needs_manual_review=review_flag or "no",
            )
        if context_quality_flags:
            add_log(
                logs,
                sheet_name=sheet2_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="context_quality_flags",
                original_value="",
                new_value=context_quality_flags,
                issue_type="quality_annotation",
                correction_reason="结构性检测记录上下文风险标记。",
                source_url=source_url,
                evidence_ref_used=evidence_ref,
                confidence="medium",
                needs_manual_review=review_flag or "no",
            )
        if review_flag == "yes":
            review_rows.append(
                {
                    "sheet_name": sheet2_name,
                    "row_number": row_num,
                    "primary_key": primary_key,
                    "issue_summary": correction_reason,
                    "source_url": source_url,
                    "evidence_ref_used": evidence_ref,
                    "confidence": confidence or "low",
                }
            )

    for row_num in range(2, ws3c.max_row + 1):
        seq = as_text(ws3c.cell(row=row_num, column=map3["seq"] + 1).value)
        entity_id = as_text(ws3c.cell(row=row_num, column=map3["entity_id"] + 1).value)
        entity_name = entities.get(entity_id).name if entity_id in entities else ""
        timestamp = ws3c.cell(row=row_num, column=map3["timestamp"] + 1).value
        hist_loc = as_text(ws3c.cell(row=row_num, column=map3["hist_loc"] + 1).value)
        current_loc = as_text(ws3c.cell(row=row_num, column=map3["current_loc"] + 1).value)
        event_name = as_text(ws3c.cell(row=row_num, column=map3["event"] + 1).value)
        original_date = as_text(timestamp)
        original_location = " / ".join([v for v in [hist_loc, current_loc] if v])
        cluster = event_clusters.get(event_name, {"dates": set(), "locs": set(), "rows": []})

        conflict_flags: List[str] = []
        if is_placeholder_jan1(timestamp):
            conflict_flags.append("placeholder_jan1")
        if len(cluster["dates"]) > 1:
            conflict_flags.append("event_multi_dates")
        if len(cluster["locs"]) > 1:
            conflict_flags.append("event_multi_locations")

        corrected_date = original_date
        date_precision = infer_date_precision(timestamp)
        historical_location = hist_loc
        current_address = current_loc
        standard_event_name = event_name
        corrected_persons = ""
        entity_role = "unclear"
        correction_reason = ""
        source_url = ""
        confidence = "low"
        needs_manual_review = "no"

        rule = VERIFIED_EVENT_RULES.get(event_name)
        if rule:
            standard_event_name = rule.get("standard_event_name") or event_name
            if rule.get("corrected_date") is not None:
                corrected_date = rule.get("corrected_date", "")
            if rule.get("date_precision") is not None:
                date_precision = rule.get("date_precision", "")
            if rule.get("historical_location"):
                historical_location = rule["historical_location"]
            if rule.get("current_address"):
                current_address = rule["current_address"]
            corrected_persons = rule.get("corrected_persons", "")
            correction_reason = rule.get("correction_reason", "")
            source_url = rule.get("source_url", "")
            confidence = rule.get("confidence", "medium")
            needs_manual_review = "yes" if rule.get("manual_review_default") else "no"

            if event_name == "左联成立大会":
                entity_role = "待核"
                if entity_name not in {"鲁迅", "柔石", "潘汉年", "蒋光慈", "钱杏邨", "李求实", "殷夫", "艾芜"}:
                    needs_manual_review = "yes"
            elif event_name == "五烈士遇难":
                participants = rule.get("participant_names", set())
                if entity_name in participants:
                    entity_role = "直接参与者"
                else:
                    entity_role = "关联人物"
                    needs_manual_review = "yes"
                    conflict_flags.append("entity_not_direct_participant")
            elif event_name == "鲁迅与柔石会面":
                entity_role = "直接参与者" if entity_name in {"鲁迅", "柔石"} else "关联人物"
                if entity_role != "直接参与者":
                    needs_manual_review = "yes"
                    conflict_flags.append("entity_not_direct_participant")
            else:
                entity_role = "待核"

        if row_num in ROW_EVENT_OVERRIDES:
            override = ROW_EVENT_OVERRIDES[row_num]
            corrected_date = override.get("corrected_date", corrected_date)
            date_precision = override.get("date_precision", date_precision)
            correction_reason = override.get("correction_reason", correction_reason)
            source_url = override.get("source_url", source_url)
            confidence = override.get("confidence", confidence)
            needs_manual_review = override.get("needs_manual_review", needs_manual_review)
            conflict_flags.append("row_override_review")

        event_year = extract_year(timestamp)
        entity_info = entities.get(entity_id)
        if entity_info and entity_info.death_year and event_year and event_year > entity_info.death_year and is_personal_activity_event(event_name):
            conflict_flags.append("after_death_conflict")
            corrected_date = ""
            date_precision = ""
            correction_reason = (correction_reason + "；" if correction_reason else "") + f"事件年份 {event_year} 晚于人物卒年 {entity_info.death_year}，且事件命名属于个人活动型事件，无法高置信保留。"
            confidence = "high"
            needs_manual_review = "yes"
            entity_role = "冲突"

        if not rule and is_placeholder_jan1(timestamp) and "after_death_conflict" not in conflict_flags:
            year = extract_year(timestamp)
            corrected_date = str(year) if year else ""
            date_precision = "年" if year else ""
            correction_reason = "原始日期疑似以1月1日填补未知月份和日期，现仅保留年份精度。"
            confidence = "medium"
            needs_manual_review = "yes"

        if not rule and (len(cluster["dates"]) > 1 or len(cluster["locs"]) > 1):
            needs_manual_review = "yes"
            correction_reason = (correction_reason + "；" if correction_reason else "") + "同名事件存在时间或地点冲突。"

        if corrected_date and re.fullmatch(r"\d{4}", corrected_date):
            date_precision = "年"
        elif corrected_date and re.fullmatch(r"\d{4}-\d{2}", corrected_date):
            date_precision = "月"
        elif corrected_date and re.fullmatch(r"\d{4}-\d{2}-\d{2}", corrected_date):
            date_precision = "日"

        row_changed = any(
            [
                corrected_date != original_date,
                historical_location != hist_loc,
                current_address != current_loc,
                standard_event_name != event_name,
                corrected_persons != "",
                correction_reason != "",
            ]
        )
        if row_changed:
            sheet3_auto_corrected_rows.add(row_num)
        if needs_manual_review == "yes":
            sheet3_manual_review_rows.add(row_num)

        if not correction_reason and conflict_flags:
            correction_reason = "发现事件结构性冲突或精度问题。"

        for header, value in {
            "entity_name": entity_name,
            "original_date": original_date,
            "corrected_date": corrected_date,
            "date_precision": date_precision,
            "original_location": original_location,
            "historical_location": historical_location,
            "current_address": current_address,
            "standard_event_name": standard_event_name,
            "corrected_persons": corrected_persons,
            "entity_role_in_event": entity_role,
            "event_conflict_flags": ",".join(dict.fromkeys(conflict_flags)),
            "correction_reason": correction_reason,
            "source_url": source_url,
            "confidence": confidence,
            "needs_manual_review": needs_manual_review,
        }.items():
            write_value(ws3c, row_num, sheet3_cols[header], value)

        primary_key = f"{seq}|{entity_id}|{event_name}"
        if corrected_date != original_date:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Timestamp",
                original_value=original_date,
                new_value=corrected_date,
                issue_type="date_correction",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if historical_location != hist_loc:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Hist_Loc",
                original_value=hist_loc,
                new_value=historical_location,
                issue_type="location_correction",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if current_address != current_loc:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Current_Loc",
                original_value=current_loc,
                new_value=current_address,
                issue_type="location_correction",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if standard_event_name != event_name:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Event",
                original_value=event_name,
                new_value=standard_event_name,
                issue_type="event_standardization",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if corrected_persons:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="corrected_persons",
                original_value="",
                new_value=corrected_persons,
                issue_type="participant_annotation",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if needs_manual_review == "yes":
            review_rows.append(
                {
                    "sheet_name": sheet3_name,
                    "row_number": row_num,
                    "primary_key": primary_key,
                    "issue_summary": correction_reason or ",".join(conflict_flags),
                    "source_url": source_url,
                    "evidence_ref_used": "",
                    "confidence": confidence,
                }
            )

    field_map_ws = create_aux_sheet(wb, "field_mapping")
    field_map_ws.append(["sheet_name", "logical_field", "column_header", "column_index"])
    for logical, idx in map1.items():
        field_map_ws.append([sheet1_name, logical, headers1[idx], idx + 1])
    for logical, idx in map2.items():
        field_map_ws.append([sheet2_name, logical, headers2[idx], idx + 1])
    for logical, idx in map3.items():
        field_map_ws.append([sheet3_name, logical, headers3[idx], idx + 1])

    conflict_ws = create_aux_sheet(wb, "event_conflicts")
    conflict_ws.append(["event_name", "row_count", "unique_dates", "unique_locations", "sample_rows"])
    for event_name, data in sorted(event_clusters.items(), key=lambda kv: (-len(kv[1]["rows"]), kv[0])):
        if len(data["dates"]) > 1 or len(data["locs"]) > 1:
            conflict_ws.append([event_name, len(data["rows"]), len(data["dates"]), len(data["locs"]), ",".join(str(r) for r in data["rows"][:10])])

    sources_ws = create_aux_sheet(wb, "verification_sources")
    sources_ws.append(["title", "url"])
    for title, url in VERIFICATION_SOURCES:
        sources_ws.append([title, url])

    summary_ws = create_aux_sheet(wb, "summary")
    summary_ws.append(["metric", "value"])
    summary_rows = [
        ("input_file", str(input_path)),
        ("output_file", str(output_path)),
        ("log_file", str(log_path)),
        ("review_file", str(review_path)),
        ("sheet2_total_records", ws2.max_row - 1),
        ("sheet2_auto_corrected_rows", len(sheet2_auto_corrected_rows)),
        ("sheet2_manual_review_rows", len(sheet2_manual_review_rows)),
        ("sheet3_total_records", ws3.max_row - 1),
        ("sheet3_auto_corrected_rows", len(sheet3_auto_corrected_rows)),
        ("sheet3_manual_review_rows", len(sheet3_manual_review_rows)),
        ("duplicate_event_clusters", duplicate_event_clusters),
        ("high_risk_person_relations", high_risk_relation_count),
    ]
    for metric, value in summary_rows:
        summary_ws.append([metric, value])
    wb.save(output_path)

    log_wb = Workbook()
    log_ws = log_wb.active
    log_ws.title = "modification_log"
    log_headers = [
        "sheet_name",
        "row_number",
        "primary_key",
        "column_name",
        "original_value",
        "new_value",
        "issue_type",
        "correction_reason",
        "source_url",
        "evidence_ref_used",
        "confidence",
        "needs_manual_review",
    ]
    log_ws.append(log_headers)
    for item in logs:
        log_ws.append([item[h] for h in log_headers])

    log_summary_ws = log_wb.create_sheet("summary")
    log_summary_ws.append(["metric", "value"])
    for metric, value in summary_rows:
        log_summary_ws.append([metric, value])

    review_ws = log_wb.create_sheet("review_needed")
    review_headers = ["sheet_name", "row_number", "primary_key", "issue_summary", "source_url", "evidence_ref_used", "confidence"]
    review_ws.append(review_headers)
    for item in review_rows:
        review_ws.append([item[h] for h in review_headers])
    log_wb.save(log_path)

    with review_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=review_headers)
        writer.writeheader()
        for item in review_rows:
            writer.writerow(item)

    print(f"INPUT={input_path}")
    print(f"OUTPUT={output_path}")
    print(f"LOG={log_path}")
    print(f"REVIEW={review_path}")
    print(f"SHEET2_TOTAL={ws2.max_row - 1}")
    print(f"SHEET2_AUTO_CORRECTED={len(sheet2_auto_corrected_rows)}")
    print(f"SHEET2_MANUAL_REVIEW={len(sheet2_manual_review_rows)}")
    print(f"SHEET3_TOTAL={ws3.max_row - 1}")
    print(f"SHEET3_AUTO_CORRECTED={len(sheet3_auto_corrected_rows)}")
    print(f"SHEET3_MANUAL_REVIEW={len(sheet3_manual_review_rows)}")
    print(f"DUPLICATE_EVENT_CLUSTERS={duplicate_event_clusters}")
    print(f"HIGH_RISK_PERSON_RELATIONS={high_risk_relation_count}")


if __name__ == "__main__":
    main()
