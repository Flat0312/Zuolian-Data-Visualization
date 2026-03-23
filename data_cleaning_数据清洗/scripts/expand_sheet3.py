import pandas as pd
import re
from collections import OrderedDict

# ══════════════════════════════════════════════════════════
# 配置
# ══════════════════════════════════════════════════════════
INPUT  = '数据/输出结果/final_fixed_zuolian.xlsx'
OUTPUT = '数据/输出结果/final_fixed_zuolian.xlsx'
CORE_ENTITIES = {f'ZLH-{i:03d}' for i in range(1, 51)}  # 优先处理前50人

# ── 1930年代上海历史地点 → 现代地址 + 坐标 ──
LOCATION_DB = {
    # 核心左联地点
    '公啡咖啡馆':     {'current': '四川北路多伦路口',         'coord': '121.480, 31.272'},
    '左联会址':       {'current': '多伦路201弄2号',          'coord': '121.481, 31.272'},
    '左联成立大会':   {'current': '窦乐安路233号（多伦路201弄）', 'coord': '121.481, 31.272'},
    '中华艺术大学':   {'current': '多伦路201弄2号',          'coord': '121.481, 31.272'},
    # 多伦路文化街
    '多伦路':         {'current': '虹口区多伦路',            'coord': '121.481, 31.272'},
    '景云里':         {'current': '虹口区横浜路景云里',       'coord': '121.479, 31.270'},
    '拉摩斯公寓':     {'current': '四川北路2079号',          'coord': '121.482, 31.270'},
    # 内山书店与鲁迅相关
    '内山书店':       {'current': '四川北路2050号',          'coord': '121.482, 31.251'},
    '内山完造':       {'current': '四川北路2050号',          'coord': '121.482, 31.251'},
    '鲁迅故居':       {'current': '山阴路132弄9号',          'coord': '121.480, 31.270'},
    '大陆新村':       {'current': '山阴路132弄',             'coord': '121.480, 31.270'},
    '北四川路':       {'current': '四川北路',                'coord': '121.482, 31.260'},
    '四川北路':       {'current': '四川北路',                'coord': '121.482, 31.260'},
    # 虹口区
    '虹口':           {'current': '虹口区',                  'coord': '121.485, 31.265'},
    '虹口区':         {'current': '虹口区',                  'coord': '121.485, 31.265'},
    '北虹口':         {'current': '虹口区北部',              'coord': '121.485, 31.275'},
    # 出版机构
    '开明书店':       {'current': '福州路近河南路',           'coord': '121.484, 31.238'},
    '良友图书':       {'current': '北四川路851号',           'coord': '121.482, 31.255'},
    '光华书局':       {'current': '四马路（福州路）',         'coord': '121.484, 31.238'},
    '北新书局':       {'current': '四川北路',                'coord': '121.482, 31.260'},
    '生活书店':       {'current': '福州路',                  'coord': '121.484, 31.238'},
    '现代书局':       {'current': '四马路',                  'coord': '121.484, 31.238'},
    '亚东图书馆':     {'current': '五马路棋盘街',            'coord': '121.480, 31.237'},
    '商务印书馆':     {'current': '河南路宝山路',            'coord': '121.480, 31.253'},
    '中华书局':       {'current': '静安寺路',                'coord': '121.450, 31.230'},
    '创造社':         {'current': '闸北宝山路',              'coord': '121.473, 31.260'},
    '太阳社':         {'current': '闸北区',                  'coord': '121.460, 31.255'},
    # 刊物编辑部
    '萌芽月刊':       {'current': '闸北区',                  'coord': '121.460, 31.255'},
    '北斗杂志':       {'current': '虹口区',                  'coord': '121.485, 31.265'},
    '文学月报':       {'current': '虹口区',                  'coord': '121.485, 31.265'},
    '十字街头':       {'current': '虹口区',                  'coord': '121.485, 31.265'},
    '前哨':           {'current': '虹口区',                  'coord': '121.485, 31.265'},
    # 政治相关
    '龙华监狱':       {'current': '龙华烈士陵园',            'coord': '121.452, 31.182'},
    '龙华':           {'current': '龙华烈士陵园',            'coord': '121.452, 31.182'},
    '龙华淞沪警备':   {'current': '龙华烈士陵园',            'coord': '121.452, 31.182'},
    '提篮桥监狱':     {'current': '提篮桥',                  'coord': '121.499, 31.257'},
    '法租界':         {'current': '原法租界（黄浦/卢湾）',    'coord': '121.470, 31.221'},
    '公共租界':       {'current': '原公共租界（虹口/静安）',  'coord': '121.480, 31.245'},
    '工部局':         {'current': '外滩（原公共租界）',       'coord': '121.490, 31.240'},
    # 活动场所
    '东方旅社':       {'current': '南京路近河南路',           'coord': '121.483, 31.240'},
    '大世界':         {'current': '西藏南路1号',             'coord': '121.479, 31.232'},
    '大世界游乐场':   {'current': '西藏南路1号',             'coord': '121.479, 31.232'},
    '上海大戏院':     {'current': '南京西路',                'coord': '121.460, 31.235'},
    '明星影片公司':   {'current': '枫林桥徐家汇路',          'coord': '121.467, 31.210'},
    # 学校
    '上海劳动大学':   {'current': '江湾',                    'coord': '121.500, 31.310'},
    '暨南大学':       {'current': '真如',                    'coord': '121.410, 31.250'},
    '复旦大学':       {'current': '邯郸路220号',             'coord': '121.506, 31.300'},
    '上海大学':       {'current': '宝山路',                  'coord': '121.473, 31.260'},
    # 其他区域
    '闸北':           {'current': '静安区（原闸北）',         'coord': '121.460, 31.255'},
    '南京路':         {'current': '南京东路',                'coord': '121.483, 31.240'},
    '外滩':           {'current': '外滩',                    'coord': '121.490, 31.240'},
    '法华路':         {'current': '新华路',                  'coord': '121.428, 31.210'},
    '霞飞路':         {'current': '淮海中路',                'coord': '121.460, 31.220'},
    '上海':           {'current': '上海',                    'coord': '121.473, 31.230'},
}
DEFAULT_COORD = '121.47, 31.23'

# ── 事件类型推断 ──
EVENT_RULES = [
    ('左联成立大会', ['左联成立', '成立大会', '3月2日.*成立']),
    ('逮捕/牺牲',   ['被捕', '逮捕', '就义', '牺牲', '杀害', '遇害', '枪杀', '被害', '被押', '入狱']),
    ('秘密会议',     ['秘密.*会', '地下.*会', '密谋', '秘密集会', '秘密联络']),
    ('签名/抗议',    ['签名', '宣言', '抗议', '联署', '声明', '通电']),
    ('文学活动',     ['创作', '发表', '出版', '写作', '编辑', '杂志', '刊物', '月刊', '丛书']),
    ('会议/集会',    ['会议', '大会', '集会', '座谈', '讨论会', '开会']),
    ('文学论战',     ['论战', '辩论', '争论', '论争', '批判', '笔战']),
    ('社会活动',     ['演讲', '报告', '纪念', '追悼', '祝贺', '庆祝', '欢迎']),
    ('交往/拜访',    ['拜访', '访问', '会面', '会见', '相识', '结识', '来访']),
    ('教学活动',     ['授课', '讲课', '教学', '任教', '讲座', '培训']),
    ('通信/联络',    ['通信', '书信', '来信', '复信', '致函', '电报', '信件']),
]

def infer_event(context, evidence_ref):
    text = (str(context) if pd.notna(context) else '') + ' ' + (str(evidence_ref) if pd.notna(evidence_ref) else '')
    for event_type, keywords in EVENT_RULES:
        for kw in keywords:
            if re.search(kw, text):
                return event_type
    return '一般活动'


# ══════════════════════════════════════════════════════════
# 日期提取
# ══════════════════════════════════════════════════════════
def extract_dates(text):
    if not text or pd.isna(text):
        return []
    text = str(text)
    dates = []

    # 完整日期: 1930年3月2日
    for m in re.finditer(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', text):
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1920 <= y <= 1945 and 1 <= mo <= 12 and 1 <= d <= 31:
            dates.append(f'{y:04d}-{mo:02d}-{d:02d}')

    # 年月: 1930年3月
    if not dates:
        for m in re.finditer(r'(\d{4})\s*年\s*(\d{1,2})\s*月', text):
            y, mo = int(m.group(1)), int(m.group(2))
            if 1920 <= y <= 1945 and 1 <= mo <= 12:
                dates.append(f'{y:04d}-{mo:02d}-01')

    # 仅年份: 1930年
    if not dates:
        for m in re.finditer(r'(\d{4})\s*年', text):
            y = int(m.group(1))
            if 1920 <= y <= 1945:
                dates.append(f'{y:04d}-01-01')

    return list(OrderedDict.fromkeys(dates))[:3]  # 去重，最多3个


# ══════════════════════════════════════════════════════════
# 地点提取
# ══════════════════════════════════════════════════════════
# 按名称长度降序排列，优先匹配更具体的地名
LOCATION_KEYS = sorted(LOCATION_DB.keys(), key=len, reverse=True)

def extract_locations(text):
    if not text or pd.isna(text):
        return []
    text = str(text)
    found = []
    for loc in LOCATION_KEYS:
        if loc in text and loc not in found:
            found.append(loc)
    return found[:3]  # 最多3个


# ══════════════════════════════════════════════════════════
# 主处理逻辑
# ══════════════════════════════════════════════════════════
print("加载数据...")
df1 = pd.read_excel(INPUT, sheet_name='Sheet1')
df2 = pd.read_excel(INPUT, sheet_name='Sheet2')
df3 = pd.read_excel(INPUT, sheet_name='Sheet3')

valid_ids = set(df1['Entity_ID'])
new_rows = []
warnings = []

print(f"Sheet2 共 {len(df2)} 行，开始提取时空数据...")

for idx, row in df2.iterrows():
    src, tgt = row['Source_ID'], row['Target_ID']
    ctx = str(row['Context']) if pd.notna(row['Context']) else ''
    ref = str(row['Evidence_Ref']) if pd.notna(row['Evidence_Ref']) else ''
    combined = ctx + ' ' + ref

    # 提取日期和地点
    dates = extract_dates(combined)
    locs = extract_locations(combined)

    if not dates or not locs:
        continue

    event = infer_event(ctx, ref)

    # 决定为哪些实体生成行
    entities = []
    if src in valid_ids:
        entities.append(src)
    if tgt in valid_ids and tgt != src:
        entities.append(tgt)

    # 优先处理核心实体
    core = [e for e in entities if e in CORE_ENTITIES]
    non_core = [e for e in entities if e not in CORE_ENTITIES]

    for date in dates:
        for loc in locs:
            loc_info = LOCATION_DB.get(loc, None)
            if loc_info:
                current = loc_info['current']
                coord = loc_info['coord']
            else:
                current = loc
                coord = DEFAULT_COORD
                warnings.append(f"未知地点: {loc}")

            # 核心实体全部生成
            for eid in core:
                new_rows.append({
                    'Entity_ID': eid,
                    'Timestamp': date,
                    'Hist_Loc': loc,
                    'Current_Loc': current,
                    'Coord_XY': coord,
                    'Event': event,
                })

            # 非核心实体只在核心为空时生成（避免过多行）
            if not core:
                for eid in non_core:
                    new_rows.append({
                        'Entity_ID': eid,
                        'Timestamp': date,
                        'Hist_Loc': loc,
                        'Current_Loc': current,
                        'Coord_XY': coord,
                        'Event': event,
                    })

print(f"提取到 {len(new_rows)} 条候选记录")

# 合并并去重
new_df = pd.DataFrame(new_rows)
combined_df = pd.concat([df3, new_df], ignore_index=True)
combined_df = combined_df.drop_duplicates(subset=['Entity_ID', 'Timestamp', 'Hist_Loc', 'Event'])
combined_df = combined_df.reset_index(drop=True)
combined_df['序号'] = range(1, len(combined_df) + 1)

print(f"去重后: {len(combined_df)} 行")

# ══════════════════════════════════════════════════════════
# 统计
# ══════════════════════════════════════════════════════════
print(f"\n=== 统计 ===")
print(f"总行数: {len(combined_df)}")
print(f"涉及实体数: {combined_df['Entity_ID'].nunique()}")
print(f"\n时间分布:")
combined_df['_year'] = combined_df['Timestamp'].str[:4]
print(combined_df['_year'].value_counts().sort_index().to_string())
print(f"\n地点分布 (前15):")
print(combined_df['Hist_Loc'].value_counts().head(15).to_string())
print(f"\n事件类型分布:")
print(combined_df['Event'].value_counts().to_string())
print(f"\n实体分布 (前10):")
id_to_name = dict(zip(df1['Entity_ID'], df1['True_Name']))
entity_counts = combined_df['Entity_ID'].value_counts().head(10)
for eid, cnt in entity_counts.items():
    print(f"  {id_to_name.get(eid, eid)}: {cnt}")

# 去掉临时列
combined_df = combined_df.drop(columns=['_year'], errors='ignore')

if warnings:
    unique_warnings = list(OrderedDict.fromkeys(warnings))
    print(f"\n⚠ 未映射地点 ({len(unique_warnings)}):")
    for w in unique_warnings[:10]:
        print(f"  {w}")

# ══════════════════════════════════════════════════════════
# 保存
# ══════════════════════════════════════════════════════════
import shutil
if INPUT != OUTPUT:
    shutil.copy2(INPUT, OUTPUT)

from openpyxl import load_workbook
wb = load_workbook(OUTPUT)
if 'Sheet3' in wb.sheetnames:
    del wb['Sheet3']
wb.save(OUTPUT)

with pd.ExcelWriter(OUTPUT, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    combined_df.to_excel(writer, sheet_name='Sheet3', index=False)

print(f"\n✓ 已保存至: {OUTPUT}")
