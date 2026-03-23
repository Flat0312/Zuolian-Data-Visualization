"""
write_back_to_xlsx.py
════════════════════════════════════════════════════════════
将处理结果写回 《左联相关档案资源目录》.xlsx：
  - Sheet1：原样保留（来自 xlsx XML 直解）
  - Sheet2：写入解耦后的 Relation_Type + LLM 权重
使用 openpyxl 重建 xlsx，规避 sharedStrings 兼容性问题。
════════════════════════════════════════════════════════════
"""

import re
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────────────────
# 路径配置
# ────────────────────────────────────────────────────────────
XLSX_PATH    = Path(r"d:\1大创\《左联相关档案资源目录》.xlsx")
SHEET2_CSV   = Path(r"d:\1大创\context_extracted.csv")   # 使用重提取后的 Context
INTERIM_CSV  = Path(r"d:\1大创\_llm_weights_interim.csv")
OUTPUT_XLSX  = Path(r"d:\1大创\《左联相关档案资源目录》_更新版.xlsx")   # 写入新文件，请关闭原文件后自行替换

# 真实亲属白名单
KINSHIP_WHITELIST: list[frozenset] = []

# ────────────────────────────────────────────────────────────
# 1. 读取工具（XML 直解，兼容 WPS sharedStrings 缺失）
# ────────────────────────────────────────────────────────────
_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def cell_value(elem) -> str:
    is_t = elem.findall(".//main:is//main:t", _NS)
    if is_t:
        return "".join(e.text or "" for e in is_t)
    v = elem.find("main:v", _NS)
    return v.text if v is not None else ""


def read_sheet_xml(xlsx_path: Path, sheet_xml: str) -> pd.DataFrame:
    with zipfile.ZipFile(xlsx_path) as z:
        with z.open(sheet_xml) as f:
            root = ET.parse(f).getroot()

    rows = []
    for row_elem in root.findall(".//main:sheetData/main:row", _NS):
        rows.append([cell_value(c) for c in row_elem.findall("main:c", _NS)])

    if not rows:
        return pd.DataFrame()
    headers = rows[0]
    data    = rows[1:]
    ncols   = len(headers)
    data    = [r + [""] * (ncols - len(r)) if len(r) < ncols else r[:ncols] for r in data]
    return pd.DataFrame(data, columns=headers)


# ────────────────────────────────────────────────────────────
# 2. 阶段二逻辑：标签解耦（从 llm_weight_decouple.py 提取）
# ────────────────────────────────────────────────────────────
PREFIX_RE = re.compile(r"^(强|弱)(关联|关系)\s*[-—–]\s*")
INTERNAL  = {"核心领导", "普通成员"}
EXTERNAL  = {"外围联络人", "相关人士"}


def decouple_relation_type(df: pd.DataFrame, role_map: dict) -> pd.DataFrame:
    """剥离强/弱前缀，并对裸「亲属」按 Role 进行逻辑修正。"""
    df = df.copy()
    df["Relation_Type"] = df["Relation_Type"].astype(str).apply(
        lambda x: PREFIX_RE.sub("", x).strip()
    )

    def fix_kinship(row) -> str:
        if row["Relation_Type"] != "亲属":
            return row["Relation_Type"]
        src, tgt = str(row["Source_ID"]), str(row["Target_ID"])
        if frozenset([src, tgt]) in KINSHIP_WHITELIST:
            return "亲属"
        sr, tr = role_map.get(src, "未知"), role_map.get(tgt, "未知")
        if sr in INTERNAL and tr in INTERNAL:
            return "组织隶属"
        elif sr in EXTERNAL or tr in EXTERNAL:
            return "交游"
        return "交游"

    df["Relation_Type"] = df.apply(fix_kinship, axis=1)
    return df


# ────────────────────────────────────────────────────────────
# 3. 将 LLM 权重映射回行级数据
# ────────────────────────────────────────────────────────────

def map_llm_weights(df: pd.DataFrame, df_weights: pd.DataFrame) -> pd.DataFrame:
    """
    通过双向归一的 (id_a, id_b) 将 LLM weight 映射回每一行，
    覆盖原 Weight 列，新增 LLM_Reason 列。
    """
    df = df.copy()
    df["_id_a"] = df.apply(lambda r: min(str(r["Source_ID"]), str(r["Target_ID"])), axis=1)
    df["_id_b"] = df.apply(lambda r: max(str(r["Source_ID"]), str(r["Target_ID"])), axis=1)

    lookup = df_weights.set_index(["id_a", "id_b"])[["llm_weight", "llm_reason"]].to_dict("index")

    df["Weight"]     = df.apply(lambda r: lookup.get((r["_id_a"], r["_id_b"]), {}).get("llm_weight", ""), axis=1)
    df["LLM_Reason"] = df.apply(lambda r: lookup.get((r["_id_a"], r["_id_b"]), {}).get("llm_reason", ""), axis=1)
    df.drop(columns=["_id_a", "_id_b"], inplace=True)
    return df


# ────────────────────────────────────────────────────────────
# 4. 用 openpyxl 将 DataFrame 写入 xlsx 工作表
# ────────────────────────────────────────────────────────────

# 表头样式：深色背景 + 白色粗体
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)


def write_df_to_sheet(ws, df: pd.DataFrame):
    """将 DataFrame 写入已有 openpyxl worksheet。"""
    # 写表头
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # 写数据行
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            # 数值型直接写 float/int，否则写字符串
            if value != "" and str(value).replace(".", "", 1).lstrip("-").isdigit():
                try:
                    ws.cell(row=row_idx, column=col_idx, value=float(value) if "." in str(value) else int(value))
                except ValueError:
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else None)
            ws.cell(row=row_idx, column=col_idx).alignment = CELL_ALIGN

    # 自动列宽（取前200行最大宽度，上限60字符）
    for col_idx, col_name in enumerate(df.columns, start=1):
        sample_vals = df.iloc[:200, col_idx - 1].astype(str)
        max_len = max(sample_vals.str.len().max(), len(col_name))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # 冻结首行
    ws.freeze_panes = "A2"


# ────────────────────────────────────────────────────────────
# 5. 主流程
# ────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  将处理结果写回 xlsx")
    print("=" * 60)

    # ── 读取 Sheet1 ──────────────────────────────────────────
    print("\n[1/5] 读取 Sheet1（实体表）...")
    df_s1       = read_sheet_xml(XLSX_PATH, "xl/worksheets/sheet1.xml")
    role_map    = dict(zip(df_s1["Entity_ID"], df_s1["Role"]))
    print(f"      {len(df_s1)} 行，{len(df_s1.columns)} 列")

    # ── 读取 cleaned Sheet2 ───────────────────────────────────
    print("[2/5] 读取 cleaned_Sheet2.csv...")
    df_s2 = pd.read_csv(SHEET2_CSV, encoding="utf-8-sig")
    # 去除脚本附加的辅助列
    drop_cols = [c for c in ["Context_raw", "NER_persons", "NER_places"] if c in df_s2.columns]
    if drop_cols:
        df_s2.drop(columns=drop_cols, inplace=True)
    print(f"      {len(df_s2)} 行，列：{list(df_s2.columns)}")

    # ── 标签解耦 ─────────────────────────────────────────────
    print("[3/5] 执行标签解耦（剥离强/弱前缀 + 亲属修正）...")
    df_s2 = decouple_relation_type(df_s2, role_map)
    print("      Relation_Type 分布：")
    for rel, cnt in df_s2["Relation_Type"].value_counts().items():
        print(f"        {rel:20s}  {cnt}")

    # ── 映射 LLM 权重 ─────────────────────────────────────────
    print("[4/5] 映射 LLM 权重...")
    df_weights = pd.read_csv(INTERIM_CSV, encoding="utf-8-sig")
    covered    = df_weights[df_weights["llm_weight"] > 0]
    print(f"      权重缓存：{len(df_weights)} 对，有效（weight>0）：{len(covered)} 对")
    df_s2 = map_llm_weights(df_s2, df_weights)
    missing = (df_s2["Weight"] == "").sum()
    print(f"      映射完成，未匹配行：{missing}")

    # ── 写入 xlsx ─────────────────────────────────────────────
    print("[5/5] 写入 xlsx...")
    wb = Workbook()

    # Sheet1
    ws1 = wb.active
    ws1.title = "Sheet1"
    write_df_to_sheet(ws1, df_s1)
    print(f"      Sheet1 写入完成（{len(df_s1)} 行）")

    # Sheet2
    ws2 = wb.create_sheet("Sheet2")
    write_df_to_sheet(ws2, df_s2)
    print(f"      Sheet2 写入完成（{len(df_s2)} 行，含 LLM_Reason 列）")

    wb.save(OUTPUT_XLSX)
    print(f"\n      保存成功：{OUTPUT_XLSX}")
    print("=" * 60)


if __name__ == "__main__":
    main()
