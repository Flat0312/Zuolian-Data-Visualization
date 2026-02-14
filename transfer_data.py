# Transfer data from Word document to Excel spreadsheet
from docx import Document
from openpyxl import load_workbook

# Read Word document
doc = Document(r'c:\Users\33158\Desktop\大创\姓名.docx')
table = doc.tables[0]

# Extract data from Word table (skip header row)
word_data = []
for i, row in enumerate(table.rows):
    if i == 0:  # Skip header
        continue
    cells = [cell.text.strip() for cell in row.cells]
    if len(cells) >= 4 and cells[0]:  # Has valid data
        word_data.append({
            'True_Name': cells[0],      # 姓名
            'Alias': cells[1],          # 著名笔名/化名/代号
            'Birth_Death': cells[2],    # 生卒年
            'Role': cells[3]            # 在左联中的角色
        })

print(f"Extracted {len(word_data)} records from Word document")
print("\nFirst 10 records:")
for i, record in enumerate(word_data[:10]):
    print(f"{i+1}. {record['True_Name']} | {record['Alias']} | {record['Birth_Death']} | {record['Role']}")

# Load Excel workbook
wb = load_workbook(r'c:\Users\33158\Desktop\大创\大创数据收集.xlsx')
sheet = wb['Sheet1']

# Get header row to find column indices
headers = [cell.value for cell in sheet[1]]
print(f"\nExcel headers: {headers}")

# Find column indices
col_map = {}
for i, header in enumerate(headers, 1):
    if header == 'True_Name':
        col_map['True_Name'] = i
    elif header == 'Alias':
        col_map['Alias'] = i
    elif header == 'Birth_Death':
        col_map['Birth_Death'] = i
    elif header == 'Role':
        col_map['Role'] = i

print(f"Column mapping: {col_map}")

# Find rows that need data (by True_Name match or empty rows)
updated_count = 0
for row_idx in range(2, sheet.max_row + 1):
    true_name_cell = sheet.cell(row=row_idx, column=col_map['True_Name'])
    excel_name = true_name_cell.value
    
    if excel_name:
        # Find matching record in word_data
        for record in word_data:
            if record['True_Name'] == excel_name:
                # Update other columns if they're empty
                alias_cell = sheet.cell(row=row_idx, column=col_map['Alias'])
                birth_cell = sheet.cell(row=row_idx, column=col_map['Birth_Death'])
                role_cell = sheet.cell(row=row_idx, column=col_map['Role'])
                
                if not alias_cell.value and record['Alias']:
                    alias_cell.value = record['Alias']
                if not birth_cell.value and record['Birth_Death']:
                    birth_cell.value = record['Birth_Death']
                if not role_cell.value and record['Role']:
                    role_cell.value = record['Role']
                    
                updated_count += 1
                break

print(f"\nUpdated {updated_count} existing rows")

# Save workbook
wb.save(r'c:\Users\33158\Desktop\大创\大创数据收集.xlsx')
print("\nExcel file saved successfully!")
