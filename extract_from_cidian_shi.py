# -*- coding: utf-8 -*-
"""
从《左联词典》和《左联史》TXT文件中提取人际关系数据并追加到Sheet2
按页分割文本，对每页进行人物共现分析
"""

import openpyxl
import re
from collections import defaultdict
from itertools import combinations

# ===== 人物映射表（复用自 extract_zuolian.py）=====

PERSON_MAP = {
    "鲁迅": "ZLH-001", "周树人": "ZLH-001", "豫才": "ZLH-001",
    "茅盾": "ZLH-002", "沈雁冰": "ZLH-002", "玄珠": "ZLH-002", "方璧": "ZLH-002",
    "瞿秋白": "ZLH-003", "维它": "ZLH-003", "史铁儿": "ZLH-003",
    "夏衍": "ZLH-004", "沈端先": "ZLH-004", "黄子布": "ZLH-004",
    "冯雪峰": "ZLH-005", "画室": "ZLH-005", "洛扬": "ZLH-005", "雪峰": "ZLH-005",
    "潘汉年": "ZLH-006", "萧叔安": "ZLH-006", "小潘": "ZLH-006",
    "周扬": "ZLH-007", "周起应": "ZLH-007",
    "田汉": "ZLH-008", "寿昌": "ZLH-008",
    "郑伯奇": "ZLH-009", "伯奇": "ZLH-009",
    "洪深": "ZLH-010", "洪伯飞": "ZLH-010",
    "钱杏邨": "ZLH-011", "阿英": "ZLH-011", "魏如晦": "ZLH-011",
    "蒋光慈": "ZLH-012", "蒋光赤": "ZLH-012", "光慈": "ZLH-012",
    "阳翰笙": "ZLH-013", "欧阳继修": "ZLH-013", "华汉": "ZLH-013",
    "冯乃超": "ZLH-014", "冯其英": "ZLH-014",
    "郁达夫": "ZLH-015", "郁文": "ZLH-015", "达夫": "ZLH-015",
    "柔石": "ZLH-016", "赵平复": "ZLH-016",
    "胡也频": "ZLH-017", "胡崇轩": "ZLH-017", "也频": "ZLH-017",
    "冯铿": "ZLH-018", "岭梅": "ZLH-018",
    "殷夫": "ZLH-019", "徐白": "ZLH-019", "白莽": "ZLH-019",
    "李求实": "ZLH-020", "李伟森": "ZLH-020",
    "丁玲": "ZLH-021", "蒋冰之": "ZLH-021",
    "张天翼": "ZLH-022", "张元定": "ZLH-022",
    "叶紫": "ZLH-023", "余鹤林": "ZLH-023",
    "艾芜": "ZLH-024", "汤道耕": "ZLH-024",
    "沙汀": "ZLH-025", "杨朝熙": "ZLH-025",
    "胡风": "ZLH-026", "张光人": "ZLH-026", "谷非": "ZLH-026",
    "萧军": "ZLH-027", "刘鸿霖": "ZLH-027",
    "萧红": "ZLH-028", "张迺莹": "ZLH-028", "悄吟": "ZLH-028",
    "聂绀弩": "ZLH-029", "绀弩": "ZLH-029",
    "吴奚如": "ZLH-030", "吴席儒": "ZLH-030",
    "葛琴": "ZLH-031", "葛爱琴": "ZLH-031",
    "徐懋庸": "ZLH-032", "徐茂荣": "ZLH-032", "懋庸": "ZLH-032",
    "周文": "ZLH-033",
    "彭康": "ZLH-034", "彭坚": "ZLH-034",
    "李辉英": "ZLH-035", "李公白": "ZLH-035",
    "穆木天": "ZLH-036", "穆敬熙": "ZLH-036",
    "楼适夷": "ZLH-037", "楼锡春": "ZLH-037", "适夷": "ZLH-037", "楼建南": "ZLH-037",
    "陈荒煤": "ZLH-038", "陈光亮": "ZLH-038",
    "欧阳山": "ZLH-039", "杨凤岐": "ZLH-039",
    "草明": "ZLH-040", "郑晓光": "ZLH-040",
    "骆宾基": "ZLH-041", "张璞君": "ZLH-041",
    "魏金枝": "ZLH-042", "魏金志": "ZLH-042",
    "端木蕻良": "ZLH-043", "曹京平": "ZLH-043", "曹汉文": "ZLH-043",
    "蒋锡金": "ZLH-044", "蒋锡珍": "ZLH-044",
    "周立波": "ZLH-045", "周绍仪": "ZLH-045",
    "邵荃麟": "ZLH-046", "邵全麟": "ZLH-046",
    "林淡秋": "ZLH-047", "林秉钧": "ZLH-047",
    "王任叔": "ZLH-048", "巴人": "ZLH-048",
    "舒群": "ZLH-049", "许春元": "ZLH-049",
    "于黑丁": "ZLH-050", "于黑庭": "ZLH-050",
    "邱东平": "ZLH-051", "邱东萍": "ZLH-051",
    "梅志": "ZLH-052", "屠稚眉": "ZLH-052",
    "黄源": "ZLH-053", "黄河清": "ZLH-053",
    "许幸之": "ZLH-054", "许兴之": "ZLH-054",
    "郑君里": "ZLH-055", "郑君礼": "ZLH-055",
    "白杨": "ZLH-056", "杨成芳": "ZLH-056",
    "赵丹": "ZLH-057", "赵凤翱": "ZLH-057",
    "郭沫若": "ZLH-058", "郭开贞": "ZLH-058", "沫若": "ZLH-058",
    "白薇": "ZLH-059", "黄彰": "ZLH-059",
    "鲁彦": "ZLH-060", "王鲁彦": "ZLH-060",
    "朱镜我": "ZLH-061", "朱德源": "ZLH-061",
    "李一芒": "ZLH-062", "李太白": "ZLH-062",
    "任白戈": "ZLH-063",
    "杜埃": "ZLH-064", "杜秉正": "ZLH-064",
    "柯仲平": "ZLH-065", "柯维翰": "ZLH-065",
    "黄其奎": "ZLH-066",
    "任钧": "ZLH-067", "任君": "ZLH-067",
    "温梓川": "ZLH-068",
    "张庚": "ZLH-069", "张三庚": "ZLH-069",
    "塞克": "ZLH-070", "陈凝秋": "ZLH-070",
    "李霁野": "ZLH-071",
    "丛深": "ZLH-072",
    "王余杞": "ZLH-073",
    "严文井": "ZLH-074", "严文锦": "ZLH-074",
    "黄肃秋": "ZLH-075",
    "马宁": "ZLH-076",
    "柳倩": "ZLH-077", "柳自强": "ZLH-077",
    "王一知": "ZLH-078",
    "谢和赓": "ZLH-079",
    "陈企霞": "ZLH-080",
    "陈白尘": "ZLH-081", "陈炳泉": "ZLH-081",
    "宋之的": "ZLH-082",
    "林焕平": "ZLH-083",
    "陆万美": "ZLH-084",
    "方殷": "ZLH-085",
    "葛一虹": "ZLH-086",
    "于伶": "ZLH-087", "任郁伶": "ZLH-087",
    "徐芳": "ZLH-088",
    "唐弢": "ZLH-089",
    "吴晓邦": "ZLH-090",
    "叶以群": "ZLH-091",
    "蒋天佐": "ZLH-092",
    "王亚平": "ZLH-093",
    "吴亮平": "ZLH-094",
    "何稼夫": "ZLH-095",
    "王统照": "ZLH-096",
    "王西彦": "ZLH-097",
    "沈从文": "ZLH-098",
    "吴奔星": "ZLH-099",
    "姚蓬子": "ZLH-100", "蓬子": "ZLH-100",
    "徐迟": "ZLH-101",
    "张映纶": "ZLH-102",
    "王莹": "ZLH-103", "喻志华": "ZLH-103",
    "叶灵凤": "ZLH-104",
    "沈起予": "ZLH-105",
    "张光年": "ZLH-106", "光未然": "ZLH-106",
    "孙师毅": "ZLH-107",
    "陈鲤庭": "ZLH-108",
    "王林": "ZLH-109",
    "章泯": "ZLH-110",
    "朱静子": "ZLH-111",
    "戴望舒": "ZLH-112",
    "施蛰存": "ZLH-113",
    "刘呐鸥": "ZLH-114",
    "穆时英": "ZLH-115",
    "孙席珍": "ZLH-116",
    "陈波儿": "ZLH-117",
    "吴组缃": "ZLH-118",
    "何家槐": "ZLH-119",
    "夏征农": "ZLH-120",
    "内山完造": "ZLH-121",
    "许广平": "ZLH-122", "景宋": "ZLH-122",
    "邹韬奋": "ZLH-123", "邹恩润": "ZLH-123",
    "李小峰": "ZLH-124", "李致广": "ZLH-124",
    "陆志庠": "ZLH-125", "陆志祥": "ZLH-125",
    "沈兼士": "ZLH-126",
    "曹靖华": "ZLH-127", "曹联亚": "ZLH-127",
    "林语堂": "ZLH-128",
    "胡愈之": "ZLH-129",
    "杨杏佛": "ZLH-130",
    "史沫特莱": "ZLH-131",
    "伊罗生": "ZLH-132",
    "斯诺": "ZLH-133",
    "鹿地亘": "ZLH-134",
    "池田幸子": "ZLH-135",
    "宋庆龄": "ZLH-136",
    "蔡元培": "ZLH-137",
    "何香凝": "ZLH-138",
    "陈望道": "ZLH-139",
    "鲁维查": "ZLH-140",
    "高尔基": "ZLH-141",
    "法捷耶夫": "ZLH-142",
    "罗曼罗兰": "ZLH-143",
    "巴比塞": "ZLH-144",
    "萧伯纳": "ZLH-145",
    "马尔罗": "ZLH-146",
    "辛克莱": "ZLH-147",
    "片山潜": "ZLH-148",
    "艾黎": "ZLH-149",
    "萧三": "ZLH-150", "萧植蕃": "ZLH-150",
    # 额外别名补充（书中常见的OCR变体）
    "钱杏邮": "ZLH-011", "钱杏部": "ZLH-011", "钱杏郭": "ZLH-011", "钱杏邻": "ZLH-011",
    "洪灵菲": "ZLH-010",  # 注意：洪灵菲不等于洪深，他不在150人中，需排除
    "戴平万": "ZLH-048",  # 戴平万也不在列表中，排除
    "成仿吾": "ZLH-014",  # 成仿吾不是冯乃超，需排除
    "林伯修": "ZLH-061",  # 林伯修=杜国庫=朱镜我？不对，需查
}

# 修正：移除不正确的映射
# 洪灵菲、戴平万、成仿吾、林伯修 不在150人列表中，移除其错误映射
for wrong_name in ["洪灵菲", "戴平万", "成仿吾", "林伯修"]:
    if wrong_name in PERSON_MAP:
        del PERSON_MAP[wrong_name]

# 需要排除的过短或易误匹配的别名（单字或常见词）
SHORT_NAMES_TO_EXCLUDE = {
    "L.X.", "丁一", "画室", "巴人", "周文", "草明", "白杨",
    "王林", "马宁", "徐芳", "徐迟", "王莹", "杜埃", "于伶",
    "方殷", "梅志", "舒群", "塞克", "丛深",
}

# ID到主要名字的映射
ID_TO_NAME = {
    "ZLH-001": "鲁迅", "ZLH-002": "茅盾", "ZLH-003": "瞿秋白", "ZLH-004": "夏衍",
    "ZLH-005": "冯雪峰", "ZLH-006": "潘汉年", "ZLH-007": "周扬", "ZLH-008": "田汉",
    "ZLH-009": "郑伯奇", "ZLH-010": "洪深", "ZLH-011": "钱杏邨", "ZLH-012": "蒋光慈",
    "ZLH-013": "阳翰笙", "ZLH-014": "冯乃超", "ZLH-015": "郁达夫", "ZLH-016": "柔石",
    "ZLH-017": "胡也频", "ZLH-018": "冯铿", "ZLH-019": "殷夫", "ZLH-020": "李求实",
    "ZLH-021": "丁玲", "ZLH-022": "张天翼", "ZLH-023": "叶紫", "ZLH-024": "艾芜",
    "ZLH-025": "沙汀", "ZLH-026": "胡风", "ZLH-027": "萧军", "ZLH-028": "萧红",
    "ZLH-029": "聂绀弩", "ZLH-030": "吴奚如", "ZLH-031": "葛琴", "ZLH-032": "徐懋庸",
    "ZLH-033": "周文", "ZLH-034": "彭康", "ZLH-035": "李辉英", "ZLH-036": "穆木天",
    "ZLH-037": "楼适夷", "ZLH-038": "陈荒煤", "ZLH-039": "欧阳山", "ZLH-040": "草明",
    "ZLH-041": "骆宾基", "ZLH-042": "魏金枝", "ZLH-043": "端木蕻良", "ZLH-044": "蒋锡金",
    "ZLH-045": "周立波", "ZLH-046": "邵荃麟", "ZLH-047": "林淡秋", "ZLH-048": "王任叔",
    "ZLH-049": "舒群", "ZLH-050": "于黑丁", "ZLH-051": "邱东平", "ZLH-052": "梅志",
    "ZLH-053": "黄源", "ZLH-054": "许幸之", "ZLH-055": "郑君里", "ZLH-056": "白杨",
    "ZLH-057": "赵丹", "ZLH-058": "郭沫若", "ZLH-059": "白薇", "ZLH-060": "鲁彦",
    "ZLH-061": "朱镜我", "ZLH-062": "李一芒", "ZLH-063": "任白戈", "ZLH-064": "杜埃",
    "ZLH-065": "柯仲平", "ZLH-066": "黄其奎", "ZLH-067": "任钧", "ZLH-068": "温梓川",
    "ZLH-069": "张庚", "ZLH-070": "塞克", "ZLH-071": "李霁野", "ZLH-072": "丛深",
    "ZLH-073": "王余杞", "ZLH-074": "严文井", "ZLH-075": "黄肃秋", "ZLH-076": "马宁",
    "ZLH-077": "柳倩", "ZLH-078": "王一知", "ZLH-079": "谢和赓", "ZLH-080": "陈企霞",
    "ZLH-081": "陈白尘", "ZLH-082": "宋之的", "ZLH-083": "林焕平", "ZLH-084": "陆万美",
    "ZLH-085": "方殷", "ZLH-086": "葛一虹", "ZLH-087": "于伶", "ZLH-088": "徐芳",
    "ZLH-089": "唐弢", "ZLH-090": "吴晓邦", "ZLH-091": "叶以群", "ZLH-092": "蒋天佐",
    "ZLH-093": "王亚平", "ZLH-094": "吴亮平", "ZLH-095": "何稼夫", "ZLH-096": "王统照",
    "ZLH-097": "王西彦", "ZLH-098": "沈从文", "ZLH-099": "吴奔星", "ZLH-100": "姚蓬子",
    "ZLH-101": "徐迟", "ZLH-102": "张映纶", "ZLH-103": "王莹", "ZLH-104": "叶灵凤",
    "ZLH-105": "沈起予", "ZLH-106": "张光年", "ZLH-107": "孙师毅", "ZLH-108": "陈鲤庭",
    "ZLH-109": "王林", "ZLH-110": "章泯", "ZLH-111": "朱静子", "ZLH-112": "戴望舒",
    "ZLH-113": "施蛰存", "ZLH-114": "刘呐鸥", "ZLH-115": "穆时英", "ZLH-116": "孙席珍",
    "ZLH-117": "陈波儿", "ZLH-118": "吴组缃", "ZLH-119": "何家槐", "ZLH-120": "夏征农",
    "ZLH-121": "内山完造", "ZLH-122": "许广平", "ZLH-123": "邹韬奋", "ZLH-124": "李小峰",
    "ZLH-125": "陆志庠", "ZLH-126": "沈兼士", "ZLH-127": "曹靖华", "ZLH-128": "林语堂",
    "ZLH-129": "胡愈之", "ZLH-130": "杨杏佛", "ZLH-131": "史沫特莱", "ZLH-132": "伊罗生",
    "ZLH-133": "斯诺", "ZLH-134": "鹿地亘", "ZLH-135": "池田幸子", "ZLH-136": "宋庆龄",
    "ZLH-137": "蔡元培", "ZLH-138": "何香凝", "ZLH-139": "陈望道", "ZLH-140": "鲁维查",
    "ZLH-141": "高尔基", "ZLH-142": "法捷耶夫", "ZLH-143": "罗曼·罗兰", "ZLH-144": "巴比塞",
    "ZLH-145": "萧伯纳", "ZLH-146": "马尔罗", "ZLH-147": "辛克莱", "ZLH-148": "片山潜",
    "ZLH-149": "路易·艾黎", "ZLH-150": "萧三",
}

# 左联核心成员（用于判断组织关系）
LEFT_LEAGUE_CORE = set([f"ZLH-{i:03d}" for i in range(1, 51)])


def split_pages(text):
    """将OCR文本按页分割，返回 [(page_num, page_text), ...]"""
    pages = []
    # 页分隔符格式: ────────────────────────────────────────
    # 第 X 页
    # ────────────────────────────────────────
    pattern = r'第\s*(\d+)\s*页'
    
    # 找到所有页标记
    page_markers = list(re.finditer(pattern, text))
    
    for i, marker in enumerate(page_markers):
        page_num = int(marker.group(1))
        start = marker.end()
        if i + 1 < len(page_markers):
            end = page_markers[i + 1].start()
            # 回退到分隔线之前
            sep_pos = text.rfind('────', start, end)
            if sep_pos > start:
                end = sep_pos
        else:
            end = len(text)
        
        page_text = text[start:end].strip()
        if page_text:
            pages.append((page_num, page_text))
    
    return pages


def find_persons_in_text(text):
    """在文本中查找所有人物，返回 {entity_id: [matched_name, ...]}"""
    found = defaultdict(set)
    
    for name, entity_id in PERSON_MAP.items():
        # 跳过过短的名字（容易误匹配）
        if len(name) < 2:
            continue
        # 跳过已知的易误匹配名字
        if name in SHORT_NAMES_TO_EXCLUDE and len(name) <= 2:
            continue
        
        # 用空格分隔的OCR文本需要特殊处理
        # 尝试原始匹配
        if name in text:
            found[entity_id].add(name)
            continue
        
        # 尝试字之间有空格的匹配（OCR特征：字之间插入空格）
        if len(name) >= 2:
            spaced_name = r'\s*'.join(re.escape(c) for c in name)
            if re.search(spaced_name, text):
                found[entity_id].add(name)
    
    return found


def extract_context(text, name1, name2, max_len=80):
    """提取两个人名之间的上下文"""
    # 清理OCR空格
    clean_text = re.sub(r'\s+', '', text)
    
    pos1 = clean_text.find(name1)
    pos2 = clean_text.find(name2)
    
    if pos1 == -1 or pos2 == -1:
        return clean_text[:max_len]
    
    start = min(pos1, pos2)
    end = max(pos1 + len(name1), pos2 + len(name2))
    
    ctx_start = max(0, start - 15)
    ctx_end = min(len(clean_text), end + 15)
    context = clean_text[ctx_start:ctx_end]
    
    if len(context) > max_len:
        context = context[:max_len]
    
    return context


def determine_relation(text, source_id, target_id):
    """根据页面文本内容判断关系类型和权重"""
    clean = re.sub(r'\s+', '', text)
    
    # 组织隶属
    org_kw = ["左联", "左翼作家联盟", "执委", "常委", "党团", "盟员", "入盟",
              "加入", "成员", "大会", "成立", "组织", "领导"]
    # 合作
    coop_kw = ["合作", "共同", "编辑", "主编", "创办", "合译", "合著", "同编",
               "一起", "同去", "同往", "共事"]
    # 亲属（用更精确的多字词避免误匹配）
    family_kw = ["妻子", "丈夫", "夫妇", "爱人", "伴侣", "父亲", "母亲",
                 "兄弟", "姐妹", "结婚", "夫妻", "配偶"]
    # 论战
    debate_kw = ["论战", "批评", "批判", "争论", "辩论", "反驳", "论争", "攻击"]
    # 通信
    comm_kw = ["来信", "致信", "通信", "书信", "函"]
    # 会面
    meet_kw = ["会见", "见面", "拜访", "访问", "来访", "会晤", "相见", "造访",
               "看望", "座谈", "宴"]
    
    for kw in family_kw:
        if kw in clean:
            return "强关联-亲属", 5
    
    for kw in org_kw:
        if kw in clean:
            return "强关联-组织隶属", 4
    
    for kw in debate_kw:
        if kw in clean:
            return "弱关联-论战", 3
    
    for kw in coop_kw:
        if kw in clean:
            return "弱关联-合作", 3
    
    for kw in comm_kw:
        if kw in clean:
            return "弱关联-通信", 3
    
    for kw in meet_kw:
        if kw in clean:
            return "弱关联-时空共现", 3
    
    # 默认：如果都是核心成员，认为是组织关系
    if source_id in LEFT_LEAGUE_CORE and target_id in LEFT_LEAGUE_CORE:
        return "强关联-组织隶属", 3
    
    return "弱关联-时空共现", 2


def load_existing_pairs(excel_path):
    """加载Sheet2中已有的Source_ID+Target_ID组合"""
    existing = set()
    max_seq = 0
    try:
        wb = openpyxl.load_workbook(excel_path)
        if 'Sheet2' in wb.sheetnames:
            sheet = wb['Sheet2']
            for row in range(2, sheet.max_row + 1):
                seq = sheet.cell(row=row, column=1).value
                src = sheet.cell(row=row, column=2).value
                tgt = sheet.cell(row=row, column=3).value
                if src and tgt:
                    existing.add((src, tgt))
                    existing.add((tgt, src))  # 双向去重
                if seq and isinstance(seq, (int, float)):
                    max_seq = max(max_seq, int(seq))
        wb.close()
    except Exception as e:
        print(f"加载现有数据出错: {e}")
    return existing, max_seq


def process_text_file(filepath, source_label, existing_pairs):
    """处理一个TXT文件，提取所有页内共现关系"""
    print(f"\n读取 {filepath}...")
    with open(filepath, 'r', encoding='utf-8') as f:
        text = f.read()
    
    pages = split_pages(text)
    print(f"  共分割为 {len(pages)} 页")
    
    # 收集所有关系 key=(src, tgt) sorted pair
    all_rels = {}  # key -> {relation_type, contexts, evidence_refs, weight}
    
    pages_with_persons = 0
    total_pairs = 0
    
    for page_num, page_text in pages:
        persons = find_persons_in_text(page_text)
        
        if len(persons) < 2:
            continue
        
        pages_with_persons += 1
        entity_ids = list(persons.keys())
        
        # 生成所有两两组合
        for id1, id2 in combinations(sorted(entity_ids), 2):
            if id1 == id2:
                continue
            
            key = (id1, id2)
            
            # 获取匹配到的名字
            name1 = list(persons[id1])[0]
            name2 = list(persons[id2])[0]
            
            # 提取上下文
            context = extract_context(page_text, name1, name2)
            evidence = f"{source_label} 第{page_num}页"
            
            # 判断关系类型  
            rel_type, weight = determine_relation(page_text, id1, id2)
            
            if key in all_rels:
                rel = all_rels[key]
                # 合并：追加证据，取最高权重和最强关系
                if evidence not in rel['evidence_refs']:
                    rel['evidence_refs'].append(evidence)
                if len(rel['contexts']) < 3:  # 最多保留3条上下文
                    rel['contexts'].append(context)
                if weight > rel['weight']:
                    rel['weight'] = weight
                    rel['relation_type'] = rel_type
            else:
                all_rels[key] = {
                    'relation_type': rel_type,
                    'contexts': [context],
                    'evidence_refs': [evidence],
                    'weight': weight
                }
                total_pairs += 1
    
    print(f"  找到 {pages_with_persons} 页包含2+个实体")
    print(f"  提取到 {total_pairs} 对关系")
    
    # 过滤掉已存在的关系
    new_rels = {}
    for (src, tgt), data in all_rels.items():
        if (src, tgt) not in existing_pairs and (tgt, src) not in existing_pairs:
            new_rels[(src, tgt)] = data
    
    print(f"  去除已有后剩余 {len(new_rels)} 对新关系")
    return new_rels


def main():
    excel_path = r'd:\1大创\大创数据收集1.xlsx'
    cidian_path = r'd:\1大创\左联词典.txt'
    shi_path = r'd:\1大创\左联史.txt'
    
    print("=" * 60)
    print("从《左联词典》和《左联史》提取人际关系数据")
    print("=" * 60)
    
    # 1. 加载已有数据
    print("\n[1/4] 加载Sheet2现有数据...")
    existing_pairs, max_seq = load_existing_pairs(excel_path)
    print(f"  现有关系对: {len(existing_pairs) // 2} 条")
    print(f"  当前最大序号: {max_seq}")
    
    # 2. 处理左联词典
    print("\n[2/4] 处理《左联词典》...")
    cidian_rels = process_text_file(cidian_path, "左联词典", existing_pairs)
    
    # 更新existing_pairs以避免词典和史的重复
    for key in cidian_rels:
        existing_pairs.add(key)
        existing_pairs.add((key[1], key[0]))
    
    # 3. 处理左联史
    print("\n[3/4] 处理《左联史》...")
    shi_rels = process_text_file(shi_path, "左联史", existing_pairs)
    
    # 合并两本书的结果
    all_new_rels = {}
    all_new_rels.update(cidian_rels)
    
    for key, data in shi_rels.items():
        if key in all_new_rels:
            # 合并
            old = all_new_rels[key]
            old['evidence_refs'].extend(data['evidence_refs'])
            old['contexts'].extend(data['contexts'][:2])
            if data['weight'] > old['weight']:
                old['weight'] = data['weight']
                old['relation_type'] = data['relation_type']
        else:
            all_new_rels[key] = data
    
    print(f"\n合计新关系: {len(all_new_rels)} 条")
    
    if not all_new_rels:
        print("没有新关系可以添加。")
        return
    
    # 4. 写入Excel
    print("\n[4/4] 写入Excel Sheet2...")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['Sheet2']
    
    # 找到已有数据的末尾行
    last_row = sheet.max_row
    
    seq = max_seq
    added = 0
    
    for (src, tgt), data in sorted(all_new_rels.items()):
        seq += 1
        row = last_row + 1 + added
        
        # 格式化上下文和证据
        context_str = "; ".join(data['contexts'][:3])
        if len(context_str) > 200:
            context_str = context_str[:200]
        
        evidence_str = "; ".join(data['evidence_refs'][:5])
        if len(evidence_str) > 200:
            evidence_str = evidence_str[:200]
        
        sheet.cell(row=row, column=1, value=seq)
        sheet.cell(row=row, column=2, value=src)
        sheet.cell(row=row, column=3, value=tgt)
        sheet.cell(row=row, column=4, value=data['relation_type'])
        sheet.cell(row=row, column=5, value=context_str)
        sheet.cell(row=row, column=6, value=evidence_str)
        sheet.cell(row=row, column=7, value=data['weight'])
        
        added += 1
    
    wb.save(excel_path)
    print(f"\n✓ 成功追加 {added} 条新关系到 Sheet2")
    print(f"  序号范围: {max_seq + 1} ~ {seq}")
    print(f"  保存到: {excel_path}")
    
    # 统计
    print("\n" + "=" * 60)
    print("关系类型分布:")
    type_count = defaultdict(int)
    for data in all_new_rels.values():
        type_count[data['relation_type']] += 1
    for rt, cnt in sorted(type_count.items(), key=lambda x: -x[1]):
        print(f"  {rt}: {cnt}")
    
    print("\n涉及实体频次（前20）:")
    entity_count = defaultdict(int)
    for (src, tgt) in all_new_rels:
        entity_count[src] += 1
        entity_count[tgt] += 1
    for eid, cnt in sorted(entity_count.items(), key=lambda x: -x[1])[:20]:
        name = ID_TO_NAME.get(eid, eid)
        print(f"  {name} ({eid}): {cnt} 次")
    
    # 检查原先孤立的实体是否被覆盖
    print("\n新覆盖的实体（之前在Sheet2中无记录）:")
    original_entities = set()
    for (s, t) in existing_pairs:
        original_entities.add(s)
        original_entities.add(t)
    
    new_entities = set()
    for (s, t) in all_new_rels:
        new_entities.add(s)
        new_entities.add(t)
    
    newly_covered = new_entities - original_entities
    for eid in sorted(newly_covered):
        name = ID_TO_NAME.get(eid, eid)
        print(f"  {name} ({eid})")


if __name__ == '__main__':
    main()
