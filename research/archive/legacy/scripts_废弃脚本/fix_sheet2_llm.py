  """
使用 GPT-4 对高风险条目（亲属、论战、师生）进行定向复核与修正
"""
import pandas as pd
import json
import time
import re
from openai import OpenAI

# ══════════════════════════════════════════════════════════
# 配置
# ══════════════════════════════════════════════════════════
client = OpenAI(
    api_key="sk-fe66vznYArBkw4gFrEkM5vPkXvQgzdDVFInQ5OrAwfIagZDH",
    base_url="https://yunwu.ai/v1",
)
MODEL = "gpt-4o"

INPUT = '数据/输出结果/fixed_rules_zuolian.xlsx'
OUTPUT = '数据/输出结果/final_fixed_zuolian.xlsx'
BATCH_SIZE = 8

# ══════════════════════════════════════════════════════════
# 修正指南
# ══════════════════════════════════════════════════════════
SYSTEM_PROMPT = """你是一位中国现代文学史专家，正在审核“左联”成员关系数据库。
请核查以下[亲属关系]、[文学论战]或[师生关系]条目是否准确。

判断标准：
1. 亲属关系：必须有明确血缘或婚姻证据。同辈作家的“笔头兄弟”应判为“交游”或“创作合作”。
2. 文学论战：必须有明确的观点冲突、批判、驳斥。单纯的“讨论”应判为“交游”。
3. 师生关系：必须有明确的授课、指导或师承记录。

输出格式（JSON数组）：
[{"序号": 序号, "verdict": "correct" 或 "incorrect", "corrected_type": "修正后的类型 (若正确则为null)", "reason": "15字以内说明"}]"""

def call_llm(prompt):
    try:
        resp = client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": prompt}
            ],
            temperature=0,
            response_format={"type": "json_object"} if "gpt-4" in MODEL else None
        )
        content = resp.choices[0].message.content
        # 尝试剥离 Markdown 包装
        content = re.sub(r'```json\s*|\s*```', '', content).strip()
        
        data = json.loads(content)
        # 提取数组
        results = []
        if isinstance(data, list):
            results = data
        elif isinstance(data, dict):
            # 如果是 {"results": [...]}
            for k, v in data.items():
                if isinstance(v, list):
                    results = v
                    break
            if not results: results = [data] # 单个对象

        # 过滤非字典项
        return [r for r in results if isinstance(r, dict)]
    except Exception as e:
        print(f"API Error: {e}")
        return None

# ══════════════════════════════════════════════════════════
# 执行逻辑
# ══════════════════════════════════════════════════════════
print("加载数据...")
df1 = pd.read_excel(INPUT, sheet_name='Sheet1')
df2 = pd.read_excel(INPUT, sheet_name='Sheet2')
id_to_name = dict(zip(df1['Entity_ID'], df1['True_Name']))

# 筛选高风险条目
high_risk = df2[df2['Relation_Type'].isin(['亲属关系', '文学论战', '师生关系'])].copy()
print(f"待复核条目总数: {len(high_risk)}")

corrections = {}

for i in range(0, len(high_risk), BATCH_SIZE):
    batch = high_risk.iloc[i:i+BATCH_SIZE]
    print(f"处理批次 {i//BATCH_SIZE + 1} ({i+1}-{min(i+BATCH_SIZE, len(high_risk))})...")
    
    items_text = []
    for _, row in batch.iterrows():
        s = id_to_name.get(row['Source_ID'], row['Source_ID'])
        t = id_to_name.get(row['Target_ID'], row['Target_ID'])
        items_text.append(f"序号{row['序号']}: {s} 与 {t} 为 [{row['Relation_Type']}]。上下文: {row['Context']}")
    
    prompt = "请核查以下条目：\n" + "\n".join(items_text)
    
    results = call_llm(prompt)
    if results:
        for res in results:
            if res.get('verdict') == 'incorrect' and res.get('corrected_type'):
                corrections[res['序号']] = res['corrected_type']
                print(f"  序号{res['序号']} -> {res['corrected_type']} ({res.get('reason')})")
    
    time.sleep(1) # 频率控制

# ══════════════════════════════════════════════════════════
# 应用修正
# ══════════════════════════════════════════════════════════
print(f"\n共应用 {len(corrections)} 条 LLM 修正。")
def apply_llm_fix(row):
    return corrections.get(row['序号'], row['Relation_Type'])

df2['Relation_Type'] = df2.apply(apply_llm_fix, axis=1)

# 保存
import shutil
shutil.copy2(INPUT, OUTPUT)
from openpyxl import load_workbook
wb = load_workbook(OUTPUT)
if 'Sheet2' in wb.sheetnames: del wb['Sheet2']
wb.save(OUTPUT)

with pd.ExcelWriter(OUTPUT, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df2.to_excel(writer, sheet_name='Sheet2', index=False)

print(f"已保存至 {OUTPUT}")
