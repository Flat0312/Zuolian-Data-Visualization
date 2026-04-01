import pandas as pd
import re

# ── 1. 加载数据 ──
INPUT = '数据/输出结果/《左联相关档案资源目录》.xlsx'
OUTPUT = '数据/输出结果/modified_zuolian.xlsx'

df1 = pd.read_excel(INPUT, sheet_name='Sheet1')
df = pd.read_excel(INPUT, sheet_name='Sheet2')

valid_ids = set(df1['Entity_ID'])
print(f"原始行数: {len(df)}")

# ── 2. 删除无效 ID 行 ──
mask_valid = df['Source_ID'].isin(valid_ids) & df['Target_ID'].isin(valid_ids)
removed_invalid = (~mask_valid).sum()
df = df[mask_valid].copy()
print(f"删除无效ID行: {removed_invalid}")

# ── 3. 基于 Context 细化 Relation_Type ──
# 规则优先级：从高到低，后面的不会覆盖前面已匹配的
# 只对当前为"组织隶属"或"交游"的行进行细化
RECLASSIFY_RULES = [
    # (新类型, 关键词列表, 适用的原类型)
    ('亲属关系', ['妻子', '丈夫', '夫人', '结婚', '夫妇', '妻', '兄弟', '姐妹', '父亲', '母亲', '嫁给', '婚后', '携妻', '携夫', '遗孀'], None),
    ('师生关系', ['老师', '学生', '师从', '受教', '教导', '指导', '从学', '受业', '拜师', '门生'], None),
    ('文学论战', ['论战', '辩论', '争论', '论争', '批判', '驳斥', '反驳', '笔战', '攻击', '论敌', '批评.*回应', '回击'], None),
    ('地下通讯', ['秘密', '地下', '密信', '暗号', '接头', '掩护', '转移', '潜伏', '伪装身份', '秘密联络', '秘密通信'], None),
    ('创作合作', ['合编', '合著', '联名签署', '共同编辑', '同编', '合作创作', '合作翻译', '共同翻译'], ['组织隶属', '交游']),
    ('签名联署', ['签名', '联名', '宣言', '声明', '抗议.*签', '联署', '通电'], ['组织隶属']),
    ('空间共现', ['同行', '同住', '同居一室', '聚会', '宴会', '聚餐', '同赴', '一起.*到', '一起.*去', '同往', '共赴'], ['组织隶属', '交游']),
]

def reclassify(row):
    ctx = str(row['Context']) if pd.notna(row['Context']) else ''
    orig_type = row['Relation_Type']
    for new_type, keywords, apply_to in RECLASSIFY_RULES:
        if apply_to and orig_type not in apply_to:
            continue
        for kw in keywords:
            if re.search(kw, ctx):
                return new_type
    return orig_type

df['Relation_Type'] = df.apply(reclassify, axis=1)

# ── 4. 更新 Weight ──
# 策略: 基于 Evidence_Ref 的来源数量 + Context 丰富度
def compute_weight(row):
    ref = str(row['Evidence_Ref']) if pd.notna(row['Evidence_Ref']) else ''
    ctx = str(row['Context']) if pd.notna(row['Context']) else ''

    # 基础分: 来源数量（;分隔的数量）
    ref_count = len([r for r in ref.split(';') if r.strip()])

    # Context 长度加分
    ctx_len = len(ctx)

    if ref_count >= 4:
        w = 5
    elif ref_count >= 3:
        w = 4
    elif ref_count >= 2:
        w = 3
    elif ctx_len > 100:
        w = 2
    else:
        w = 1

    # 特殊关系类型加分（亲属/论战天然证据强）
    if row['Relation_Type'] in ('亲属关系', '文学论战', '师生关系'):
        w = max(w, 2)

    return min(w, 5)

df['Weight'] = df.apply(compute_weight, axis=1)

# ── 5. 清理 Evidence_Ref ──
def clean_evidence_ref(ref):
    if pd.isna(ref) or not str(ref).strip():
        return ref
    ref = str(ref)
    # 按分号拆分，去重，去空白
    parts = [p.strip() for p in ref.split(';') if p.strip()]
    # 去重（保持顺序）
    seen = []
    for p in parts:
        if p not in seen:
            seen.append(p)
    return '; '.join(seen)

df['Evidence_Ref'] = df['Evidence_Ref'].apply(clean_evidence_ref)

# ── 6. 去重 ──
# 同 Source_ID + Target_ID + Relation_Type 的重复行，保留 Evidence_Ref 最丰富的那条
before_dedup = len(df)

def merge_duplicates(group):
    if len(group) == 1:
        return group
    # 选 Evidence_Ref 最长（来源最多）的行
    best_idx = group['Evidence_Ref'].fillna('').str.len().idxmax()
    best = group.loc[[best_idx]].copy()
    # 合并所有不重复的 Evidence_Ref
    all_refs = set()
    for ref in group['Evidence_Ref'].dropna():
        for p in str(ref).split(';'):
            if p.strip():
                all_refs.add(p.strip())
    best['Evidence_Ref'] = '; '.join(sorted(all_refs))
    # 合并 Context: 取最长的
    best_ctx_idx = group['Context'].fillna('').str.len().idxmax()
    best['Context'] = group.loc[best_ctx_idx, 'Context']
    # Weight 取最大
    best['Weight'] = group['Weight'].max()
    return best

df = df.groupby(['Source_ID', 'Target_ID', 'Relation_Type'], group_keys=False).apply(merge_duplicates)
df = df.reset_index(drop=True)
print(f"去重: {before_dedup} -> {len(df)}")

# ── 7. 重新编号 ──
df['序号'] = range(1, len(df) + 1)

# ── 8. 统计结果 ──
print(f"\n=== 处理后统计 ===")
print(f"总行数: {len(df)}")
print(f"\nRelation_Type 分布:")
print(df['Relation_Type'].value_counts().to_string())
print(f"\nWeight 分布:")
print(df['Weight'].value_counts().sort_index().to_string())

# ── 9. 保存 ──
from openpyxl import load_workbook

# 先复制原文件
import shutil
shutil.copy2(INPUT, OUTPUT)

# 用 openpyxl 打开并替换 Sheet2
wb = load_workbook(OUTPUT)
if 'Sheet2' in wb.sheetnames:
    del wb['Sheet2']
wb.save(OUTPUT)

with pd.ExcelWriter(OUTPUT, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name='Sheet2', index=False)

print(f"\n✓ 已保存至: {OUTPUT}")
