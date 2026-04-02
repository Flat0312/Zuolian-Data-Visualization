from __future__ import annotations

import csv
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


TARGET_BASENAME = "《左联相关档案资源目录》.xlsx"
OUTPUT_BASENAME = "《左联相关档案资源目录》_修正版.xlsx"
LOG_BASENAME = "《左联相关档案资源目录》_修改日志.xlsx"
REVIEW_BASENAME = "review_needed.csv"

SHEET_ALIASES = {
    "sheet1": ["sheet1", "sheet 1", "节点", "人物", "entity", "entities"],
    "sheet2": ["sheet2", "sheet 2", "关系", "relation", "relations", "边"],
    "sheet3": ["sheet3", "sheet 3", "事件", "event", "events"],
}

SHEET1_COLS = {
    "entity_id": ["entity_id", "id", "实体id", "人物id", "entity id"],
    "true_name": ["true_name", "label", "name", "true name", "真实姓名", "姓名"],
    "birth_death": ["birth_death", "birth-death", "birth death", "生卒", "生卒年"],
    "role": ["role", "角色"],
}

SHEET2_COLS = {
    "seq": ["序号", "id", "relation_id"],
    "source_id": ["source_id", "source", "from_id"],
    "target_id": ["target_id", "target", "to_id"],
    "relation_type": ["relation_type", "relation", "关系类型"],
    "context": ["context", "证据", "说明"],
    "evidence_ref": ["evidence_ref", "evidence", "来源", "出处"],
    "weight": ["weight", "score", "权重"],
}

SHEET3_COLS = {
    "seq": ["序号", "id", "event_id"],
    "entity_id": ["entity_id", "entity", "person_id"],
    "timestamp": ["timestamp", "date", "time", "日期"],
    "hist_loc": ["hist_loc", "historical_location", "historical location", "历史地点"],
    "current_loc": ["current_loc", "current_location", "current location", "今址", "现址"],
    "coord_xy": ["coord_xy", "coord", "coordinates", "坐标"],
    "event": ["event", "event_name", "事件", "事件名称"],
}

ORG_KEYWORDS = ["左联", "联盟", "成员", "执行委", "执行委员会", "常委", "书记", "秘书处", "发起人", "同盟", "社"]
VISIT_KEYWORDS = ["来访", "会见", "会面", "拜访", "同往", "看屋", "看房", "同赴"]
COMM_KEYWORDS = ["来信", "收信", "寄信", "通信", "寄函", "收函"]
COOP_KEYWORDS = ["联名", "签署", "共同", "合编", "合作", "编辑", "创办", "发表"]
DEBATE_KEYWORDS = ["论战", "论争", "批判", "批评", "驳", "争论"]
MEMORIAL_KEYWORDS = ["悼", "纪念", "追悼", "悼念", "缅怀"]
KINSHIP_KEYWORDS = ["妻", "夫人", "丈夫", "母", "父", "子", "女", "兄", "弟", "姐", "妹", "伴侣"]
TEACH_KEYWORDS = ["老师", "先生", "学生", "授业", "师从", "弟子", "讲课", "指导"]
ACTIVE_EVENT_KEYWORDS = ["文学活动", "上海文学活动", "被捕事件", "被捕", "秘密会议", "成立大会", "活动", "交流会", "会面"]

VERIFIED_EVENT_RULES: Dict[str, Dict[str, Any]] = {
    "左联成立大会": {
        "standard_event_name": "中国左翼作家联盟成立大会",
        "corrected_date": "1930-03-02",
        "date_precision": "日",
        "historical_location": "中华艺术大学教室（原窦乐安路233弄）",
        "current_address": "上海市虹口区多伦路201弄2号",
        "corrected_persons": "",
        "correction_reason": "公开资料表明左联成立大会于1930年3月2日在中华艺术大学借用教室召开，今址为左联会址纪念馆；原表存在公啡咖啡馆、前哨、复旦大学等冲突地点。",
        "source_url": "https://www.shhk.gov.cn/xwzx/002008/002008040/20221031/bd8cb3ee-198a-431a-adf7-781e9fc5185d.html ; https://www.shhk.gov.cn/xwzx/002003/20250303/ec139c7d-8fa3-4970-a5dd-3248468989c8.html",
        "confidence": "high",
        "manual_review_default": False,
    },
    "五烈士遇难": {
        "standard_event_name": "左联五烈士遇难",
        "corrected_date": "1931-02-07",
        "date_precision": "日",
        "historical_location": "上海龙华淞沪警备司令部刑场",
        "current_address": "上海龙华烈士陵园（门牌待复核）",
        "corrected_persons": "柔石；胡也频；李伟森；冯铿；殷夫",
        "correction_reason": "官方公开资料明确左联五烈士于1931年2月7日在上海龙华遇难；原表将日期统一写成1931-01-01，且今址门牌存在多个冲突版本。",
        "source_url": "https://www.shhk.gov.cn/slh/038001/20260302/5f983b4c-74f7-4a6a-a2b4-930dedf99970.html ; https://www.shhk.gov.cn/xwzx/002006/20210722/07f62353-471f-40d9-880e-ea82be5da936.html",
        "confidence": "high",
        "manual_review_default": False,
        "participant_names": {"柔石", "胡也频", "李伟森", "冯铿", "殷夫"},
    },
    "冯铿被捕事件": {
        "standard_event_name": "冯铿被捕事件",
        "corrected_date": "1931-01-17",
        "date_precision": "日",
        "historical_location": "上海东方旅社",
        "current_address": "上海市黄浦区汉口路（东方旅社旧址门牌待复核）",
        "corrected_persons": "冯铿",
        "correction_reason": "上海党史网资料记载，冯铿于1931年1月17日下午在公共租界三马路（今汉口路）东方旅社举行秘密会议时被捕；原表大量以1月1日占位且地点冲突，现优先保留公开史料可核的具体日期。",
        "source_url": "https://www.ccphistory.org.cn/shds/ssjc/content/cd2b9f77-aa37-488f-beb2-4853b6290e42.html",
        "confidence": "high",
        "manual_review_default": False,
        "participant_names": {"冯铿"},
    },
    "殷夫被捕事件": {
        "standard_event_name": "殷夫被捕事件",
        "corrected_date": "1931-01-17",
        "date_precision": "日",
        "historical_location": "上海东方旅社31号房间",
        "current_address": "上海市黄浦区汉口路（东方旅社旧址门牌待复核）",
        "corrected_persons": "殷夫",
        "correction_reason": "公开资料记载，殷夫于1931年1月17日13时40分在东方旅社31号房间参加党的会议时被捕；原表将日期写成1月1日占位，现按公开史料修正到日。",
        "source_url": "https://df.bjsjw.gov.cn/n299/20210409/i41861.html",
        "confidence": "high",
        "manual_review_default": False,
        "participant_names": {"殷夫"},
    },
    "李求实被捕事件": {
        "standard_event_name": "李求实被捕事件",
        "canonical_event_key": "李求实被捕事件|ZLH-020",
        "corrected_date": "1931-01-17",
        "date_precision": "日",
        "historical_location": "上海东方旅社、中山旅社秘密会议",
        "current_address": "上海市黄浦区汉口路、天津路一带",
        "corrected_persons": "李求实",
        "correction_reason": "中国军网资料记载，1931年1月17日李求实（李伟森）与何孟雄、林育南等在上海东方旅社和中山旅社秘密开会，后因叛徒告密被捕；原表把日期统一写成1月1日，现先按可公开核实的日期修正。",
        "source_url": "https://www.81.cn/yljnt/2013-11/04/content_5630422.htm",
        "confidence": "medium",
        "manual_review_default": True,
        "participant_names": {"李求实"},
    },
    "丁玲被捕事件": {
        "standard_event_name": "丁玲被捕事件",
        "corrected_date": "1933-05-14",
        "date_precision": "日",
        "historical_location": "虹口昆山花园路7号寓所",
        "current_address": "上海市虹口区昆山路（门牌待复核）",
        "corrected_persons": "丁玲",
        "correction_reason": "政协相关报道指出，丁玲1933年5月14日在昆山花园路7号寓所被绑架；原表将日期写成1月1日占位，现按公开可核史料改为具体日。",
        "source_url": "https://www.icppcc.cn/newsDetail_1059710 ; https://www.ccphistory.org.cn/shds/zxdt100zn/content/115347fd-09db-4528-b823-2d056c333340.html",
        "confidence": "high",
        "manual_review_default": False,
        "participant_names": {"丁玲"},
    },
    "龙华二十四烈士遇难": {
        "standard_event_name": "龙华二十四烈士遇难",
        "corrected_date": "1931-02-07",
        "date_precision": "日",
        "historical_location": "上海龙华淞沪警备司令部刑场",
        "current_address": "上海龙华烈士陵园（门牌待复核）",
        "corrected_persons": "",
        "correction_reason": "中国共产党新闻网资料表明龙华二十四烈士于1931年2月7日被秘密集体枪杀；原表日期为占位值1931-01-01，且部分关联人物并非可直接确认的二十四烈士成员。",
        "source_url": "https://cpc.people.com.cn/n1/2022/1209/c443712-32583679.html",
        "confidence": "medium",
        "manual_review_default": True,
    },
    "内山书店秘密会议": {
        "standard_event_name": "内山书店秘密会议",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "内山书店旧址",
        "current_address": "四川北路2050号",
        "corrected_persons": "",
        "correction_reason": "公开资料可确认内山书店旧址位于四川北路2050号；原表在2040、2044-2048、2050等多个地址之间冲突，但暂未检得同等级公开来源可精确确认该次“秘密会议”的具体日期与参与人。",
        "source_url": "https://www.shhk.gov.cn/xwzx/002006/20210930/96ecb0ec-79e3-49ef-a097-a89c5a5dbc40.html ; https://www.shhk.gov.cn/xwzx/002008/002008040/20240425/3e9546e4-0e0f-409d-a0e2-91bc115f8f66.html",
        "confidence": "medium",
        "manual_review_default": True,
    },
    "鲁迅与柔石会面": {
        "standard_event_name": "鲁迅与柔石会面",
        "corrected_date": None,
        "date_precision": None,
        "historical_location": "北四川路一带",
        "current_address": "",
        "corrected_persons": "鲁迅；柔石",
        "correction_reason": "公开转录材料可支持1930年3月28日与3月31日相关看屋活动；1929年8月29日与1930年3月30日条目仍需回查《鲁迅日记》原文，不宜强行改写。",
        "source_url": "https://al3tai.nenzhu.com/news-id-2373.html",
        "confidence": "medium",
        "manual_review_default": True,
    },
}

ROW_EVENT_OVERRIDES: Dict[int, Dict[str, Any]] = {
    7: {
        "corrected_date": "",
        "date_precision": "",
        "correction_reason": "公开可检索转录材料更明确出现的是1930-03-31同柔石看屋，当前1930-03-30记录未找到同等级公开佐证，保留待人工复核。",
        "source_url": "https://al3tai.nenzhu.com/news-id-2373.html",
        "confidence": "low",
        "needs_manual_review": "yes",
    },
    8: {
        "corrected_date": "",
        "date_precision": "",
        "correction_reason": "公开可检索转录材料更明确出现的是1930-03-31同柔石看屋，当前1930-03-30记录未找到同等级公开佐证，保留待人工复核。",
        "source_url": "https://al3tai.nenzhu.com/news-id-2373.html",
        "confidence": "low",
        "needs_manual_review": "yes",
    },
    42: {
        "standard_event_name": "瞿秋白到上海大学任教",
        "canonical_event_key": "ZLH-003|瞿秋白到上海大学任教|1923-07",
        "corrected_date": "1923-07",
        "date_precision": "月",
        "historical_location": "上海大学",
        "current_address": "上海（具体校址待核）",
        "event_scope": "entity",
        "correction_reason": "党史资料记载瞿秋白于1923年7月到上海大学任教务长兼社会学系主任，现将该条泛化活动记录改写为可唯一识别的任教事件。",
        "display_note": "1923年7月，瞿秋白到上海大学任教并主持社会学系工作，当前条目按这一可公开核实的在沪活动节点展示。",
        "source_url": "https://cpc.people.com.cn/BIG5/n1/2024/1006/c443712-40333498.html",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    43: {
        "standard_event_name": "丁玲抵沪就读平民女校",
        "canonical_event_key": "ZLH-021|丁玲抵沪就读平民女校|1922-02",
        "corrected_date": "1922-02",
        "date_precision": "月",
        "historical_location": "平民女校",
        "current_address": "上海（校址待核）",
        "event_scope": "entity",
        "correction_reason": "澎湃文章《成为“丁玲”之前，和上海的三次际会》记载丁玲于1922年2月到上海平民女校求学，现将该条泛化活动记录改写为抵沪入学事件。",
        "display_note": "1922年2月，丁玲抵沪进入平民女子学校求学，当前条目按这一可核实的上海起点事件展示。",
        "source_url": "https://m.thepaper.cn/newsDetail_forward_28996791",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    60: {
        "standard_event_name": "瞿秋白长汀被俘",
        "canonical_event_key": "ZLH-003|瞿秋白长汀被俘|1935-02-24",
        "corrected_date": "1935-02-24",
        "date_precision": "日",
        "historical_location": "福建长汀县水口镇小迳村附近",
        "current_address": "福建省龙岩市长汀县水口镇小迳村",
        "event_scope": "entity",
        "correction_reason": "上海虹口资料与福建党史资料均指出，瞿秋白于1935年2月24日在福建长汀水口镇小迳村附近被俘；原表把该条误写成1934年上海内山书店被捕事件，现改写为可核实的被俘事件。",
        "display_note": "1935年2月24日，瞿秋白在福建长汀县水口镇小迳村附近突围时被俘，当前条目按公开可核的被俘时间与地点展示。",
        "source_url": "https://www.shhk.gov.cn/xwzx/002009/002009002/20100105/f83bfbd8-323c-4316-bcea-d15f2082a896.html ; https://www.fjdsfzw.org.cn/2023-12-19/content_133925.html",
        "confidence": "high",
        "needs_manual_review": "no",
    },
    192: {
        "standard_event_name": "冯铿元宵抵沪",
        "canonical_event_key": "ZLH-018|冯铿元宵抵沪|1929-02-24",
        "corrected_date": "1929-02-24",
        "date_precision": "日",
        "historical_location": "上海",
        "current_address": "上海",
        "event_scope": "entity",
        "correction_reason": "上海党史网记载冯铿于1929年2月24日元宵节抵达上海，现将该条泛化社交活动记录改写为抵沪事件。",
        "display_note": "1929年2月24日，冯铿抵达上海，当前条目按这一明确可核的来沪时间展示。",
        "source_url": "https://www.ccphistory.org.cn/shds/dsxx/content/cd2b9f77-aa37-488f-beb2-4853b6290e42.html",
        "confidence": "high",
        "needs_manual_review": "no",
    },
    194: {
        "standard_event_name": "楼适夷因支持《文学》月刊被捕",
        "canonical_event_key": "ZLH-037|楼适夷因支持《文学》月刊被捕|1933-09-17",
        "corrected_date": "1933-09-17",
        "date_precision": "日",
        "historical_location": "上海",
        "current_address": "上海（具体抓捕地点待核）",
        "event_scope": "entity",
        "correction_reason": "中国作家网资料记载楼适夷于1933年9月17日因支持《文学》月刊被捕；原表1923/1931日期互相冲突，现将其改写为单一可识别事件。",
        "display_note": "1933年9月17日，楼适夷在上海因参与和支持《文学》相关工作被捕，当前条目按公开可核日期展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0813/c404063-31820381.html",
        "confidence": "high",
        "needs_manual_review": "yes",
    },
    196: {
        "standard_event_name": "楼适夷因支持《文学》月刊被捕",
        "canonical_event_key": "ZLH-037|楼适夷因支持《文学》月刊被捕|1933-09-17",
        "corrected_date": "1933-09-17",
        "date_precision": "日",
        "historical_location": "上海",
        "current_address": "上海（具体抓捕地点待核）",
        "event_scope": "entity",
        "correction_reason": "中国作家网资料记载楼适夷于1933年9月17日因支持《文学》月刊被捕；原表1923/1931日期互相冲突，现将其改写为单一可识别事件。",
        "display_note": "1933年9月17日，楼适夷在上海因参与和支持《文学》相关工作被捕，当前条目按公开可核日期展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0813/c404063-31820381.html",
        "confidence": "high",
        "needs_manual_review": "yes",
    },
    198: {
        "standard_event_name": "楼适夷因支持《文学》月刊被捕",
        "canonical_event_key": "ZLH-037|楼适夷因支持《文学》月刊被捕|1933-09-17",
        "corrected_date": "1933-09-17",
        "date_precision": "日",
        "historical_location": "上海",
        "current_address": "上海（具体抓捕地点待核）",
        "event_scope": "entity",
        "correction_reason": "中国作家网资料记载楼适夷于1933年9月17日因支持《文学》月刊被捕；原表1923/1931日期互相冲突，现将其改写为单一可识别事件。",
        "display_note": "1933年9月17日，楼适夷在上海因参与和支持《文学》相关工作被捕，当前条目按公开可核日期展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0813/c404063-31820381.html",
        "confidence": "high",
        "needs_manual_review": "yes",
    },
    200: {
        "standard_event_name": "楼适夷因支持《文学》月刊被捕",
        "canonical_event_key": "ZLH-037|楼适夷因支持《文学》月刊被捕|1933-09-17",
        "corrected_date": "1933-09-17",
        "date_precision": "日",
        "historical_location": "上海",
        "current_address": "上海（具体抓捕地点待核）",
        "event_scope": "entity",
        "correction_reason": "中国作家网资料记载楼适夷于1933年9月17日因支持《文学》月刊被捕；原表1923/1931日期互相冲突，现将其改写为单一可识别事件。",
        "display_note": "1933年9月17日，楼适夷在上海因参与和支持《文学》相关工作被捕，当前条目按公开可核日期展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0813/c404063-31820381.html",
        "confidence": "high",
        "needs_manual_review": "yes",
    },
    202: {
        "standard_event_name": "楼适夷因支持《文学》月刊被捕",
        "canonical_event_key": "ZLH-037|楼适夷因支持《文学》月刊被捕|1933-09-17",
        "corrected_date": "1933-09-17",
        "date_precision": "日",
        "historical_location": "上海",
        "current_address": "上海（具体抓捕地点待核）",
        "event_scope": "entity",
        "correction_reason": "中国作家网资料记载楼适夷于1933年9月17日因支持《文学》月刊被捕；原表1923/1931日期互相冲突，现将其改写为单一可识别事件。",
        "display_note": "1933年9月17日，楼适夷在上海因参与和支持《文学》相关工作被捕，当前条目按公开可核日期展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0813/c404063-31820381.html",
        "confidence": "high",
        "needs_manual_review": "yes",
    },
    204: {
        "standard_event_name": "楼适夷因支持《文学》月刊被捕",
        "canonical_event_key": "ZLH-037|楼适夷因支持《文学》月刊被捕|1933-09-17",
        "corrected_date": "1933-09-17",
        "date_precision": "日",
        "historical_location": "上海",
        "current_address": "上海（具体抓捕地点待核）",
        "event_scope": "entity",
        "correction_reason": "中国作家网资料记载楼适夷于1933年9月17日因支持《文学》月刊被捕；原表1923/1931日期互相冲突，现将其改写为单一可识别事件。",
        "display_note": "1933年9月17日，楼适夷在上海因参与和支持《文学》相关工作被捕，当前条目按公开可核日期展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0813/c404063-31820381.html",
        "confidence": "high",
        "needs_manual_review": "yes",
    },
    243: {
        "standard_event_name": "丁玲进入上海大学中国文学系",
        "canonical_event_key": "ZLH-021|丁玲进入上海大学中国文学系|1923-08",
        "corrected_date": "1923-08",
        "date_precision": "月",
        "historical_location": "上海大学中国文学系",
        "current_address": "上海（具体校址待核）",
        "event_scope": "entity",
        "correction_reason": "澎湃与上海大学校史资料均记载丁玲于1923年8月进入上海大学中国文学系，现将该条泛化文学活动记录改写为入学事件。",
        "display_note": "1923年8月，丁玲进入上海大学中国文学系旁听和学习，当前条目按这一较为明确的文学活动节点展示。",
        "source_url": "https://m.thepaper.cn/newsDetail_forward_28996791 ; https://museum.shu.edu.cn/info/1034/1373.htm",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    215: {
        "standard_event_name": "胡风在日本因反日活动被捕",
        "canonical_event_key": "ZLH-026|胡风在日本因反日活动被捕|1933",
        "corrected_date": "1933",
        "date_precision": "年",
        "historical_location": "日本",
        "current_address": "日本（具体城市待核）",
        "event_scope": "entity",
        "correction_reason": "湖北政协文史资料记载胡风因参加反日与左翼文化活动遭日本政府逮捕并驱逐回国；原表将该条泛写为1933年上海被捕事件，现先改回到可公开核实的国家与事件性质。",
        "display_note": "1933年，胡风因参与反日与左翼文化活动在日本遭逮捕并被驱逐回国，当前条目按公开可核的年级别事实展示。",
        "source_url": "https://hbzx.gov.cn/49/2014-09-15/5811.html",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    225: {
        "standard_event_name": "李求实被捕事件",
        "canonical_event_key": "李求实被捕事件|ZLH-020",
        "corrected_date": "1931-01-17",
        "date_precision": "日",
        "historical_location": "上海东方旅社、中山旅社秘密会议",
        "current_address": "上海市黄浦区汉口路、天津路一带",
        "event_scope": "entity",
        "correction_reason": "中国军网资料记载，1931年1月17日李求实（李伟森）与何孟雄、林育南等在上海东方旅社和中山旅社秘密开会，后因叛徒告密被捕；原表把日期统一写成1月1日，现先按可公开核实的日期修正。",
        "display_note": "1931年1月17日，李求实在上海东方旅社和中山旅社秘密会议链条中被捕，当前条目按这一可公开核实的日期展示。",
        "source_url": "https://www.81.cn/yljnt/2013-11/04/content_5630422.htm",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    226: {
        "standard_event_name": "艾芜被捕事件",
        "canonical_event_key": "ZLH-024|艾芜被捕事件|1931",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "上海",
        "current_address": "上海（具体被捕地点待考）",
        "event_scope": "entity",
        "correction_reason": "原始记录在“上海”“龙华”“龙华淞沪警备司令部”等地点之间冲突，且缺少可公开核实的具体被捕日期，现保守收口为1931年艾芜在上海遭拘押的相关记录。",
        "display_note": "当前仅能确认 1931 年艾芜在上海有遭拘押相关记录，具体被捕时间和地点待考。",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    227: {
        "standard_event_name": "艾芜被捕事件",
        "canonical_event_key": "ZLH-024|艾芜被捕事件|1931",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "上海",
        "current_address": "上海（具体被捕地点待考）",
        "event_scope": "entity",
        "correction_reason": "原始记录在“上海”“龙华”“龙华淞沪警备司令部”等地点之间冲突，且缺少可公开核实的具体被捕日期，现保守收口为1931年艾芜在上海遭拘押的相关记录。",
        "display_note": "当前仅能确认 1931 年艾芜在上海有遭拘押相关记录，具体被捕时间和地点待考。",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    228: {
        "standard_event_name": "艾芜被捕事件",
        "canonical_event_key": "ZLH-024|艾芜被捕事件|1931",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "上海",
        "current_address": "上海（具体被捕地点待考）",
        "event_scope": "entity",
        "correction_reason": "原始记录在“上海”“龙华”“龙华淞沪警备司令部”等地点之间冲突，且缺少可公开核实的具体被捕日期，现保守收口为1931年艾芜在上海遭拘押的相关记录。",
        "display_note": "当前仅能确认 1931 年艾芜在上海有遭拘押相关记录，具体被捕时间和地点待考。",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    229: {
        "standard_event_name": "沙汀被捕事件",
        "canonical_event_key": "ZLH-025|沙汀被捕事件|1931",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "上海",
        "current_address": "上海（具体被捕地点待考）",
        "event_scope": "entity",
        "correction_reason": "原始记录在“上海”“龙华”“龙华淞沪警备司令部”等地点之间冲突，且缺少可公开核实的具体被捕日期，现保守收口为1931年沙汀在上海遭拘押的相关记录。",
        "display_note": "当前仅能确认 1931 年沙汀在上海有遭拘押相关记录，具体被捕时间和地点待考。",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    230: {
        "standard_event_name": "沙汀被捕事件",
        "canonical_event_key": "ZLH-025|沙汀被捕事件|1931",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "上海",
        "current_address": "上海（具体被捕地点待考）",
        "event_scope": "entity",
        "correction_reason": "原始记录在“上海”“龙华”“龙华淞沪警备司令部”等地点之间冲突，且缺少可公开核实的具体被捕日期，现保守收口为1931年沙汀在上海遭拘押的相关记录。",
        "display_note": "当前仅能确认 1931 年沙汀在上海有遭拘押相关记录，具体被捕时间和地点待考。",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    231: {
        "standard_event_name": "沙汀被捕事件",
        "canonical_event_key": "ZLH-025|沙汀被捕事件|1931",
        "corrected_date": "1931",
        "date_precision": "年",
        "historical_location": "上海",
        "current_address": "上海（具体被捕地点待考）",
        "event_scope": "entity",
        "correction_reason": "原始记录在“上海”“龙华”“龙华淞沪警备司令部”等地点之间冲突，且缺少可公开核实的具体被捕日期，现保守收口为1931年沙汀在上海遭拘押的相关记录。",
        "display_note": "当前仅能确认 1931 年沙汀在上海有遭拘押相关记录，具体被捕时间和地点待考。",
        "confidence": "medium",
        "needs_manual_review": "yes",
    },
    185: {
        "standard_event_name": "柔石在东方旅社31号房间被捕",
        "canonical_event_key": "ZLH-016|柔石在东方旅社31号房间被捕|1931-01-17",
        "corrected_date": "1931-01-17",
        "date_precision": "日",
        "historical_location": "上海东方旅社31号房间",
        "current_address": "上海市黄浦区汉口路一带",
        "event_scope": "entity",
        "correction_reason": "中国作家网与人民网党史资料均记载，柔石于1931年1月17日下午1时40分在上海东方旅社31号房间参加秘密会议时被捕；原表仅保留1931年和笼统地点，现改写为可公开核实的具体事件。",
        "display_note": "1931年1月17日下午1时40分，柔石在上海东方旅社31号房间参加秘密会议时被捕，当前条目按公开可核日期和地点展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0701/c404019-31765706.html ; https://dangshi.people.com.cn/n1/2018/0208/c85037-29812774.html",
        "confidence": "high",
        "needs_manual_review": "no",
    },
    186: {
        "standard_event_name": "柔石在东方旅社31号房间被捕",
        "canonical_event_key": "ZLH-016|柔石在东方旅社31号房间被捕|1931-01-17",
        "corrected_date": "1931-01-17",
        "date_precision": "日",
        "historical_location": "上海东方旅社31号房间",
        "current_address": "上海市黄浦区汉口路一带",
        "event_scope": "entity",
        "correction_reason": "中国作家网与人民网党史资料均记载，柔石于1931年1月17日下午1时40分在上海东方旅社31号房间参加秘密会议时被捕；原表仅保留1931年和笼统地点，现改写为可公开核实的具体事件。",
        "display_note": "1931年1月17日下午1时40分，柔石在上海东方旅社31号房间参加秘密会议时被捕，当前条目按公开可核日期和地点展示。",
        "source_url": "https://www.chinawriter.com.cn/n1/2020/0701/c404019-31765706.html ; https://dangshi.people.com.cn/n1/2018/0208/c85037-29812774.html",
        "confidence": "high",
        "needs_manual_review": "no",
    },
    400: {
        "standard_event_name": "宋庆龄在莫利爱路寓所宴请萧伯纳",
        "canonical_event_key": "ZLH-145|宋庆龄在莫利爱路寓所宴请萧伯纳|1933-02-17",
        "corrected_date": "1933-02-17",
        "date_precision": "日",
        "historical_location": "宋庆龄莫利爱路寓所",
        "current_address": "上海市徐汇区武康路184号",
        "event_scope": "entity",
        "correction_reason": "宋庆龄故居纪念馆资料记载，1933年2月17日宋庆龄在莫利爱路寓所宴请萧伯纳，并与鲁迅、蔡元培等会见合影；原表将地点泛记为孙中山故居，现修正为莫利爱路寓所。",
        "display_note": "1933年2月17日，宋庆龄在莫利爱路寓所宴请萧伯纳，鲁迅、蔡元培等亦到场会见，当前条目按公开可核日期和地点展示。",
        "source_url": "https://sswgw.org.cn/wwdc/sqlww/9329.htm ; https://www.51ldb.com/shsldb/xf/content/017d83aaa824c0010000df844d7e124a.htm",
        "confidence": "high",
        "needs_manual_review": "no",
    },
}

VERIFICATION_SOURCES = [
    ("左联会址纪念馆与成立大会旧址", "https://www.shhk.gov.cn/xwzx/002008/002008040/20221031/bd8cb3ee-198a-431a-adf7-781e9fc5185d.html"),
    ("左联成立95周年主题活动", "https://www.shhk.gov.cn/xwzx/002003/20250303/ec139c7d-8fa3-4970-a5dd-3248468989c8.html"),
    ("左联五烈士专题纪念", "https://www.shhk.gov.cn/slh/038001/20260302/5f983b4c-74f7-4a6a-a2b4-930dedf99970.html"),
    ("左联会址纪念馆五烈士介绍", "https://www.shhk.gov.cn/xwzx/002006/20210722/07f62353-471f-40d9-880e-ea82be5da936.html"),
    ("龙华二十四烈士资料", "https://cpc.people.com.cn/n1/2022/1209/c443712-32583679.html"),
    ("内山书店旧址说明", "https://www.shhk.gov.cn/xwzx/002006/20210930/96ecb0ec-79e3-49ef-a097-a89c5a5dbc40.html"),
    ("内山书店今址活动页", "https://www.shhk.gov.cn/xwzx/002008/002008040/20240425/3e9546e4-0e0f-409d-a0e2-91bc115f8f66.html"),
    ("鲁迅与柔石看屋转录材料", "https://al3tai.nenzhu.com/news-id-2373.html"),
    ("公啡咖啡馆与左联筹备会", "https://www.thepaper.cn/newsDetail_forward_8389375"),
    ("澎湃新闻：成为“丁玲”之前，和上海的三次际会", "https://m.thepaper.cn/newsDetail_forward_28996791"),
    ("上海大学校史馆：历史上的上海大学（1923年）", "https://museum.shu.edu.cn/info/1034/1373.htm"),
    ("人民网党史频道：上海大学与瞿秋白", "https://cpc.people.com.cn/BIG5/n1/2024/1006/c443712-40333498.html"),
    ("中国作家网：鲁迅帮助狱中的楼适夷", "https://www.chinawriter.com.cn/n1/2020/0813/c404063-31820381.html"),
]


@dataclass
class EntityInfo:
    entity_id: str
    name: str
    birth_year: Optional[int]
    death_year: Optional[int]
    role: str


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_\-]+", "", str(value or "").strip().lower())


def find_input_file(cwd: Path) -> Path:
    preferred = cwd / "知识库" / "data" / TARGET_BASENAME
    if preferred.exists():
        return preferred
    exact = [p for p in cwd.rglob(TARGET_BASENAME) if "_修正版" not in p.name and "_修改日志" not in p.name]
    if exact:
        return sorted(exact, key=lambda p: len(str(p)))[0]
    candidates = [p for p in cwd.rglob("*.xlsx") if "左联相关档案资源目录" in p.name and "_修正版" not in p.name and "_修改日志" not in p.name]
    if candidates:
        return sorted(candidates, key=lambda p: len(str(p)))[0]
    raise FileNotFoundError(f"未找到输入文件：{TARGET_BASENAME}")


def map_sheet_name(sheet_names: Iterable[str], aliases: List[str]) -> str:
    alias_set = {normalize_header(a) for a in aliases}
    for name in sheet_names:
        if normalize_header(name) in alias_set:
            return name
    raise KeyError(f"未找到匹配 sheet：{aliases}")


def map_columns(headers: List[Any], alias_map: Dict[str, List[str]]) -> Dict[str, int]:
    normalized_headers = {normalize_header(h): idx for idx, h in enumerate(headers)}
    result: Dict[str, int] = {}
    for logical_field, aliases in alias_map.items():
        found = None
        for alias in aliases:
            alias_norm = normalize_header(alias)
            if alias_norm in normalized_headers:
                found = normalized_headers[alias_norm]
                break
        if found is None:
            raise KeyError(f"未找到列：{logical_field}")
        result[logical_field] = found
    return result


def parse_birth_death(text: Any) -> Tuple[Optional[int], Optional[int]]:
    if text is None:
        return None, None
    m = re.match(r"^\s*(\d{4}|\?)\s*-\s*(\d{4}|\?)\s*$", str(text))
    if not m:
        return None, None
    birth = int(m.group(1)) if m.group(1).isdigit() else None
    death = int(m.group(2)) if m.group(2).isdigit() else None
    return birth, death


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def extract_year(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year
    s = as_text(value)
    m = re.match(r"^(\d{4})", s)
    return int(m.group(1)) if m else None


def infer_date_precision(value: Any) -> str:
    s = as_text(value)
    if not s:
        return ""
    if isinstance(value, datetime):
        return "日"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return "年" if s.endswith("-01-01") else "日"
    if re.fullmatch(r"\d{4}-\d{2}", s):
        return "月"
    if re.fullmatch(r"\d{4}", s):
        return "年"
    return ""


def is_placeholder_jan1(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, datetime):
        return value.month == 1 and value.day == 1
    return bool(re.fullmatch(r"\d{4}-01-01", as_text(value)))


def contains_any(text: str, keywords: Iterable[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def is_generic_activity_event_name(event_name: str) -> bool:
    if not event_name:
        return False
    if contains_any(event_name, ["成立大会", "遇难", "被捕", "秘密会议", "会面", "论战", "集会"]):
        return False
    return contains_any(
        event_name,
        ["文学活动", "交往活动", "社交活动", "社会活动", "一般活动", "交流活动", "文学交流", "上海活动", "活动"],
    )


def is_generic_arrest_event_name(event_name: str) -> bool:
    if not event_name:
        return False
    return contains_any(event_name, ["被捕事件", "上海被捕事件", "被俘事件", "逮捕事件", "被捕", "逮捕"])


def is_generic_visit_event_name(event_name: str) -> bool:
    if not event_name:
        return False
    return contains_any(event_name, ["访华会", "访华"])


def should_split_event_by_year(event_name: str) -> bool:
    if not event_name:
        return False
    return (
        is_generic_activity_event_name(event_name)
        or is_generic_arrest_event_name(event_name)
        or is_generic_visit_event_name(event_name)
        or event_name.endswith("记录")
    )


def should_convert_to_annual_record(event_name: str, date_precision: str) -> bool:
    if date_precision != "年" or not event_name:
        return False
    return (
        is_generic_activity_event_name(event_name)
        or is_generic_arrest_event_name(event_name)
        or is_generic_visit_event_name(event_name)
    )


def normalize_event_name_alias(event_name: str) -> str:
    cleaned = as_text(event_name)
    if "内山完造秘密会议" in cleaned:
        return cleaned.replace("内山完造秘密会议", "内山书店秘密会议")
    return cleaned


def normalize_event_location_aliases(event_name: str, historical_location: str, current_address: str) -> Tuple[str, str, str]:
    normalized_event = normalize_event_name_alias(event_name)
    hist = as_text(historical_location)
    cur = as_text(current_address)
    if normalized_event.endswith("内山书店秘密会议") and "四川北路2050号" in cur:
        hist = "内山书店旧址"
    if "四川北路2050号" in cur and "内山完造" in hist:
        hist = "内山书店旧址" if "秘密会议" in normalized_event else "内山书店"
    return normalized_event, hist, cur


def strip_entity_prefix(event_name: str, entity_name: str) -> str:
    cleaned_event = as_text(event_name)
    cleaned_entity = as_text(entity_name)
    if cleaned_entity and cleaned_event.startswith(cleaned_entity):
        return cleaned_event[len(cleaned_entity) :].strip()
    return cleaned_event


def build_annual_record_label(base: str) -> str:
    cleaned_base = as_text(base)
    if cleaned_base in {"被捕事件", "逮捕事件"}:
        return "遭拘押相关记录"
    if cleaned_base == "被捕":
        return "被捕相关记录"
    if cleaned_base.endswith("访华会"):
        return cleaned_base[:-1] + "相关记录"
    return cleaned_base if cleaned_base.endswith("记录") else f"{cleaned_base}记录"


def describe_annual_record_topic(source_event_name: str, entity_name: str) -> str:
    base = strip_entity_prefix(source_event_name, entity_name) or source_event_name
    if base in {"被捕事件", "逮捕事件"}:
        return "遭拘押"
    if base == "被捕":
        return "被捕"
    if base.endswith("访华会"):
        return base[:-1]
    return f"“{base}”"


def build_annual_record_name(event_name: str, entity_name: str, event_scope: str, corrected_date: str) -> str:
    year = corrected_date[:4] if corrected_date else ""
    if event_scope == "entity" and entity_name:
        base = strip_entity_prefix(event_name, entity_name) or event_name
        prefix = f"{entity_name}{year}年" if year else entity_name
        suffix = build_annual_record_label(base)
        return f"{prefix}{suffix}"
    prefix = f"{year}年" if year else ""
    cleaned_event = normalize_event_name_alias(event_name)
    return f"{prefix}{build_annual_record_label(cleaned_event)}"


def build_annual_record_note(
    *,
    entity_name: str,
    source_event_name: str,
    corrected_date: str,
    historical_location: str,
    current_address: str,
    event_scope: str,
) -> str:
    year = corrected_date[:4] if corrected_date else "相关年份"
    location = historical_location or current_address or "相关地点"
    topic = describe_annual_record_topic(source_event_name, entity_name)
    if event_scope == "entity" and entity_name:
        return f"当前仅能确认 {year} 年 {entity_name} 在 {location} 有{topic}相关记录，尚不足以定位到单一具体事件。"
    return f"当前仅能确认 {year} 年在 {location} 存在{topic}相关记录，尚不足以定位到单一具体事件。"


def coarsen_annual_record_location(source_event_name: str, historical_location: str, current_address: str) -> Tuple[str, str]:
    source = normalize_event_name_alias(source_event_name)
    if "创造社" in source:
        return "创造社（宝山路）", "闸北区宝山路一带"
    if "光华书局" in source:
        return "光华书局（四马路）", "福州路（四马路）"
    if "明星影片公司" in source:
        return "明星影片公司", "徐家汇路一带"
    if is_generic_arrest_event_name(source):
        return "上海", "上海（具体拘押地点待考）"
    if is_generic_visit_event_name(source):
        return "上海", "上海（具体活动地点待考）"
    return historical_location, current_address


def build_default_event_cluster_key(entity_id: str, event_name: str, timestamp: Any) -> str:
    normalized_event_name = normalize_event_name_alias(event_name)
    if entity_id and should_split_event_by_year(normalized_event_name):
        year = extract_year(timestamp)
        suffix = str(year) if year else "unknown"
        return f"{entity_id}|{normalized_event_name}|{suffix}"
    return normalized_event_name


def resolve_event_cluster_key(row_num: int, entity_id: str, event_name: str, timestamp: Any) -> str:
    override = ROW_EVENT_OVERRIDES.get(row_num, {})
    if override.get("canonical_event_key"):
        return as_text(override["canonical_event_key"])
    if override.get("standard_event_name"):
        return build_default_event_cluster_key(
            entity_id,
            as_text(override.get("standard_event_name")),
            override.get("corrected_date") or timestamp,
        )

    rule = VERIFIED_EVENT_RULES.get(event_name, {})
    if rule.get("canonical_event_key"):
        return as_text(rule["canonical_event_key"])
    if rule.get("standard_event_name"):
        return build_default_event_cluster_key(
            entity_id,
            as_text(rule.get("standard_event_name")),
            rule.get("corrected_date") or timestamp,
        )
    return build_default_event_cluster_key(entity_id, normalize_event_name_alias(event_name), timestamp)


def build_display_note(
    *,
    entity_name: str,
    standard_event_name: str,
    corrected_date: str,
    date_precision: str,
    historical_location: str,
    current_address: str,
    correction_reason: str,
    explicit_note: str,
) -> str:
    if explicit_note:
        return explicit_note

    location = historical_location or current_address or "相关地点"
    subject = entity_name if entity_name and entity_name in standard_event_name else standard_event_name

    if "原始日期疑似以1月1日填补未知月份和日期" in correction_reason:
        if date_precision == "年" and corrected_date:
            return f"当前仅能确认 {corrected_date} 年 {subject} 在 {location} 有相关记录，具体月份和日期待考。"
        if date_precision == "月" and corrected_date:
            return f"当前可确认 {subject} 在 {corrected_date} 于 {location} 有相关活动，更精确日期待考。"

    if "同名事件存在时间或地点冲突" in correction_reason:
        if corrected_date and date_precision == "日":
            return f"当前条目按已核实的日期 {corrected_date} 展示；原始同名记录在时间或地点上仍存在版本冲突。"
        if corrected_date and date_precision == "月":
            return f"当前条目按 {corrected_date} 的月份信息展示；同名记录在具体时间或地点上仍存在分歧。"
        if corrected_date and date_precision == "年":
            if standard_event_name.endswith("记录"):
                return f"当前仅能确认 {corrected_date} 年在 {location} 存在相关年度记录，原始同名条目在时间或地点上仍有分歧。"
            return f"当前仅能确认 {corrected_date} 年 {subject} 在 {location} 有相关记录，同名条目在时间或地点上仍有分歧。"
        return "同名记录在时间或地点上存在分歧，当前条目按保守口径展示。"

    if correction_reason:
        return correction_reason
    if corrected_date:
        return f"{subject}于{corrected_date}在{location}有相关记录。"
    return f"当前仅能确认 {subject} 与 {location} 存在相关记录，具体时间待考。"


def is_list_like(text: str) -> bool:
    if not text:
        return False
    score = text.count("、") + text.count(";") + text.count("；") + text.count(",")
    return score >= 6 or ("等" in text and text.count("、") >= 3)


def has_ocr_noise(text: str) -> bool:
    if not text:
        return False
    patterns = [r"[`]{2,}", r"[_]{2,}", r"[“”\"]{3,}", r"[A-Za-z]{6,}", r"[…]{2,}", r"[)\(]{2,}", r"�"]
    return any(re.search(pattern, text) for pattern in patterns)


def relation_risk_assessment(relation_type: str, context: str) -> Tuple[int, str, List[str]]:
    score = 0
    reasons: List[str] = []
    if relation_type in {"组织隶属", "亲属关系", "师生关系"}:
        score += 45
        reasons.append("人物-人物关系使用强证据型关系标签")
    if relation_type == "组织隶属":
        score += 15
        reasons.append("人物-人物之间出现“组织隶属”语义不自然")
    if is_list_like(context):
        score += 25
        reasons.append("证据更像长名单并列共现")
    if has_ocr_noise(context):
        score += 20
        reasons.append("存在明显OCR噪声")
    if relation_type == "亲属关系" and not contains_any(context, KINSHIP_KEYWORDS):
        score += 20
        reasons.append("缺少亲属关键词")
    if relation_type == "师生关系" and not contains_any(context, TEACH_KEYWORDS):
        score += 20
        reasons.append("缺少师生关键词")
    if relation_type == "组织隶属" and not contains_any(context, ORG_KEYWORDS):
        score += 15
        reasons.append("缺少组织归属关键词")
    if score >= 80:
        return score, "critical", reasons
    if score >= 60:
        return score, "high", reasons
    if score >= 35:
        return score, "medium", reasons
    return score, "low", reasons


def infer_relation_type(relation_type: str, context: str, risk_level: str) -> Tuple[str, str, str, str]:
    text = context or ""
    if relation_type == "组织隶属":
        if contains_any(text, COMM_KEYWORDS):
            return "通信", "relation_type_mismatch", "medium", "no"
        if contains_any(text, VISIT_KEYWORDS):
            return "交往", "relation_type_mismatch", "medium", "no"
        if contains_any(text, COOP_KEYWORDS):
            return "合作", "relation_type_mismatch", "medium", "no"
        if contains_any(text, DEBATE_KEYWORDS):
            return "论战", "relation_type_mismatch", "medium", "no"
        if contains_any(text, MEMORIAL_KEYWORDS):
            return "纪念/悼念", "relation_type_mismatch", "medium", "no"
        if contains_any(text, ORG_KEYWORDS):
            return "同属组织", "person_person_org_relation", "low", "yes"
        return "待核验", "weak_evidence", "low", "yes"
    if relation_type in {"亲属关系", "师生关系"}:
        if relation_type == "亲属关系" and contains_any(text, KINSHIP_KEYWORDS):
            return relation_type, "verified_direct_evidence", "medium", "no"
        if relation_type == "师生关系" and contains_any(text, TEACH_KEYWORDS):
            return relation_type, "verified_direct_evidence", "medium", "no"
        if contains_any(text, COMM_KEYWORDS):
            return "通信", "relation_type_mismatch", "low", "yes"
        if contains_any(text, VISIT_KEYWORDS):
            return "交往", "relation_type_mismatch", "low", "yes"
        return "待核验", "strong_claim_without_support", "low", "yes"
    if risk_level in {"critical", "high"} and is_list_like(text):
        return "待核验", "weak_evidence", "low", "yes"
    return relation_type, "", "", ""


def is_personal_activity_event(event_name: str) -> bool:
    return any(keyword in event_name for keyword in ACTIVE_EVENT_KEYWORDS)


def infer_event_scope(event_name: str, entity_name: str) -> str:
    cleaned_event = as_text(event_name)
    cleaned_entity = as_text(entity_name)
    if cleaned_event and cleaned_entity and cleaned_entity in cleaned_event:
        return "entity"
    return "collective"


def build_canonical_event_key(event_name: str, entity_id: str, event_scope: str, timestamp: Any = None) -> str:
    cleaned_event = normalize_event_name_alias(event_name)
    cleaned_entity = as_text(entity_id)
    if event_scope == "entity" and cleaned_entity:
        if should_split_event_by_year(cleaned_event):
            year = extract_year(timestamp)
            suffix = str(year) if year else "unknown"
            return f"{cleaned_event}|{cleaned_entity}|{suffix}"
        return f"{cleaned_event}|{cleaned_entity}"
    return cleaned_event


def clone_sheet(wb, source_name: str, target_name: str):
    if target_name in wb.sheetnames:
        del wb[target_name]
    ws = wb.copy_worksheet(wb[source_name])
    ws.title = target_name
    return ws


def append_headers(ws, headers: List[str]) -> Dict[str, int]:
    start_col = ws.max_column + 1
    positions: Dict[str, int] = {}
    for offset, header in enumerate(headers):
        col = start_col + offset
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = copy(ws.cell(row=1, column=1).font) if ws.cell(row=1, column=1).font else Font(bold=True)
        positions[header] = col
    return positions


def write_value(ws, row_num: int, col_idx: int, value: Any):
    ws.cell(row=row_num, column=col_idx, value=value)


def add_log(logs: List[Dict[str, Any]], **kwargs):
    if as_text(kwargs["original_value"]) == as_text(kwargs["new_value"]):
        return
    logs.append(
        {
            "sheet_name": kwargs["sheet_name"],
            "row_number": kwargs["row_number"],
            "primary_key": kwargs["primary_key"],
            "column_name": kwargs["column_name"],
            "original_value": as_text(kwargs["original_value"]),
            "new_value": as_text(kwargs["new_value"]),
            "issue_type": kwargs["issue_type"],
            "correction_reason": kwargs["correction_reason"],
            "source_url": kwargs["source_url"],
            "evidence_ref_used": kwargs["evidence_ref_used"],
            "confidence": kwargs["confidence"],
            "needs_manual_review": kwargs["needs_manual_review"],
        }
    )


def create_aux_sheet(wb, title: str):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def build_entity_map(ws, mapping: Dict[str, int]) -> Dict[str, EntityInfo]:
    entities: Dict[str, EntityInfo] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        entity_id = as_text(row[mapping["entity_id"]])
        if not entity_id:
            continue
        birth_year, death_year = parse_birth_death(row[mapping["birth_death"]])
        entities[entity_id] = EntityInfo(entity_id, as_text(row[mapping["true_name"]]), birth_year, death_year, as_text(row[mapping["role"]]))
    return entities


def main():
    cwd = Path.cwd()
    input_path = find_input_file(cwd)
    output_dir = cwd
    output_path = output_dir / OUTPUT_BASENAME
    log_path = output_dir / LOG_BASENAME
    review_path = output_dir / REVIEW_BASENAME

    wb = load_workbook(input_path)
    sheet1_name = map_sheet_name(wb.sheetnames, SHEET_ALIASES["sheet1"])
    sheet2_name = map_sheet_name(wb.sheetnames, SHEET_ALIASES["sheet2"])
    sheet3_name = map_sheet_name(wb.sheetnames, SHEET_ALIASES["sheet3"])

    ws1 = wb[sheet1_name]
    ws2 = wb[sheet2_name]
    ws3 = wb[sheet3_name]

    headers1 = [c.value for c in ws1[1]]
    headers2 = [c.value for c in ws2[1]]
    headers3 = [c.value for c in ws3[1]]

    map1 = map_columns(headers1, SHEET1_COLS)
    map2 = map_columns(headers2, SHEET2_COLS)
    map3 = map_columns(headers3, SHEET3_COLS)
    entities = build_entity_map(ws1, map1)

    event_clusters: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"rows": [], "dates": set(), "locs": set(), "entities": set()})
    for row_num in range(2, ws3.max_row + 1):
        event_name = as_text(ws3.cell(row=row_num, column=map3["event"] + 1).value)
        timestamp = ws3.cell(row=row_num, column=map3["timestamp"] + 1).value
        hist_loc = as_text(ws3.cell(row=row_num, column=map3["hist_loc"] + 1).value)
        cur_loc = as_text(ws3.cell(row=row_num, column=map3["current_loc"] + 1).value)
        entity_id = as_text(ws3.cell(row=row_num, column=map3["entity_id"] + 1).value)
        cluster_key = resolve_event_cluster_key(row_num, entity_id, event_name, timestamp)
        event_clusters[cluster_key]["rows"].append(row_num)
        event_clusters[cluster_key]["dates"].add(as_text(timestamp))
        event_clusters[cluster_key]["locs"].add((hist_loc, cur_loc))
        event_clusters[cluster_key]["entities"].add(entity_id)
    duplicate_event_clusters = sum(1 for data in event_clusters.values() if len(data["dates"]) > 1 or len(data["locs"]) > 1)

    ws2c = clone_sheet(wb, sheet2_name, "Sheet2_corrected")
    ws3c = clone_sheet(wb, sheet3_name, "Sheet3_corrected")

    sheet2_cols = append_headers(
        ws2c,
        [
            "source_name",
            "target_name",
            "is_person_to_person",
            "context_quality_flags",
            "relation_quality_score",
            "relation_risk_level",
            "original_relation_type",
            "corrected_relation_type",
            "correction_reason",
            "source_url",
            "evidence_ref_used",
            "confidence",
            "needs_manual_review",
        ],
    )
    sheet3_cols = append_headers(
        ws3c,
        [
            "entity_name",
            "original_date",
            "corrected_date",
            "date_precision",
            "original_location",
            "historical_location",
            "current_address",
            "standard_event_name",
            "event_scope",
            "canonical_event_key",
            "display_note",
            "corrected_persons",
            "entity_role_in_event",
            "event_conflict_flags",
            "correction_reason",
            "source_url",
            "confidence",
            "needs_manual_review",
        ],
    )

    logs: List[Dict[str, Any]] = []
    review_rows: List[Dict[str, Any]] = []
    sheet2_auto_corrected_rows, sheet2_manual_review_rows = set(), set()
    sheet3_auto_corrected_rows, sheet3_manual_review_rows = set(), set()
    high_risk_relation_count = 0

    for row_num in range(2, ws2c.max_row + 1):
        source_id = as_text(ws2c.cell(row=row_num, column=map2["source_id"] + 1).value)
        target_id = as_text(ws2c.cell(row=row_num, column=map2["target_id"] + 1).value)
        relation_type = as_text(ws2c.cell(row=row_num, column=map2["relation_type"] + 1).value)
        context = as_text(ws2c.cell(row=row_num, column=map2["context"] + 1).value)
        evidence_ref = as_text(ws2c.cell(row=row_num, column=map2["evidence_ref"] + 1).value)
        seq = as_text(ws2c.cell(row=row_num, column=map2["seq"] + 1).value)
        source_name = entities.get(source_id).name if source_id in entities else ""
        target_name = entities.get(target_id).name if target_id in entities else ""
        is_person_to_person = "yes" if source_id in entities and target_id in entities else "no"
        score, risk_level, risk_reasons = relation_risk_assessment(relation_type, context)
        if risk_level in {"critical", "high"}:
            high_risk_relation_count += 1

        corrected_relation_type, issue_type, confidence, review_flag = infer_relation_type(relation_type, context, risk_level)
        quality_flags = []
        if is_list_like(context):
            quality_flags.append("list_like_evidence")
        if has_ocr_noise(context):
            quality_flags.append("ocr_noise")
        if relation_type in {"组织隶属", "亲属关系", "师生关系"}:
            quality_flags.append("strong_claim_type")
        if not context:
            quality_flags.append("empty_context")
        context_quality_flags = ",".join(quality_flags)

        correction_reason = "结构性检测未发现需要自动改写的明确证据。"
        source_url = ""
        if corrected_relation_type != relation_type:
            parts = list(risk_reasons)
            mapping_reason = {
                "通信": "上下文直接出现来信/收信/寄信类动作",
                "交往": "上下文直接出现来访/会见/共同行动类动作",
                "合作": "上下文直接出现联名/共同编辑/合作动作",
                "论战": "上下文直接出现论争/批评/论战动作",
                "纪念/悼念": "上下文直接出现纪念/悼念动作",
                "同属组织": "证据只能支持同属组织或共同组织关联",
                "待核验": "证据不足以支撑原关系类型",
            }
            if corrected_relation_type in mapping_reason:
                parts.append(mapping_reason[corrected_relation_type])
            correction_reason = "；".join(dict.fromkeys(parts))
            sheet2_auto_corrected_rows.add(row_num)
        else:
            confidence = confidence or "medium"
            review_flag = review_flag or ("yes" if risk_level in {"critical", "high"} else "no")
        if review_flag == "yes":
            sheet2_manual_review_rows.add(row_num)

        for header, value in {
            "source_name": source_name,
            "target_name": target_name,
            "is_person_to_person": is_person_to_person,
            "context_quality_flags": context_quality_flags,
            "relation_quality_score": score,
            "relation_risk_level": risk_level,
            "original_relation_type": relation_type,
            "corrected_relation_type": corrected_relation_type,
            "correction_reason": correction_reason,
            "source_url": source_url,
            "evidence_ref_used": evidence_ref,
            "confidence": confidence or "low",
            "needs_manual_review": review_flag or "no",
        }.items():
            write_value(ws2c, row_num, sheet2_cols[header], value)

        primary_key = f"{seq}|{source_id}|{target_id}"
        if corrected_relation_type != relation_type:
            add_log(
                logs,
                sheet_name=sheet2_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Relation_Type",
                original_value=relation_type,
                new_value=corrected_relation_type,
                issue_type=issue_type or "relation_type_mismatch",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used=evidence_ref,
                confidence=confidence or "low",
                needs_manual_review=review_flag or "no",
            )
        if context_quality_flags:
            add_log(
                logs,
                sheet_name=sheet2_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="context_quality_flags",
                original_value="",
                new_value=context_quality_flags,
                issue_type="quality_annotation",
                correction_reason="结构性检测记录上下文风险标记。",
                source_url=source_url,
                evidence_ref_used=evidence_ref,
                confidence="medium",
                needs_manual_review=review_flag or "no",
            )
        if review_flag == "yes":
            review_rows.append(
                {
                    "sheet_name": sheet2_name,
                    "row_number": row_num,
                    "primary_key": primary_key,
                    "issue_summary": correction_reason,
                    "source_url": source_url,
                    "evidence_ref_used": evidence_ref,
                    "confidence": confidence or "low",
                }
            )

    for row_num in range(2, ws3c.max_row + 1):
        seq = as_text(ws3c.cell(row=row_num, column=map3["seq"] + 1).value)
        entity_id = as_text(ws3c.cell(row=row_num, column=map3["entity_id"] + 1).value)
        entity_name = entities.get(entity_id).name if entity_id in entities else ""
        timestamp = ws3c.cell(row=row_num, column=map3["timestamp"] + 1).value
        hist_loc = as_text(ws3c.cell(row=row_num, column=map3["hist_loc"] + 1).value)
        current_loc = as_text(ws3c.cell(row=row_num, column=map3["current_loc"] + 1).value)
        event_name = as_text(ws3c.cell(row=row_num, column=map3["event"] + 1).value)
        normalized_event_name, normalized_hist_loc, normalized_current_loc = normalize_event_location_aliases(
            event_name,
            hist_loc,
            current_loc,
        )
        original_date = as_text(timestamp)
        original_location = " / ".join([v for v in [hist_loc, current_loc] if v])
        raw_event_scope = infer_event_scope(normalized_event_name, entity_name)
        raw_cluster_key = resolve_event_cluster_key(row_num, entity_id, normalized_event_name, timestamp)
        cluster = event_clusters.get(raw_cluster_key, {"dates": set(), "locs": set(), "rows": []})

        conflict_flags: List[str] = []
        if is_placeholder_jan1(timestamp):
            conflict_flags.append("placeholder_jan1")
        if len(cluster["dates"]) > 1:
            conflict_flags.append("event_multi_dates")
        if len(cluster["locs"]) > 1:
            conflict_flags.append("event_multi_locations")

        corrected_date = original_date
        date_precision = infer_date_precision(timestamp)
        historical_location = normalized_hist_loc
        current_address = normalized_current_loc
        standard_event_name = normalized_event_name
        event_scope = raw_event_scope
        canonical_event_key = build_canonical_event_key(standard_event_name, entity_id, event_scope, timestamp)
        display_note = ""
        corrected_persons = ""
        entity_role = "unclear"
        correction_reason = ""
        source_url = ""
        confidence = "low"
        needs_manual_review = "no"

        rule = VERIFIED_EVENT_RULES.get(event_name) or VERIFIED_EVENT_RULES.get(normalized_event_name)
        explicit_canonical_key = False
        if rule:
            standard_event_name = rule.get("standard_event_name") or event_name
            if rule.get("corrected_date") is not None:
                corrected_date = rule.get("corrected_date", "")
            if rule.get("date_precision") is not None:
                date_precision = rule.get("date_precision", "")
            if rule.get("historical_location"):
                historical_location = rule["historical_location"]
            if rule.get("current_address"):
                current_address = rule["current_address"]
            if rule.get("event_scope"):
                event_scope = rule["event_scope"]
            if rule.get("canonical_event_key"):
                canonical_event_key = rule["canonical_event_key"]
                explicit_canonical_key = True
            display_note = rule.get("display_note", display_note)
            corrected_persons = rule.get("corrected_persons", "")
            correction_reason = rule.get("correction_reason", "")
            source_url = rule.get("source_url", "")
            confidence = rule.get("confidence", "medium")
            needs_manual_review = "yes" if rule.get("manual_review_default") else "no"

            if event_name == "左联成立大会":
                entity_role = "待核"
                if entity_name not in {"鲁迅", "柔石", "潘汉年", "蒋光慈", "钱杏邨", "李求实", "殷夫", "艾芜"}:
                    needs_manual_review = "yes"
            elif event_name == "五烈士遇难":
                participants = rule.get("participant_names", set())
                if entity_name in participants:
                    entity_role = "直接参与者"
                else:
                    entity_role = "关联人物"
                    needs_manual_review = "yes"
                    conflict_flags.append("entity_not_direct_participant")
            elif event_name in {"冯铿被捕事件", "殷夫被捕事件", "李求实被捕事件", "丁玲被捕事件"}:
                participants = rule.get("participant_names", set())
                if entity_name in participants:
                    entity_role = "直接参与者"
                else:
                    entity_role = "关联人物"
                    needs_manual_review = "yes"
                    conflict_flags.append("entity_not_direct_participant")
            elif event_name == "鲁迅与柔石会面":
                entity_role = "直接参与者" if entity_name in {"鲁迅", "柔石"} else "关联人物"
                if entity_role != "直接参与者":
                    needs_manual_review = "yes"
                    conflict_flags.append("entity_not_direct_participant")
            else:
                entity_role = "待核"

        override = ROW_EVENT_OVERRIDES.get(row_num)
        if override:
            corrected_date = override.get("corrected_date", corrected_date)
            date_precision = override.get("date_precision", date_precision)
            standard_event_name = override.get("standard_event_name", standard_event_name)
            historical_location = override.get("historical_location", historical_location)
            current_address = override.get("current_address", current_address)
            corrected_persons = override.get("corrected_persons", corrected_persons)
            entity_role = override.get("entity_role_in_event", entity_role)
            event_scope = override.get("event_scope", event_scope)
            canonical_event_key = override.get("canonical_event_key", canonical_event_key)
            display_note = override.get("display_note", display_note)
            correction_reason = override.get("correction_reason", correction_reason)
            source_url = override.get("source_url", source_url)
            confidence = override.get("confidence", confidence)
            needs_manual_review = override.get("needs_manual_review", needs_manual_review)
            if override.get("canonical_event_key"):
                explicit_canonical_key = True
            conflict_flags.append("row_override_review")

        if not explicit_canonical_key:
            canonical_event_key = build_canonical_event_key(standard_event_name, entity_id, event_scope, corrected_date or timestamp)
        if entity_role == "unclear" and event_scope == "entity" and entity_id:
            entity_role = "直接参与者"

        event_year = extract_year(timestamp)
        entity_info = entities.get(entity_id)
        if entity_info and entity_info.death_year and event_year and event_year > entity_info.death_year and is_personal_activity_event(event_name):
            conflict_flags.append("after_death_conflict")
            corrected_date = ""
            date_precision = ""
            correction_reason = (correction_reason + "；" if correction_reason else "") + f"事件年份 {event_year} 晚于人物卒年 {entity_info.death_year}，且事件命名属于个人活动型事件，无法高置信保留。"
            confidence = "high"
            needs_manual_review = "yes"
            entity_role = "冲突"

        if not rule and not override and is_placeholder_jan1(timestamp) and "after_death_conflict" not in conflict_flags:
            year = extract_year(timestamp)
            corrected_date = str(year) if year else ""
            date_precision = "年" if year else ""
            correction_reason = "原始日期疑似以1月1日填补未知月份和日期，现仅保留年份精度。"
            confidence = "medium"
            needs_manual_review = "yes"

        if not rule and not override and (len(cluster["dates"]) > 1 or len(cluster["locs"]) > 1):
            needs_manual_review = "yes"
            correction_reason = (correction_reason + "；" if correction_reason else "") + "同名事件存在时间或地点冲突。"

        if corrected_date and re.fullmatch(r"\d{4}", corrected_date):
            date_precision = "年"
        elif corrected_date and re.fullmatch(r"\d{4}-\d{2}", corrected_date):
            date_precision = "月"
        elif corrected_date and re.fullmatch(r"\d{4}-\d{2}-\d{2}", corrected_date):
            date_precision = "日"

        annual_record_source_name = standard_event_name
        if should_convert_to_annual_record(standard_event_name, date_precision):
            standard_event_name = build_annual_record_name(standard_event_name, entity_name, event_scope, corrected_date)
            historical_location, current_address = coarsen_annual_record_location(
                annual_record_source_name,
                historical_location,
                current_address,
            )
            if not display_note:
                display_note = build_annual_record_note(
                    entity_name=entity_name,
                    source_event_name=annual_record_source_name,
                    corrected_date=corrected_date,
                    historical_location=historical_location,
                    current_address=current_address,
                    event_scope=event_scope,
                )
            if not explicit_canonical_key:
                canonical_event_key = build_canonical_event_key(standard_event_name, entity_id, event_scope, corrected_date or timestamp)

        display_note = build_display_note(
            entity_name=entity_name,
            standard_event_name=standard_event_name,
            corrected_date=corrected_date,
            date_precision=date_precision,
            historical_location=historical_location,
            current_address=current_address,
            correction_reason=correction_reason,
            explicit_note=display_note,
        )

        row_changed = any(
            [
                corrected_date != original_date,
                historical_location != hist_loc,
                current_address != current_loc,
                standard_event_name != event_name,
                canonical_event_key != raw_cluster_key,
                corrected_persons != "",
                correction_reason != "",
                display_note != "",
            ]
        )
        if row_changed:
            sheet3_auto_corrected_rows.add(row_num)
        if needs_manual_review == "yes":
            sheet3_manual_review_rows.add(row_num)

        if not correction_reason and conflict_flags:
            correction_reason = "发现事件结构性冲突或精度问题。"

        for header, value in {
            "entity_name": entity_name,
            "original_date": original_date,
            "corrected_date": corrected_date,
            "date_precision": date_precision,
            "original_location": original_location,
            "historical_location": historical_location,
            "current_address": current_address,
            "standard_event_name": standard_event_name,
            "event_scope": event_scope,
            "canonical_event_key": canonical_event_key,
            "display_note": display_note,
            "corrected_persons": corrected_persons,
            "entity_role_in_event": entity_role,
            "event_conflict_flags": ",".join(dict.fromkeys(conflict_flags)),
            "correction_reason": correction_reason,
            "source_url": source_url,
            "confidence": confidence,
            "needs_manual_review": needs_manual_review,
        }.items():
            write_value(ws3c, row_num, sheet3_cols[header], value)

        primary_key = f"{seq}|{entity_id}|{event_name}"
        if corrected_date != original_date:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Timestamp",
                original_value=original_date,
                new_value=corrected_date,
                issue_type="date_correction",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if historical_location != hist_loc:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Hist_Loc",
                original_value=hist_loc,
                new_value=historical_location,
                issue_type="location_correction",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if current_address != current_loc:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Current_Loc",
                original_value=current_loc,
                new_value=current_address,
                issue_type="location_correction",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if standard_event_name != event_name:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="Event",
                original_value=event_name,
                new_value=standard_event_name,
                issue_type="event_standardization",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if corrected_persons:
            add_log(
                logs,
                sheet_name=sheet3_name,
                row_number=row_num,
                primary_key=primary_key,
                column_name="corrected_persons",
                original_value="",
                new_value=corrected_persons,
                issue_type="participant_annotation",
                correction_reason=correction_reason,
                source_url=source_url,
                evidence_ref_used="",
                confidence=confidence,
                needs_manual_review=needs_manual_review,
            )
        if needs_manual_review == "yes":
            review_rows.append(
                {
                    "sheet_name": sheet3_name,
                    "row_number": row_num,
                    "primary_key": primary_key,
                    "issue_summary": correction_reason or ",".join(conflict_flags),
                    "source_url": source_url,
                    "evidence_ref_used": "",
                    "confidence": confidence,
                }
            )

    field_map_ws = create_aux_sheet(wb, "field_mapping")
    field_map_ws.append(["sheet_name", "logical_field", "column_header", "column_index"])
    for logical, idx in map1.items():
        field_map_ws.append([sheet1_name, logical, headers1[idx], idx + 1])
    for logical, idx in map2.items():
        field_map_ws.append([sheet2_name, logical, headers2[idx], idx + 1])
    for logical, idx in map3.items():
        field_map_ws.append([sheet3_name, logical, headers3[idx], idx + 1])

    conflict_ws = create_aux_sheet(wb, "event_conflicts")
    conflict_ws.append(["event_name", "row_count", "unique_dates", "unique_locations", "sample_rows"])
    for event_name, data in sorted(event_clusters.items(), key=lambda kv: (-len(kv[1]["rows"]), kv[0])):
        if len(data["dates"]) > 1 or len(data["locs"]) > 1:
            conflict_ws.append([event_name, len(data["rows"]), len(data["dates"]), len(data["locs"]), ",".join(str(r) for r in data["rows"][:10])])

    sources_ws = create_aux_sheet(wb, "verification_sources")
    sources_ws.append(["title", "url"])
    for title, url in VERIFICATION_SOURCES:
        sources_ws.append([title, url])

    summary_ws = create_aux_sheet(wb, "summary")
    summary_ws.append(["metric", "value"])
    summary_rows = [
        ("input_file", str(input_path)),
        ("output_file", str(output_path)),
        ("log_file", str(log_path)),
        ("review_file", str(review_path)),
        ("sheet2_total_records", ws2.max_row - 1),
        ("sheet2_auto_corrected_rows", len(sheet2_auto_corrected_rows)),
        ("sheet2_manual_review_rows", len(sheet2_manual_review_rows)),
        ("sheet3_total_records", ws3.max_row - 1),
        ("sheet3_auto_corrected_rows", len(sheet3_auto_corrected_rows)),
        ("sheet3_manual_review_rows", len(sheet3_manual_review_rows)),
        ("duplicate_event_clusters", duplicate_event_clusters),
        ("high_risk_person_relations", high_risk_relation_count),
    ]
    for metric, value in summary_rows:
        summary_ws.append([metric, value])
    wb.save(output_path)

    log_wb = Workbook()
    log_ws = log_wb.active
    log_ws.title = "modification_log"
    log_headers = [
        "sheet_name",
        "row_number",
        "primary_key",
        "column_name",
        "original_value",
        "new_value",
        "issue_type",
        "correction_reason",
        "source_url",
        "evidence_ref_used",
        "confidence",
        "needs_manual_review",
    ]
    log_ws.append(log_headers)
    for item in logs:
        log_ws.append([item[h] for h in log_headers])

    log_summary_ws = log_wb.create_sheet("summary")
    log_summary_ws.append(["metric", "value"])
    for metric, value in summary_rows:
        log_summary_ws.append([metric, value])

    review_ws = log_wb.create_sheet("review_needed")
    review_headers = ["sheet_name", "row_number", "primary_key", "issue_summary", "source_url", "evidence_ref_used", "confidence"]
    review_ws.append(review_headers)
    for item in review_rows:
        review_ws.append([item[h] for h in review_headers])
    log_wb.save(log_path)

    with review_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=review_headers)
        writer.writeheader()
        for item in review_rows:
            writer.writerow(item)

    print(f"INPUT={input_path}")
    print(f"OUTPUT={output_path}")
    print(f"LOG={log_path}")
    print(f"REVIEW={review_path}")
    print(f"SHEET2_TOTAL={ws2.max_row - 1}")
    print(f"SHEET2_AUTO_CORRECTED={len(sheet2_auto_corrected_rows)}")
    print(f"SHEET2_MANUAL_REVIEW={len(sheet2_manual_review_rows)}")
    print(f"SHEET3_TOTAL={ws3.max_row - 1}")
    print(f"SHEET3_AUTO_CORRECTED={len(sheet3_auto_corrected_rows)}")
    print(f"SHEET3_MANUAL_REVIEW={len(sheet3_manual_review_rows)}")
    print(f"DUPLICATE_EVENT_CLUSTERS={duplicate_event_clusters}")
    print(f"HIGH_RISK_PERSON_RELATIONS={high_risk_relation_count}")


if __name__ == "__main__":
    main()
