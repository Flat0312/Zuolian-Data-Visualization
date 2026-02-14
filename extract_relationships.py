# -*- coding: utf-8 -*-
"""
从鲁迅日记中提取人际关系数据并写入Excel
更新版 - 修正日期格式和关系类型分类
"""

import openpyxl
import re

# 人物对照表 - 包含所有别名映射
PERSON_MAP = {
    # ZLH-002 茅盾
    "茅盾": "ZLH-002", "沈雁冰": "ZLH-002", "玄珠": "ZLH-002", "方璧": "ZLH-002", "微明": "ZLH-002",
    # ZLH-003 瞿秋白
    "瞿秋白": "ZLH-003", "维它": "ZLH-003", "史铁儿": "ZLH-003", "宋阳": "ZLH-003", "何凝": "ZLH-003",
    # ZLH-004 夏衍
    "夏衍": "ZLH-004", "沈端先": "ZLH-004", "黄子布": "ZLH-004", "蔡叔亮": "ZLH-004",
    # ZLH-005 冯雪峰
    "冯雪峰": "ZLH-005", "画室": "ZLH-005", "洛扬": "ZLH-005", "成休": "ZLH-005", "雪峰": "ZLH-005",
    # ZLH-006 潘汉年
    "潘汉年": "ZLH-006", "萧叔安": "ZLH-006", "胡越": "ZLH-006", "小潘": "ZLH-006", "胡阿毛": "ZLH-006", "汉年": "ZLH-006",
    # ZLH-008 田汉
    "田汉": "ZLH-008", "寿昌": "ZLH-008", "陈瑜": "ZLH-008", "伯鸿": "ZLH-008", "汉仙": "ZLH-008",
    # ZLH-009 郑伯奇
    "郑伯奇": "ZLH-009", "伯奇": "ZLH-009",
    # ZLH-012 蒋光慈
    "蒋光慈": "ZLH-012", "蒋光赤": "ZLH-012", "光赤": "ZLH-012", "光慈": "ZLH-012",
    # ZLH-015 郁达夫
    "郁达夫": "ZLH-015", "郁文": "ZLH-015", "达夫": "ZLH-015",
    # ZLH-016 柔石
    "柔石": "ZLH-016", "赵平复": "ZLH-016", "平复": "ZLH-016",
    # ZLH-021 丁玲
    "丁玲": "ZLH-021", "蒋冰之": "ZLH-021", "彬芷": "ZLH-021",
    # ZLH-060 鲁彦
    "鲁彦": "ZLH-060", "王鲁彦": "ZLH-060", "王衡": "ZLH-060",
    # ZLH-071 李霁野
    "李霁野": "ZLH-071", "霁野": "ZLH-071",
    # ZLH-121 内山完造
    "内山完造": "ZLH-121", "邬其山": "ZLH-121", "内山": "ZLH-121",
    # ZLH-122 许广平
    "许广平": "ZLH-122", "景宋": "ZLH-122", "广平": "ZLH-122",
    # ZLH-124 李小峰
    "李小峰": "ZLH-124", "李致广": "ZLH-124", "小峰": "ZLH-124",
    # ZLH-126 沈兼士
    "沈兼士": "ZLH-126", "兼士": "ZLH-126",
    # ZLH-128 林语堂
    "林语堂": "ZLH-128", "语堂": "ZLH-128", "玉堂": "ZLH-128",
    # ZLH-129 胡愈之
    "胡愈之": "ZLH-129", "愈之": "ZLH-129",
    # ZLH-139 陈望道
    "陈望道": "ZLH-139", "望道": "ZLH-139",
}

# 中文数字转阿拉伯数字
CN_TO_NUM = {
    "一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
    "六": 6, "七": 7, "八": 8, "九": 9, "十": 10,
    "十一": 11, "十二": 12, "十三": 13, "十四": 14, "十五": 15,
    "十六": 16, "十七": 17, "十八": 18, "十九": 19, "二十": 20,
    "廿一": 21, "廿二": 22, "廿三": 23, "廿四": 24, "廿五": 25,
    "廿六": 26, "廿七": 27, "廿八": 28, "廿九": 29, "三十": 30,
    "三十一": 31, "卅一": 31,
    "二十一": 21, "二十二": 22, "二十三": 23, "二十四": 24, "二十五": 25,
    "二十六": 26, "二十七": 27, "二十八": 28, "二十九": 29,
}

def match_person(name):
    """匹配人物,返回ID或UNKNOWN"""
    name = name.strip()
    if name in PERSON_MAP:
        return PERSON_MAP[name]
    return "UNKNOWN"

def parse_day(day_str):
    """解析中文日期到数字"""
    day_str = day_str.replace('日', '').strip()
    if day_str in CN_TO_NUM:
        return CN_TO_NUM[day_str]
    return None

# 强关联人物 - 亲属/伴侣
STRONG_RELATIONS = {
    "ZLH-122": "强关联-亲属",  # 许广平 - 鲁迅伴侣
}

# 左联成员(组织隶属)
LEFT_LEAGUE_MEMBERS = {
    "ZLH-002", "ZLH-003", "ZLH-004", "ZLH-005", "ZLH-006", "ZLH-007", "ZLH-008",
    "ZLH-009", "ZLH-010", "ZLH-011", "ZLH-012", "ZLH-013", "ZLH-014", "ZLH-015",
    "ZLH-016", "ZLH-017", "ZLH-018", "ZLH-019", "ZLH-020", "ZLH-021",
}

def extract_from_line(line, year, month, day):
    """从单行日记中提取关系"""
    relationships = []
    date_str = f"{year}-{month:02d}-{day:02d}"
    evidence_ref = f"鲁迅日记 {year}年{month}月{day}日"
    
    # 已匹配的人物(避免重复)
    found = set()
    
    def get_relation_type(target_id, base_type):
        """根据人物ID确定关系类型"""
        # 许广平 - 亲属关系
        if target_id == "ZLH-122":
            return "强关联-亲属", 5
        # 左联成员 - 组织隶属
        if target_id in LEFT_LEAGUE_MEMBERS:
            if base_type == "时空共现":
                return "强关联-组织隶属", 4
            else:
                return "弱关联-通信", 3
        # 其他人
        if base_type == "时空共现":
            return "弱关联-时空共现", 2
        return "弱关联-通信", 2
    
    # 1. 会面/时空共现类 - "XXX来" 或 "XXX来访"
    for m in re.finditer(r'([^\s。，、：]+?)(?:来访|来(?!信|稿|函))', line):
        name = m.group(1).strip()
        if len(name) >= 2 and len(name) <= 4:
            target_id = match_person(name)
            if target_id != "UNKNOWN":
                key = (target_id, "时空共现")
                if key not in found:
                    found.add(key)
                    context = line[max(0, m.start()-10):min(len(line), m.end()+20)]
                    rel_type, weight = get_relation_type(target_id, "时空共现")
                    relationships.append({
                        'target_id': target_id,
                        'relation_type': rel_type,
                        'context': f"{name}来访。{context}",
                        'evidence_ref': evidence_ref,
                        'date': date_str,
                        'weight': weight
                    })
    
    # 2. 通信类 - "得XXX信"
    for m in re.finditer(r'得([^\s。，、：]+?)信', line):
        name = m.group(1).strip()
        if len(name) >= 2 and len(name) <= 4:
            target_id = match_person(name)
            if target_id != "UNKNOWN":
                key = (target_id, "通信")
                if key not in found:
                    found.add(key)
                    rel_type, weight = get_relation_type(target_id, "通信")
                    relationships.append({
                        'target_id': target_id,
                        'relation_type': rel_type,
                        'context': f"收到{name}来信",
                        'evidence_ref': evidence_ref,
                        'date': date_str,
                        'weight': weight
                    })
    
    # 3. 通信类 - "寄XXX信" 或 "复XXX信"
    for m in re.finditer(r'(?:寄|复)([^\s。，、：]+?)信', line):
        name = m.group(1).strip()
        if len(name) >= 2 and len(name) <= 4:
            target_id = match_person(name)
            if target_id != "UNKNOWN":
                key = (target_id, "通信")
                if key not in found:
                    found.add(key)
                    rel_type, weight = get_relation_type(target_id, "通信")
                    relationships.append({
                        'target_id': target_id,
                        'relation_type': rel_type,
                        'context': f"寄/复信给{name}",
                        'evidence_ref': evidence_ref,
                        'date': date_str,
                        'weight': weight
                    })
    
    # 4. 合作/宴请类 - 同席/晚餐/午餐
    for m in re.finditer(r'(?:同|与|邀|偕)([^\s。，、：]+?)(?:至|往|赴|在).{0,15}(?:午餐|夜餐|晚餐|午饭|夜饭|晚饭|饮茗)', line):
        name = m.group(1).strip()
        if len(name) >= 2 and len(name) <= 4:
            target_id = match_person(name)
            if target_id != "UNKNOWN":
                key = (target_id, "时空共现")
                if key not in found:
                    found.add(key)
                    rel_type, weight = get_relation_type(target_id, "时空共现")
                    # 共餐权重+1
                    relationships.append({
                        'target_id': target_id,
                        'relation_type': rel_type,
                        'context': f"与{name}共餐/饮茗",
                        'evidence_ref': evidence_ref,
                        'date': date_str,
                        'weight': min(5, weight + 1)
                    })
    
    # 5. 拜访类 - "访XXX"
    for m in re.finditer(r'访([^\s。，、：]+?)(?:$|，|。|于|未遇)', line):
        name = m.group(1).strip()
        if len(name) >= 2 and len(name) <= 4:
            target_id = match_person(name)
            if target_id != "UNKNOWN":
                key = (target_id, "时空共现")
                if key not in found:
                    found.add(key)
                    rel_type, weight = get_relation_type(target_id, "时空共现")
                    relationships.append({
                        'target_id': target_id,
                        'relation_type': rel_type,
                        'context': f"拜访{name}",
                        'evidence_ref': evidence_ref,
                        'date': date_str,
                        'weight': weight
                    })
    
    return relationships


def main():
    # 读取日记文本
    diary_path = r'd:\1大创\日记全编：全2册 (鲁迅 著) (Z-Library).txt'
    excel_path = r'd:\1大创\大创数据收集1.xlsx'
    
    with open(diary_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    all_relationships = []
    
    # 解析日记
    current_year = None
    current_month = None
    current_day = None
    
    month_names = {
        '一月': 1, '二月': 2, '三月': 3, '四月': 4, '五月': 5, '六月': 6,
        '七月': 7, '八月': 8, '九月': 9, '十月': 10, '十一月': 11, '十二月': 12
    }
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 检测年份 - "日记十六（1927年）"
        year_match = re.search(r'日记.+（(\d{4})年）', line)
        if year_match:
            current_year = int(year_match.group(1))
            print(f"进入 {current_year} 年")
            continue
        
        # 检测月份 - 单独一行的月份
        if line in month_names:
            current_month = month_names[line]
            continue
        
        # 检测日期 - "一日　晴。..."
        day_match = re.match(r'^([一二三四五六七八九十廿卅]+日)\s*(.+)$', line)
        if day_match and current_year and current_month:
            day_str = day_match.group(1)
            content = day_match.group(2)
            current_day = parse_day(day_str)
            
            if current_day:
                # 提取关系
                rels = extract_from_line(content, current_year, current_month, current_day)
                all_relationships.extend(rels)
    
    print(f"\n总共提取到 {len(all_relationships)} 条关系")
    
    # 统计已识别和未识别的
    identified = [r for r in all_relationships if r['target_id'] != 'UNKNOWN']
    print(f"已匹配ID: {len(identified)} 条")
    
    # 加载Excel并写入Sheet2
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['Sheet2']
    
    # 清除旧数据(保留表头)
    for row in range(2, sheet.max_row + 1):
        for col in range(1, 8):
            sheet.cell(row=row, column=col, value=None)
    
    # 写入新数据
    row_num = 2
    seq = 1
    
    for rel in identified:
        sheet.cell(row=row_num, column=1, value=seq)  # 序号
        sheet.cell(row=row_num, column=2, value='ZLH-001')  # Source_ID (鲁迅)
        sheet.cell(row=row_num, column=3, value=rel['target_id'])  # Target_ID
        sheet.cell(row=row_num, column=4, value=rel['relation_type'])  # Relation_Type
        sheet.cell(row=row_num, column=5, value=rel['context'])  # Context
        sheet.cell(row=row_num, column=6, value=rel['evidence_ref'])  # Evidence_Ref
        sheet.cell(row=row_num, column=7, value=rel['weight'])  # Weight
        
        row_num += 1
        seq += 1
    
    # 保存
    wb.save(excel_path)
    print(f"\n成功写入 {seq - 1} 条关系数据到 Sheet2")
    print(f"保存到: {excel_path}")


if __name__ == '__main__':
    main()
