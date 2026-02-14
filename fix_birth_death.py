import re
from docx import Document
from openpyxl import load_workbook

# Read Word document to get original Birth_Death data
doc = Document(r'c:\Users\33158\Desktop\大创\姓名.docx')
table = doc.tables[0]

# Build mapping: name -> birth_death
name_to_birth = {}
for i, row in enumerate(table.rows):
    if i == 0:
        continue
    cells = [cell.text.strip() for cell in row.cells]
    if len(cells) >= 3 and cells[0]:
        name = cells[0]
        birth_death = cells[2]
        # Extract years and format as xxxx-xxxx
        years = re.findall(r'\d{4}', birth_death)
        if len(years) >= 2:
            formatted = f'{years[0]}-{years[1]}'
        elif len(years) == 1:
            formatted = years[0]
        else:
            formatted = birth_death  # Keep original if no years found
        name_to_birth[name] = formatted

print(f"Extracted {len(name_to_birth)} Birth_Death records")
print("Sample:", list(name_to_birth.items())[:5])

# Update Excel
wb = load_workbook(r'c:\Users\33158\Desktop\大创\大创数据收集.xlsx')
sheet = wb['Sheet1']

updated = 0
for row in range(2, sheet.max_row + 1):
    name_cell = sheet.cell(row=row, column=3)  # True_Name
    birth_cell = sheet.cell(row=row, column=6)  # Birth_Death
    
    name = name_cell.value
    if name and name in name_to_birth:
        birth_cell.value = name_to_birth[name]
        updated += 1

wb.save(r'c:\Users\33158\Desktop\大创\大创数据收集.xlsx')
print(f"Updated {updated} Birth_Death entries in Excel")
